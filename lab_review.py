#!/usr/bin/env python3
"""
Lab Report Reviewer - Ansaldo Energia

Rule-based QA review of AEG materials-lab Excel reports. Two report families
are supported:

  * Metallurgical reports  (a "MET"-style sheet): header, sample/material,
    hardness, Nominal-vs-Actual chemical composition, comment, micrographs,
    sign-off.
  * Coating reports        (Cover + assessment sheet): coating-thickness
    measurements checked against the design MIN/MAX limits.

The reviewer is deterministic and runs fully offline. Every finding carries a
severity and a plain-English reason so an engineer can see the basis for it.

Public entry point:
    review_report(filename, data_bytes) -> (report_type, parsed, findings)
        report_type : 'metallurgical' | 'coating' | 'unknown'
        parsed      : dict of the extracted facts (for on-screen display)
        findings    : list of (severity, category, message)
                      severity in {'critical', 'warning', 'info', 'pass'}

Usage (CLI):  python3 lab_review.py report.xlsx
"""
import csv
import io
import os
import re
import shlex
import shutil
import subprocess
import sys
import zipfile
from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter

# Optional OCR / imaging stack. The reviewer works without it — legend, etch
# and thickness reading from micrographs are skipped gracefully when Pillow /
# pytesseract / the Tesseract binary are not present.
try:
    from PIL import Image, ImageFilter
    _PIL_AVAILABLE = True
except Exception:
    _PIL_AVAILABLE = False

_TESSERACT_BIN = shutil.which('tesseract')
pytesseract = None
try:
    import pytesseract
    pytesseract.get_tesseract_version()
    _OCR_AVAILABLE = _PIL_AVAILABLE
except Exception:
    # The wrapper is convenient but optional. A direct stdin/stdout invocation
    # keeps the installed Tesseract binary usable in lean deployments.
    pytesseract = None
    _OCR_AVAILABLE = _PIL_AVAILABLE and bool(_TESSERACT_BIN)

# Caption / etch / heat-treatment vocabulary lives in lab_vocab; import (and so
# re-export) it so external callers can keep doing `from lab_review import …`.
from lab_vocab import (  # noqa: E402,F401
    _PICNUM, _ETCH_PAT, _UNETCHED_PAT, _ALLOY_PAT, _norm_alloy,
    _ETCHANT_VOCAB, caption_etchant, report_etchants, image_etchant,
    _HT_VOCAB, HT_ORDER, caption_ht, report_ht, image_ht,
)

# ── Reference data ────────────────────────────────────────────────────────
# Chemical-element symbols expected in composition tables. Used to tell an
# element-header cell ("Cr", "Ni", ...) apart from an alloy name ("GTD-741").
ELEMENTS = {
    'Ni', 'Cr', 'Co', 'Mo', 'W', 'Al', 'Ti', 'Ta', 'C', 'B', 'Nb', 'V', 'Fe',
    'Zr', 'Cu', 'Mn', 'Si', 'Hf', 'Re', 'Y', 'Pt', 'Pd', 'S', 'P', 'N', 'O',
    'Mg', 'Ce', 'La', 'Sn', 'Ag',
}

# Composition tolerance bands (relative deviation of Actual vs Nominal). An
# absolute floor is applied as well so trace elements (e.g. C, B) don't trip
# the check on tiny absolute differences.
COMP_WARN_REL, COMP_WARN_ABS = 15.0, 0.50     # → verification warning
COMP_CRIT_REL, COMP_CRIT_ABS = 25.0, 0.50     # UI/reference threshold

# ── Reference hardness ────────────────────────────────────────────────────
# Typical hardness of common Ni- and Co-based gas-turbine superalloys in the
# fully-heat-treated / aged condition, in HRC. ADVISORY — representative of
# published/typical data; actual values depend on the exact heat-treat cycle
# and the controlling specification, so verify before relying on them.
# Anchors from datasheets / open literature: IN738LC aged ~40-45 HRC;
# Rene 80 aged ~35 HRC (as-cast ~38); GTD-111 ~440→320 HV ≈ 44→32 HRC across
# aging; IN718 aged 36-44 HRC; Nimonic/C263 ~28 HRC.
#
# CONDITION NOTE: AEG reports record *pre-* and *post-solution* hardness. The
# solution-treated state is intentionally SOFTER than these aged ranges (re-
# aging follows), so post-solution readings below the range are EXPECTED and
# are reported as informational, never as failures.
def _hrc(lo, hi, base, note=''):
    return {'hrc': (lo, hi), 'base': base, 'note': note}

# Keys are normalised (uppercase, alphanumerics only) — see _alloy_key().
HARDNESS_REF = {
    # ── Nickel-based, precipitation (γ′) hardened ──
    'IN738':      _hrc(32, 44, 'Ni'),
    'IN738LC':    _hrc(32, 44, 'Ni'),
    'INCONEL738': _hrc(32, 44, 'Ni'),
    'IN792':      _hrc(32, 44, 'Ni'),
    'GTD111':     _hrc(32, 44, 'Ni'),
    'GTD444':     _hrc(32, 44, 'Ni'),
    'GTD741':     _hrc(30, 42, 'Ni', 'GE proprietary — typical Ni bucket range; verify.'),
    'GTD222':     _hrc(25, 38, 'Ni', 'GE cast nozzle alloy (moderate γ′, weldable) — advisory; verify.'),
    'GTD241':     _hrc(25, 40, 'Ni', 'GE cast nozzle alloy — advisory; verify.'),
    'RENE80':     _hrc(30, 40, 'Ni'),
    'RENE108':    _hrc(35, 45, 'Ni'),
    'RENE142':    _hrc(35, 45, 'Ni'),
    'RENEN5':     _hrc(35, 45, 'Ni'),
    'MARM247':    _hrc(38, 46, 'Ni'),
    'CM247LC':    _hrc(38, 46, 'Ni'),
    'IN100':      _hrc(36, 44, 'Ni'),
    'IN713':      _hrc(30, 42, 'Ni'),
    'IN713C':     _hrc(30, 42, 'Ni'),
    'WASPALOY':   _hrc(32, 42, 'Ni'),
    'UDIMET500':  _hrc(30, 40, 'Ni'),
    'U500':       _hrc(30, 40, 'Ni'),     # Udimet 500 alias
    'UDIMET520':  _hrc(30, 40, 'Ni'),
    'UDIMET720':  _hrc(36, 46, 'Ni'),
    'IN718':      _hrc(36, 44, 'Ni'),
    'INCONEL718': _hrc(36, 44, 'Ni'),
    'NIMONIC263': _hrc(20, 32, 'Ni', 'Age-hardenable Ni-Co-Cr-Mo; aged ~28 HRC.'),
    'C263':       _hrc(20, 32, 'Ni'),
    'NI263':      _hrc(20, 32, 'Ni'),
    'HAYNES263':  _hrc(20, 32, 'Ni'),
    'NIMONIC90':  _hrc(30, 42, 'Ni'),
    'NIMONIC105': _hrc(32, 42, 'Ni'),
    'NIMONIC115': _hrc(32, 42, 'Ni'),
    'HAYNES282':  _hrc(28, 38, 'Ni'),
    # ── Nickel-based, solid-solution (not age-hardened; annealed, much softer) ──
    'IN625':      _hrc(8, 25, 'Ni', 'Solid-solution; annealed ~88-96 HRB.'),
    'INCONEL625': _hrc(8, 25, 'Ni', 'Solid-solution; annealed ~88-96 HRB.'),
    'HASTELLOYX': _hrc(8, 25, 'Ni', 'Solid-solution; annealed ~90 HRB.'),
    'HASTX':      _hrc(8, 25, 'Ni', 'Hastelloy X alias; solid-solution, ~90 HRB.'),
    # ── Cobalt-based (carbide / solid-solution strengthened) ──
    'FSX414':     _hrc(25, 38, 'Co', 'Cast Co nozzle/vane alloy.'),
    'X40':        _hrc(30, 42, 'Co'),
    'X45':        _hrc(30, 42, 'Co'),
    'STELLITE31': _hrc(30, 42, 'Co'),
    'MARM509':    _hrc(30, 42, 'Co'),
    'ECY768':     _hrc(30, 42, 'Co'),
    'STELLITE6':  _hrc(36, 45, 'Co'),
    'HAYNES188':  _hrc(8, 25, 'Co', 'Solid-solution; annealed ~95 HRB.'),
    'L605':       _hrc(8, 25, 'Co', 'Solid-solution; annealed ~95-100 HRB.'),
    'HAYNES25':   _hrc(8, 25, 'Co', 'Solid-solution; annealed ~95-100 HRB.'),
}


def _alloy_key(material):
    """Normalise an alloy name for HARDNESS_REF lookup (e.g. 'GTD-741'→'GTD741')."""
    return re.sub(r'[^A-Z0-9]', '', (material or '').upper())

# Placeholder strings that mean "field not actually filled in".
_PLACEHOLDERS = {'', 'n/a', 'na', 'not provided', 'to follow', 'tbd', '-', '/'}
_EXCEL_ERRORS = {
    '#NULL!', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#NUM!', '#N/A',
    '#GETTING_DATA',
}
_NOT_QUANTIFIED = re.compile(
    r'^(?:<\s*(?:lod|loq|mdl)|(?:below|under)\s+(?:detection|quantification)'
    r'|not\s+detected|n\.?d\.?|[<≤]\s*\d[\d.,]*)$',
    re.I,
)
# Words that mark a digit as part of a cross-reference (a note/revision
# number) rather than a measurement. `_num()` must not promote these to a
# fabricated numeric result (D5).
_NUM_REJECT = re.compile(r'\b(?:note|rev|revision|see|pending)\b', re.I)
# A left-censored / below-detection-limit numeric form, e.g. '<0.01' — this
# means "not quantified", not "equal to 0.01" (D5).
_BELOW_LOD_NUMERIC = re.compile(r'^[<≤]\s*-?\d')


# ── Low-level cell helpers ────────────────────────────────────────────────
def _txt(v):
    return '' if v is None else str(v).strip()


def _find(ws, pattern, col=None, max_row=None):
    """Return (row, col) of the first cell whose text matches `pattern`."""
    rx = re.compile(pattern, re.I)
    for row in ws.iter_rows(max_row=max_row):
        for cell in row:
            if col is not None and cell.column != col:
                continue
            t = _txt(cell.value)
            if t and rx.search(t):
                return cell.row, cell.column
    return None


def _value_right_loc(ws, row, col, max_scan=12):
    """(value, (row, col)) of the first non-empty cell to the right, else (None, None)."""
    for c in range(col + 1, col + 1 + max_scan):
        t = _txt(ws.cell(row=row, column=c).value)
        if t:
            return t, (row, c)
    return None, None


def _value_below_loc(ws, row, col, max_scan=6):
    """(value, (row, col)) of the first non-empty cell below, else (None, None)."""
    for r in range(row + 1, row + 1 + max_scan):
        t = _txt(ws.cell(row=r, column=col).value)
        if t:
            return t, (r, col)
    return None, None


def _value_right(ws, row, col, max_scan=12):
    """First non-empty cell value to the right of (row, col), same row."""
    return _value_right_loc(ws, row, col, max_scan)[0]


def _value_below(ws, row, col, max_scan=6):
    """First non-empty cell value below (row, col), same column."""
    return _value_below_loc(ws, row, col, max_scan)[0]


# ── Label→value resolution engine ─────────────────────────────────────────
# `_find()` above answers "where does this pattern appear on the sheet" and is
# still right for structural markers (report-type sniffing, table anchors,
# composition markers) where the pattern is meant to hit inside a longer
# string. Individual scalar fields need a stricter question: "is this cell
# ITSELF the label" — otherwise a sentence that merely mentions the label's
# words captures the field (D2), and the value scan needs to stop at the next
# label rather than walking onto it (D9) or a neighbouring field's own value.
#
# Two different strictness levels are used deliberately:
#   * `_find_label` / `_label_match` — strict, used to LOCATE a field's own
#     label. A match must consume (almost) the whole cell; prose that merely
#     contains the label text is rejected.
#   * the boundary index built by `_build_boundary_index` — permissive by
#     design, used only to STOP a value scan early. Stopping one cell too
#     soon just yields "not found", which is far safer than the scan
#     wandering onto the wrong text (the actual shape of D9), so it is fine
#     for this set to occasionally include a false positive.
_HEADER_LABELS = {
    'customer':     r'Customer',
    'customer_ref': r'Customer\s*Ref(?:erence)?\.?\s*(?:No\.?|Number)?',
    'aeg_ref':      r'AEG\s*Ref(?:erence)?\.?\s*(?:No\.?|Number)?',
    'job':          r'AEG\s*Job\s*(?:No\.?|Number)?',
    'machine':      r'Machine\s*Type',
    'qty':          r'Quantity(?:\s*per\s*(?:set|batch))?',
    'eoh':          r'EOH',
}
_HARDNESS_LABELS = {'pre': r'Pre-?\s*Solution', 'post': r'Post-?\s*Solution'}
_COATING_LABELS = {
    'present':  r'^Coating\s*$',
    'type':     r'^Type of Coating',
    'received': r'Received\s*Coating',
    'outgoing': r'Outgoing\s*Coating',
}
_SIGNOFF_LABELS = {
    'met_lab': r'Met\.?\s*Lab',
    'mat_eng': r'(?:Mat|Met)\.?\s*Eng',
    'date':    r'^Date\s*:',
}
_COMMENT_LABEL = r'Comments?'
_SAMPLE_HEADER_LABEL = r'Sample\s*nr'
_PICTURE_LABEL = r'Picture\s*\d+'
_COMPOSITION_LABELS = (r'\(\s*(?:Nominal|Minimal)\s*\)', r'\(\s*Actual\s*\)')


def _cell_label_text(v):
    """Cell text with one optional trailing ':' removed, or None if blank."""
    t = _txt(v)
    if not t:
        return None
    return t[:-1].rstrip() if t.endswith(':') else t


def _label_match(pattern, text):
    """Whether `text` (already colon-stripped) IS the label `pattern` names,
    not merely prose that mentions it — the pattern must consume the cell;
    whatever it leaves behind must carry no further letters or digits."""
    m = re.match(pattern, text, re.I)
    return bool(m) and not re.search(r'[A-Za-z0-9]', text[m.end():])


def _find_label(ws, pattern):
    """(row, col) of the first cell that IS the label `pattern` names."""
    for row in ws.iter_rows():
        for cell in row:
            text = _cell_label_text(cell.value)
            if text and _label_match(pattern, text):
                return cell.row, cell.column
    return None


def _build_boundary_index(ws):
    """(row, col) set of every cell that looks like SOME label — the stop-set
    for value scans. Built once per sheet and shared across every field."""
    strict = (list(_HEADER_LABELS.values())
              + [_COMMENT_LABEL, _SAMPLE_HEADER_LABEL, _PICTURE_LABEL])
    loose = (list(_HARDNESS_LABELS.values()) + list(_COATING_LABELS.values())
             + list(_SIGNOFF_LABELS.values()) + list(_COMPOSITION_LABELS))
    boundary = set()
    for row in ws.iter_rows():
        for cell in row:
            text = _cell_label_text(cell.value)
            if not text:
                continue
            if any(_label_match(pattern, text) for pattern in strict):
                boundary.add((cell.row, cell.column))
                continue
            full = _txt(cell.value)
            if any(re.search(pattern, full, re.I) for pattern in loose):
                boundary.add((cell.row, cell.column))
    return boundary


def _scan_right(ws, row, col, boundary, max_scan=12):
    for c in range(col + 1, col + 1 + max_scan):
        if (row, c) in boundary:
            break
        t = _txt(ws.cell(row=row, column=c).value)
        if t:
            return t, (row, c)
    return None, None


def _scan_below(ws, row, col, boundary, max_scan=6):
    for r in range(row + 1, row + 1 + max_scan):
        if (r, col) in boundary:
            break
        t = _txt(ws.cell(row=r, column=col).value)
        if t:
            return t, (r, col)
    return None, None


def _resolve_label_value(ws, row, col, boundary, primary='right', max_right=12, max_below=6):
    """Value for the label at (row, col): the primary direction first, the
    other as a fallback (D10 — some template variants stack the value below
    the label instead of beside it). Never crosses a boundary cell (D9)."""
    order = ((_scan_right, max_right), (_scan_below, max_below))
    if primary == 'below':
        order = order[::-1]
    for scan_fn, max_scan in order:
        val, loc = scan_fn(ws, row, col, boundary, max_scan)
        if val is not None:
            return val, loc
    return None, None


def _coord(loc):
    if not loc:
        return None
    row, col = loc
    return f'{get_column_letter(col)}{row}'


def _resolve_field(ws, pattern, boundary, primary='right'):
    """A single scalar field's full record: value/raw/cell/status/confidence
    (the D1 fix) via the strict label finder and boundary-aware value scan.

    status: 'found' (label + value both present), 'empty' (label located,
    value cell genuinely blank — a report defect) or 'not_located' (label
    itself never found — an extraction failure, never a finding about the
    report).
    """
    lbl = _find_label(ws, pattern)
    if not lbl:
        return ({'value': None, 'raw': None, 'cell': None,
                 'status': 'not_located', 'confidence': 0.0},
                {'label': None, 'value': None})
    raw, vloc = _resolve_label_value(ws, lbl[0], lbl[1], boundary, primary=primary)
    if raw is None:
        return ({'value': None, 'raw': None, 'cell': None,
                 'status': 'empty', 'confidence': 1.0},
                {'label': lbl, 'value': None})
    return ({'value': raw, 'raw': raw, 'cell': _coord(vloc),
             'status': 'found', 'confidence': 1.0},
            {'label': lbl, 'value': vloc})


def _is_placeholder(v):
    text = _txt(v)
    return text.lower() in _PLACEHOLDERS or text.upper() in _EXCEL_ERRORS


def _excel_errors(wb):
    """Visible Excel error values, including cached formula failures."""
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = _txt(cell.value).upper()
                if value in _EXCEL_ERRORS:
                    out.append({
                        'sheet': ws.title,
                        'cell': cell.coordinate,
                        'row': cell.row,
                        'col': cell.column,
                        'value': value,
                    })
    return out


def _formula_issues(wb):
    """Formula constructs that make a released workbook non-self-contained."""
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                formula = _txt(cell.value)
                if cell.data_type == 'f' and re.search(r'\[[^\]]+\][^!]*!', formula):
                    out.append({
                        'sheet': ws.title,
                        'cell': cell.coordinate,
                        'row': cell.row,
                        'col': cell.column,
                        'formula': formula,
                        'kind': 'external-reference',
                    })
    return out


def _num(v):
    """Parse a float from a cell that may carry units/symbols. Handles both the
    English '12.5' / '1,234.56' and the European '12,5' / '1.234,56' forms
    (common on Italian-lab reports); returns None when there's no number.

    Deliberately refuses two shapes that used to fabricate a measurement
    (D5): a digit that is really a cross-reference ('see note 3' -> 3.0), and
    a left-censored below-detection-limit value ('<0.01' -> 0.01, silently
    losing the below-LOD meaning). Both must come back None, not a number.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    text = str(v).strip()
    if _BELOW_LOD_NUMERIC.match(text) or _NUM_REJECT.search(text):
        return None
    m = re.search(r'-?\d[\d.,  ]*\d|-?\d', text)
    if not m:
        return None
    s = re.sub(r'[ \s]', '', m.group())     # drop spaces / non-breaking spaces
    # The decimal separator is whichever of '.' / ',' appears LAST; the other is
    # a thousands grouper. (No separator ⇒ plain integer.)
    if s.rfind(',') > s.rfind('.'):
        s = s.replace('.', '').replace(',', '.')  # European: comma is the decimal
    else:
        s = s.replace(',', '')                    # English: comma groups thousands
    try:
        return float(s)
    except ValueError:
        return None


def _workbook_text(wb):
    """Visible workbook text used by evidence/completeness controls.

    The filename is deliberately excluded: a word such as ``final`` or ``R``
    in a filename is not controlled document metadata inside the report.
    """
    values = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = _txt(cell.value)
                if value:
                    values.append(value)
    return "\n".join(values)


def _page_control(wb):
    """Printed footer/page-number metadata from every worksheet."""
    records = []
    for ws in wb.worksheets:
        for kind, footer in (
                ('odd', ws.oddFooter), ('even', ws.evenFooter),
                ('first', ws.firstFooter)):
            parts = [
                _txt(getattr(footer, side).text)
                for side in ('left', 'center', 'right')
            ]
            text = " | ".join(part for part in parts if part)
            if text:
                records.append({'sheet': ws.title, 'kind': kind, 'text': text})
    return records


# ── Report-type detection ─────────────────────────────────────────────────
def detect_type(wb):
    for ws in wb.worksheets:
        if _find(ws, r'Coating\s*Coverage\s*Assessment', max_row=10):
            return 'coating'
    for ws in wb.worksheets:
        if _find(ws, r'Design\s*limit') and _find(ws, r'Measurements'):
            return 'coating'
    for ws in wb.worksheets:
        if _find(ws, r'METALLURGICAL\s+EXAMINATION', max_row=6) or _find(ws, r'Sample\s*nr'):
            return 'metallurgical'
    return 'unknown'


# ════════════════════════════════════════════════════════════════════════
# METALLURGICAL REPORTS
# ════════════════════════════════════════════════════════════════════════
def _met_sheet(wb):
    for ws in wb.worksheets:
        if _find(ws, r'Sample\s*nr') or _find(ws, r'METALLURGICAL\s+EXAMINATION', max_row=6):
            return ws
    return wb.worksheets[0]


def _header(ws, boundary):
    """Header fields, both as the back-compat {key: value} view and as full
    {key: {value, raw, cell, status, confidence}} records (the D1 fix)."""
    fields, out, loc = {}, {}, {}
    for key, pat in _HEADER_LABELS.items():
        rec, lc = _resolve_field(ws, pat, boundary, primary='right')
        fields[key] = rec
        out[key] = rec['value']
        loc[key] = lc
    return fields, out, loc


def _samples(ws, boundary):
    """Every row of the sample table, not just the first (D3).

    A row belongs to the table as long as its `sample nr` cell (the table's
    primary column) is non-empty and is not itself another label — the first
    blank or label-boundary cell in that column ends the table, so a later
    section that happens to reuse the same column is never mistaken for more
    sample rows.
    """
    hdr = _find(ws, r'Sample\s*nr')
    if not hdr:
        return [], []
    hrow = hdr[0]
    headers = {}
    for cell in ws[hrow]:
        t = _txt(cell.value).lower()
        if t:
            headers[t] = cell.column

    def col_for(substr):
        for h, c in headers.items():
            if substr in h:
                return c
        return None

    field_cols = [(key, col_for(substr)) for key, substr in (
        ('sample_no', 'sample nr'), ('description', 'description'),
        ('serial', 's/n'), ('location', 'location'),
        ('material', 'material'), ('result', 'result'))]
    field_cols = [(key, col) for key, col in field_cols if col is not None]
    if not field_cols:
        return [], []
    primary_col = field_cols[0][1]

    samples, locs = [], []
    max_row = min(ws.max_row, hrow + 500)
    for r in range(hrow + 1, max_row + 1):
        if (r, primary_col) in boundary:
            break
        if not _txt(ws.cell(row=r, column=primary_col).value):
            break
        row_vals, row_loc = {}, {}
        for key, col in field_cols:
            v = _txt(ws.cell(row=r, column=col).value)
            if v:
                row_vals[key] = v
                row_loc[key] = {'label': (hrow, col), 'value': (r, col)}
        samples.append(row_vals)
        locs.append(row_loc)
    return samples, locs


def _hardness_unit(raw):
    """Hardness scale named in a cell (HRC / HV / HBW / HRB / …), or None."""
    m = re.search(r'\b(HRC|HRB|HRA|HV|HBW|HB|HK)\b', str(raw or ''), re.I)
    return m.group(1).upper() if m else None


def _hardness(ws, boundary):
    out, loc = {}, {}
    for key, pat in _HARDNESS_LABELS.items():
        lbl = _find(ws, pat)
        if lbl:
            raw, vloc = _resolve_label_value(ws, lbl[0], lbl[1], boundary, primary='right')
            out[key] = {'raw': raw, 'value': _num(raw), 'unit': _hardness_unit(raw)}
            loc[key] = {'label': lbl, 'value': vloc}
    return out, loc


def _coating(ws, boundary):
    """Coating presence / type as recorded in the structured cells."""
    out = {'present': None, 'type': None, 'received': None, 'outgoing': None}
    loc = {}
    for key, pat in _COATING_LABELS.items():
        lbl = _find(ws, pat)
        if lbl:
            out[key], vloc = _resolve_label_value(ws, lbl[0], lbl[1], boundary, primary='below')
            loc[key] = {'label': lbl, 'value': vloc}
    return out, loc


def _find_all(ws, pattern):
    """(row, col) of every cell whose text matches `pattern` — used where the
    report legitimately has more than one instance (composition tables, D4)."""
    rx = re.compile(pattern, re.I)
    hits = []
    for row in ws.iter_rows():
        for cell in row:
            t = _txt(cell.value)
            if t and rx.search(t):
                hits.append((cell.row, cell.column))
    return hits


def _composition_one_table(ws, lbl):
    """Parse the single composition table anchored at label cell `lbl`.

    Handles the two layouts seen in practice: element headers in the
    '(Nominal/Actual)' label row with values below, OR element headers in the
    row above the label with values in the label row. Returns None if the
    label isn't actually sitting next to an element-header row (a stray
    mention rather than a real table).
    """
    lrow, lcol = lbl
    label_text = _txt(ws.cell(row=lrow, column=lcol).value)
    label_cell = ws.cell(row=lrow, column=lcol).coordinate

    def elem_cells(r):
        return [c for c in ws[r] if r >= 1 and _txt(c.value).capitalize() in ELEMENTS]

    if len(elem_cells(lrow)) >= 2:           # elements in label row, values below
        ehdr, vrow = lrow, lrow + 1
    elif lrow > 1 and len(elem_cells(lrow - 1)) >= 2:   # elements above, values in label row
        ehdr, vrow = lrow - 1, lrow
    else:
        return None

    header_cells = elem_cells(ehdr)
    counts = Counter(_txt(cell.value).capitalize() for cell in header_cells)
    duplicate_headers = sorted(el for el, count in counts.items() if count > 1)
    values, loc, entries = {}, {}, []
    for cell in header_cells:
        raw = ws.cell(row=vrow, column=cell.column).value
        val = _num(raw)
        el = _txt(cell.value).capitalize()
        entry = {
            'element': el,
            'raw': _txt(raw),
            'value': val,
            'header_cell': cell.coordinate,
            'value_cell': ws.cell(row=vrow, column=cell.column).coordinate,
            'row': vrow,
            'col': cell.column,
        }
        entries.append(entry)
        if val is not None:
            values[el] = val
            loc[el] = (vrow, cell.column)
        elif el not in loc:
            loc[el] = (vrow, cell.column)
    return {
        'label_cell': label_cell,
        'label_text': label_text,
        'header_row': ehdr,
        'value_row': vrow,
        'values': values,
        'loc': loc,
        'entries': entries,
        'duplicate_headers': duplicate_headers,
    }


def _composition_tables(ws, which):
    """Every '(Nominal/Actual)' composition table on the sheet, not just the
    first (D4). Tolerates the 'Minimal' typo for 'Nominal'."""
    pat = r'\(\s*(?:Nominal|Minimal)\s*\)' if which == 'Nominal' else r'\(\s*Actual\s*\)'
    tables = []
    for lbl in _find_all(ws, pat):
        table = _composition_one_table(ws, lbl)
        if table is not None:
            tables.append(table)
    return tables


def _comment(ws, boundary=None):
    if boundary is None:
        boundary = _build_boundary_index(ws)
    lbl = _find_label(ws, _COMMENT_LABEL)
    if not lbl:
        return None, {}
    val, vloc = _resolve_label_value(ws, lbl[0], lbl[1], boundary, primary='below')
    return val, {'label': lbl, 'value': vloc}


def _pictures(ws):
    rx = re.compile(r'Picture\s*\d+\s*:', re.I)
    pics, loc = [], []
    for row in ws.iter_rows():
        for cell in row:
            if rx.search(_txt(cell.value)):
                cap, vloc = _value_right_loc(ws, cell.row, cell.column)
                pics.append((_txt(cell.value), cap))
                loc.append({'label': (cell.row, cell.column), 'value': vloc})
    return pics, loc


def _signoff(ws, boundary):
    out, loc = {}, {}
    for key, pat in _SIGNOFF_LABELS.items():
        lbl = _find(ws, pat)
        if lbl:
            raw, vloc = _resolve_label_value(ws, lbl[0], lbl[1], boundary, primary='right')
            out[key] = raw
            loc[key] = {'label': lbl, 'value': vloc}
    return out, loc


def parse_metallurgical(wb, media=0, micrograph_count=None):
    ws = _met_sheet(wb)
    boundary = _build_boundary_index(ws)

    fields, header, lh = _header(ws, boundary)
    samples, ls = _samples(ws, boundary)
    hardness, lhd = _hardness(ws, boundary)
    nominal_tables = _composition_tables(ws, 'Nominal')
    actual_tables = _composition_tables(ws, 'Actual')
    coating, lc = _coating(ws, boundary)
    comment, lcm = _comment(ws, boundary)
    pictures, lp = _pictures(ws)
    signoff, lso = _signoff(ws, boundary)

    sample = samples[0] if samples else {}
    empty_meta = {'entries': [], 'duplicate_headers': []}
    nominal_meta = nominal_tables[0] if nominal_tables else empty_meta
    actual_meta = actual_tables[0] if actual_tables else empty_meta

    return {
        'samples':   samples,
        'fields':    fields,
        'composition': {
            'nominal_tables': nominal_tables,
            'actual_tables':  actual_tables,
        },

        # ── back-compat views, derived from the canonical data above ──────
        'header':    header,
        'sample':    sample,
        'hardness':  hardness,
        'nominal':   nominal_tables[0]['values'] if nominal_tables else {},
        'actual':    actual_tables[0]['values'] if actual_tables else {},
        'composition_meta': {
            'nominal': nominal_meta,
            'actual': actual_meta,
        },
        'coating':   coating,
        'comment':   comment,
        'pictures':  pictures,
        'signoff':   signoff,
        'media':     media,
        'micrograph_count': micrograph_count,
        'loc': {
            'sheet':    ws.title,
            'header':   lh,
            'sample':   ls[0] if ls else {},
            'samples':  ls,
            'hardness': lhd,
            'nominal':  nominal_tables[0]['loc'] if nominal_tables else {},
            'actual':   actual_tables[0]['loc'] if actual_tables else {},
            'coating':  lc,
            'comment':  lcm,
            'pictures': lp,
            'signoff':  lso,
        },
    }


def _deviation_severity(rel, dev):
    """A nominal-vs-actual deviation is a verification warning by itself.

    These reports do not state controlled limits or method uncertainty, so an
    ideal/nominal single value cannot by itself prove alloy rejection. Separate
    checks still fail gross table corruption, impossible totals, duplicate
    headers, unquantified major elements and unexpected major chemistry.
    """
    return 'warning'


def _composition_deviations(nominal, actual):
    """Elements whose Actual is out of tolerance vs Nominal.

    Returns (deviations, systemic) where deviations is a list of
    (element, nominal, actual, rel%, severity) and `systemic` is True when so
    many elements are off that the actual material likely isn't the stated
    alloy (a separate, count-based signal from any single element's own
    severity — either can independently make the overall verdict critical).
    """
    deviations = []
    if not nominal or not actual:
        return deviations, False
    common = sorted(set(nominal) & set(actual))
    for el in common:
        nom, act = nominal[el], actual[el]
        if nom == 0:
            continue
        dev = act - nom
        rel = dev / abs(nom) * 100.0
        if abs(rel) >= COMP_WARN_REL and abs(dev) >= COMP_WARN_ABS:
            deviations.append((el, nom, act, rel, _deviation_severity(rel, dev)))
    n_dev, n_common = len(deviations), len(common)
    systemic = n_dev >= 4 or (n_dev >= 3 and n_common and n_dev / n_common >= 0.5)
    return deviations, systemic


def _gross_composition_conflict(deviations, n_common, unexpected_major=None):
    """High-confidence table corruption, not ordinary service/method scatter."""
    unexpected_major = unexpected_major or {}
    severe_count = sum(abs(relative) >= 25 for _el, _nom, _act, relative, _sev in deviations)
    all_columns_conflict = (
        n_common >= 8 and len(deviations) == n_common and severe_count >= 3
    )
    # In this report family, a multi-percent unexpected vanadium result combined
    # with several severe nominal conflicts is a high-confidence sign that the
    # actual row is shifted or copied. This deliberately requires both signals;
    # ordinary service scatter must remain a verification warning.
    high_vanadium_conflict = (
        abs(unexpected_major.get('V', 0)) >= 5
        and len(deviations) >= 3
        and severe_count >= 2
    )
    return all_columns_conflict or high_vanadium_conflict


def _gross_composition_note(deviations, n_common, unexpected_major=None):
    unexpected_major = unexpected_major or {}
    worst = sorted(deviations, key=lambda d: -abs(d[3]))[:4]
    detail = ", ".join(f"{el} {rel:+.0f}%" for el, _, _, rel, _ in worst)
    extras = ", ".join(
        f"unexpected {el} {value:g} wt%"
        for el, value in sorted(unexpected_major.items())
        if abs(value) >= 0.5
    )
    evidence = f'{detail}' + (f'; {extras}' if extras else '')
    return (
        f'{len(deviations)} of {n_common} comparable elements conflict materially '
        f'with the nominal table ({evidence}) — the chemistry table appears '
        f'corrupted, misaligned or copied from another analysis.'
    )


def _review_composition(nominal, actual, composition_meta=None):
    findings = []
    if not nominal or not actual:
        findings.append(('warning', 'Composition',
                         'Could not read both Nominal and Actual composition tables.'))
        # Keep reviewing the table structure: a partially parsed table can still
        # reveal a duplicate header or an explicit <LOD result.

    common = sorted(set(nominal) & set(actual))
    deviations, systemic = _composition_deviations(nominal, actual)
    n_dev, n_common = len(deviations), len(common)
    unexpected_major = {
        el: value for el, value in actual.items()
        if el not in nominal and abs(value) >= 0.5
    }
    composition_meta = composition_meta or {}
    actual_meta = composition_meta.get('actual') or {}
    entries = actual_meta.get('entries') or []
    actual_headers = {e.get('element') for e in entries if e.get('element')}

    duplicates = actual_meta.get('duplicate_headers') or []
    if duplicates:
        findings.append(('critical', 'Composition',
                         f'Actual composition table repeats element header(s): '
                         f'{", ".join(duplicates)} — at least one column is mislabeled, so the '
                         f'chemistry cannot be accepted as reported.'))

    for entry in entries:
        raw = entry.get('raw') or ''
        el = entry.get('element')
        if not el or entry.get('value') is not None or not _NOT_QUANTIFIED.match(raw):
            continue
        nom = nominal.get(el)
        if nom is not None and nom >= 0.5:
            findings.append(('critical', 'Composition',
                             f'{el}: actual result is "{raw}" while nominal is {nom:g} wt% — '
                             f'a major alloying element was not quantified.'))
        else:
            findings.append(('warning', 'Composition',
                             f'{el}: actual result is "{raw}" — verify the analytical method '
                             f'and detection limit.'))

    nominal_meta = composition_meta.get('nominal') or {}
    nominal_entries = [e['value'] for e in nominal_meta.get('entries') or []
                       if e.get('value') is not None]
    nominal_has_balance = any(re.search(r'\b(?:bal(?:ance)?|remainder)\b',
                                        entry.get('raw') or '', re.I)
                              for entry in nominal_meta.get('entries') or [])
    if len(nominal_entries) >= 5 and not nominal_has_balance:
        nominal_total = sum(nominal_entries)
        if nominal_total < 99.0 or nominal_total > 101.0:
            severity = 'critical' if nominal_total > 101.0 else 'warning'
            findings.append((severity, 'Composition',
                             f'Nominal composition totals {nominal_total:.2f} wt%, '
                             f'outside the 99–101 wt% reconciliation band — no balance '
                             f'or range basis explains the table as written.'))

    numeric_entries = [e['value'] for e in entries if e.get('value') is not None]
    if len(numeric_entries) >= 5:
        total = sum(numeric_entries)
        if total < 98.0 or total > 102.0:
            severity = 'critical' if total < 95.0 or total > 102.0 else 'warning'
            findings.append((severity, 'Composition',
                             f'Actual composition totals {total:.2f} wt%, outside the '
                             f'98–102 wt% reconciliation band — check missing/misaligned '
                             f'columns and analytical completeness.'))

    if re.search(r'\bMinimal\b', nominal_meta.get('label_text') or '', re.I):
        findings.append(('warning', 'Composition',
                         'The composition heading says "Minimal" instead of "Nominal"; '
                         'the declared chemistry basis is ambiguous.'))

    # A few/multiple elements off requires verification against a real
    # specification. Only an all-column, multi-severe conflict is treated as a
    # corrupted/misaligned table.
    gross_conflict = _gross_composition_conflict(
        deviations, n_common, unexpected_major)
    if gross_conflict:
        findings.append(('critical', 'Composition',
                         _gross_composition_note(
                             deviations, n_common, unexpected_major)))
    else:
        if systemic:
            worst = sorted(deviations, key=lambda d: -abs(d[3]))[:4]
            detail = ", ".join(f"{el} {rel:+.0f}%" for el, _, _, rel, _ in worst)
            findings.append(('warning', 'Composition',
                             f'{n_dev} of {n_common} comparable elements differ materially '
                             f'from the nominal values ({detail}); verify against the governing '
                             f'alloy specification and analytical method.'))
        else:
            for el, nom, act, rel, sev in deviations:
                findings.append((sev, 'Composition',
                                 f'{el}: actual {act:g} vs nominal {nom:g} wt% '
                                 f'({rel:+.0f}%) — verify.'))

    only_nom = sorted(set(nominal) - (actual_headers or set(actual)))
    only_act = sorted((actual_headers or set(actual)) - set(nominal))
    if only_nom:
        missing_major = sorted(el for el in only_nom if nominal.get(el, 0) >= 1.0)
        missing_trace = sorted(set(only_nom) - set(missing_major))
        if missing_major:
            findings.append(('warning', 'Composition',
                             f'Nominal major element(s) not reported in Actual: '
                             f'{", ".join(missing_major)} — verify analytical completeness.'))
        if missing_trace:
            findings.append(('info', 'Composition',
                             f'Nominal trace element(s) not reported in Actual: '
                             f'{", ".join(missing_trace)}.'))
    if only_act:
        major = []
        minor = []
        for el in only_act:
            values = [e.get('value') for e in entries
                      if e.get('element') == el and e.get('value') is not None]
            if values and max(abs(v) for v in values) >= 0.5:
                major.append(f'{el} {max(values, key=abs):g} wt%')
            else:
                minor.append(el)
        if major:
            findings.append(('warning', 'Composition',
                             f'Materially reported element(s) absent from the nominal table: '
                             f'{", ".join(major)} — verify the headers, method and alloy basis.'))
        if minor:
            findings.append(('info', 'Composition',
                             f'Reported but not in nominal spec: {", ".join(minor)}.'))
    if nominal and actual and not deviations:
        findings.append(('pass', 'Composition',
                         f'All {len(common)} matched elements within ±{COMP_WARN_REL:g}% tolerance.'))
    return findings


_CONTROL_REFERENCE = re.compile(
    r'\bacceptance\s+(?:criteria|criterion|limits?|requirements?)\b|'
    r'\b(?:governing|applicable)\s+(?:specification|standard|procedure|drawing)\b|'
    r'\b(?:specification|standard|procedure|drawing)\s*(?:no\.?|number|:)\s*'
    r'[A-Z0-9][A-Z0-9./_-]{2,}',
    re.I,
)


def _review_acceptance_and_methods(parsed):
    """Release controls that the old template silently omitted.

    The checks are evidence-based: a future report that actually states a
    governing criterion/method is not flagged merely because it uses this
    legacy layout.
    """
    findings = []
    text = parsed.get('document_text') or ''
    if not _CONTROL_REFERENCE.search(text):
        findings.append(('critical', 'Acceptance criteria',
                         'No governing acceptance specification, limits or decision '
                         'rule is stated, so the report cannot support acceptance as issued.'))

    if parsed.get('nominal') or parsed.get('actual'):
        method_markers = {
            'test method / instrument': re.compile(
                r'\b(?:OES|ICP(?:-OES|-MS)?|XRF|EDS|EDX|spectromet\w*|'
                r'chemical\s+(?:analysis|test)\s+method|ASTM\s+E\d+)\b', re.I),
            'calibration / traceability': re.compile(r'\bcalibrat\w*|traceab\w*\b', re.I),
            'measurement uncertainty': re.compile(r'\buncertaint\w*\b', re.I),
            'governing material specification': _CONTROL_REFERENCE,
        }
        missing = [label for label, pattern in method_markers.items()
                   if not pattern.search(text)]
        if missing:
            findings.append(('warning', 'Chemistry method',
                             'Chemical results lack ' + ', '.join(missing) + '.'))

    hardness = parsed.get('hardness') or {}
    if any((hardness.get(key) or {}).get('value') is not None
           for key in ('pre', 'post')):
        missing = []
        if not re.search(r'\b(?:ASTM\s+E18|ISO\s+6508|Rockwell\s+(?:tester|method)|'
                         r'hardness\s+(?:test\s+)?method)\b', text, re.I):
            missing.append('method/standard and equipment')
        if not re.search(r'\b(?:number\s+of\s+readings|\d+\s+readings?|n\s*=|'
                         r'hardness\s+locations?|range|standard\s+deviation)\b', text, re.I):
            missing.append('reading count, locations and range/statistics')
        if not _CONTROL_REFERENCE.search(text):
            missing.append('acceptance limits')
        if missing:
            findings.append(('warning', 'Hardness evidence',
                             'Hardness values lack ' + ', '.join(missing) + '.'))
    return findings


def _review_sampling_basis(parsed):
    text = parsed.get('document_text') or ''
    # "Sampling location" identifies where a cut was taken; it is not a plan
    # that justifies how a few specimens represent the whole set.
    has_plan = re.search(
        r'\b(?:sampling\s+(?:plan|basis|procedure)|sample\s+size\s+justification|'
        r'representative\s+sample|random\s+sampling|inspection\s+level|AQL)\b',
        text, re.I)
    if has_plan:
        return []
    sample_ids = _identifier_tokens((parsed.get('sample') or {}).get('sample_no'), 'sample')
    quantity = (parsed.get('header') or {}).get('qty')
    detail = (f'Set quantity {quantity}; {len(sample_ids)} specimen(s) listed.'
              if quantity else f'{len(sample_ids)} specimen(s) listed; set quantity is blank.')
    return [('warning', 'Sampling',
             'No sampling plan or justification shows how the listed specimen(s) '
             f'represent the full set. {detail}')]


def _review_phase_evidence(parsed):
    comment = parsed.get('comment') or ''
    phase_claim = re.search(
        r'\b(?:MC|M23C6|TCP|sigma|carbides?|brittle\s+phases?)\b|'
        r'gamma\s*[- ]?prime|γ\s*[\'′-]?', comment, re.I)
    if not phase_claim:
        return []
    support = re.search(
        r'\b(?:SEM|EDS|EDX|EBSD|quantif\w*|grading\s+method|reference\s+micrograph|'
        r'controlled\s+reference|acceptance\s+(?:criteria|limit))\b',
        parsed.get('document_text') or '', re.I)
    findings = []
    if not support:
        findings.append(('warning', 'Metallography evidence',
                         'Named phase/degradation claims are qualitative and are not '
                         'supported by a grading method, controlled references, SEM/EDS '
                         'confirmation or an acceptance criterion.'))
    gamma = re.search(r'gamma\s*[- ]?prime|γ\s*[\'′-]?', comment, re.I)
    if gamma and not re.search(r'\bSEM\b|gamma\s*[- ]?prime\s+(?:size|spacing|fraction)|'
                               r'γ\s*[\'′-]?\s*(?:size|spacing|fraction)',
                               parsed.get('document_text') or '', re.I):
        findings.append(('warning', 'Metallography evidence',
                         'Optical images at the stated magnifications do not substantiate '
                         'a definitive γ′ morphology, dissolution or re-precipitation conclusion.'))
    return findings


def _review_document_control(parsed):
    text = parsed.get('document_text') or ''
    findings = []
    has_report_no = re.search(r'\b(?:report|document)\s*(?:no\.?|number|id)\s*[:#-]?\s*'
                              r'[A-Z0-9][A-Z0-9./_-]{2,}', text, re.I)
    has_revision = re.search(r'\b(?:revision|rev\.?|version|status)\s*[:#-]?\s*'
                             r'[A-Z0-9][A-Z0-9._/-]*', text, re.I)
    if not (has_report_no and has_revision):
        findings.append(('warning', 'Document control',
                         'The report body has no controlled report number and revision/status; '
                         'filename labels such as R, final, done or TBU are not approval evidence.'))

    footers = parsed.get('page_control') or []
    footer_text = [record.get('text') or '' for record in footers]
    has_dynamic = any('&P' in value or '&[Page]' in value for value in footer_text)
    static = [value for value in footer_text
              if re.search(r'\bPage\s*[-:#]?\s*\d+\b', value, re.I)]
    if static and not has_dynamic:
        findings.append(('warning', 'Pagination',
                         'The printed footer uses a hard-coded page label instead of '
                         f'sequential page numbering: {"; ".join(static)}.'))
    elif not has_dynamic:
        findings.append(('warning', 'Pagination',
                         'No sequential printed page numbering was found in the report footer.'))
    return findings


def _review_alloy_identity(parsed):
    """High-confidence internal alloy contradictions found in this report family."""
    material = (parsed.get('sample') or {}).get('material') or ''
    key = _alloy_key(material)
    nominal = parsed.get('nominal') or {}
    actual = parsed.get('actual') or {}
    text = parsed.get('document_text') or ''
    findings = []

    nominal_entries = (((parsed.get('composition_meta') or {}).get('nominal') or {})
                       .get('entries') or [])
    nominal_values = [entry.get('value') for entry in nominal_entries
                      if entry.get('value') is not None]
    nominal_total = sum(nominal_values) if len(nominal_values) >= 5 else None

    if key in {'IN738', 'INCONEL738'} and nominal_total is not None and nominal_total > 101:
        findings.append(('critical', 'Material identity',
                         f'The declared IN-738 nominal table totals {nominal_total:.2f} wt%. '
                         'IN-738 uses Ni as balance; the fixed values in this table are not '
                         'a valid controlled composition.'))

    if key in {'IN738', 'INCONEL738', 'GTD741'} and actual.get('Re', 0) >= 0.5:
        findings.append(('critical', 'Material identity',
                         f'The declared {material} chemistry contains Re {actual["Re"]:g} wt% '
                         'without a matching nominal requirement; verify the alloy identity '
                         'and analytical headers.'))

    if key in {'IN738', 'INCONEL738'} and re.search(r'\bAISI\s*[- ]?310\b', text, re.I):
        findings.append(('critical', 'Material identity',
                         'The report conflicts internally: the sample is identified as IN-738 '
                         'while the composition/narrative identifies AISI 310.'))

    if re.search(r'\b(?:update|change|replace|revise)\s+(?:this\s+)?(?:as|to|with|per)\b',
                 text, re.I):
        findings.append(('critical', 'Document control',
                         'An internal editing instruction remains visible in the issued report.'))

    # Udimet 500 has substantially more Co/Mo than the IN-738-like table used
    # by one legacy workbook. Detect the internal contradiction, not a broad
    # alloy-library comparison.
    if key in {'U500', 'UDIMET500'}:
        looks_in738 = (
            7 <= nominal.get('Co', -1) <= 10
            and 14 <= nominal.get('Cr', -1) <= 18
            and nominal.get('Mo', 99) < 2.5
            and nominal.get('Ta', 0) >= 1
            and nominal.get('Ni', 0) >= 60
        )
        if looks_in738 and actual.get('Co', 0) >= 14 and actual.get('Mo', 0) >= 3:
            findings.append(('critical', 'Material identity',
                             'The U-500 report uses an IN-738-like nominal chemistry table; '
                             'the nominal alloy basis conflicts with the actual U-500-type chemistry.'))
    return findings


# Coating-type vocabulary (tolerant of the "MCrAlY"/"MCrAIY" spelling seen in
# the sheets). Each entry maps a canonical name to a detection pattern.
_COATING_TYPE_PATS = (
    ('TBC',       r'\bTBC\b|thermal\s*barrier'),
    ('MCrAlY',    r'MCR\w*Y'),
    ('aluminide', r'alumini[sz]|\baluminide\b|\bPt[-\s]?Al\b|platinum\s*alumin'),
    ('diffusion', r'diffusion\s*coat'),
    ('chromide',  r'chromi[sz]|\bchromide\b'),
)


def _coating_types_in(text):
    """Set of canonical coating types mentioned in a piece of text."""
    t = text or ''
    return {name for name, pat in _COATING_TYPE_PATS if re.search(pat, t, re.I)}


def _review_comment(parsed):
    """Flag where the free-text comment contradicts the coating cells."""
    findings = []
    comment = parsed.get('comment') or ''
    coat = parsed.get('coating') or {}
    if not comment:
        return findings
    cl = comment.lower()

    cell_types = set()
    for key in ('type', 'received', 'outgoing'):
        cell_types |= _coating_types_in(coat.get(key))
    comment_types = _coating_types_in(comment)

    present = (coat.get('present') or '').strip().lower()
    cell_has  = present == 'yes' or bool(cell_types)
    cell_none = present == 'no' or (not cell_types and _is_placeholder(coat.get('type')))

    comment_has = bool(comment_types) or bool(re.search(
        r'received with[^.]{0,30}coating|coated with|coating (?:was |is )?(?:applied|present|intact)', cl))
    comment_none = bool(re.search(
        r'\buncoated\b|no coating|without (?:any )?coating|not coated|'
        r'coating (?:is |was )?(?:fully )?removed', cl))

    # Coating type: comment names a type the cell disagrees with.
    if cell_types and comment_types and cell_types.isdisjoint(comment_types):
        findings.append(('warning', 'Comment',
                         f'Comment mentions {"/".join(sorted(comment_types))} coating but the '
                         f'coating cell says {"/".join(sorted(cell_types))}.'))
    elif cell_types and (cell_types & comment_types):
        findings.append(('pass', 'Comment',
                         f'Comment coating type matches the coating cell '
                         f'({"/".join(sorted(cell_types & comment_types))}).'))

    # Coating presence: cell vs comment.
    if cell_none and comment_has and not comment_none:
        what = "/".join(sorted(comment_types)) if comment_types else 'a coating'
        findings.append(('warning', 'Comment',
                         f'Coating cell indicates no coating, but the comment refers to {what}.'))
    elif cell_has and comment_none and not comment_has:
        label = "/".join(sorted(cell_types)) or present
        findings.append(('warning', 'Comment',
                         f'Coating cell indicates a coating ({label}), but the comment says '
                         f'it is uncoated.'))

    # Alloy named in the comment vs the material cell.
    material = (parsed.get('sample') or {}).get('material')
    if material:
        mkey = _norm_alloy(material)
        others = sorted({m.group(0) for m in _ALLOY_PAT.finditer(comment)
                         if _norm_alloy(m.group(0)) != mkey
                         and _norm_alloy(m.group(0)) not in mkey
                         and mkey not in _norm_alloy(m.group(0))})
        if others:
            findings.append(('warning', 'Comment',
                             f'Comment mentions alloy {", ".join(others)} but the material cell '
                             f'says "{material}".'))

    # Service verdict in the comment vs the Result cell.
    result = (parsed.get('sample') or {}).get('result') or ''
    rlow = result.lower()
    neg = re.search(r'not\s+suitable|unsuitable|not\s+recommend|\breject|\bscrap|'
                    r'beyond\s+repair|non[-\s]?conform|unacceptable', cl)
    pos = re.search(r'(?<!not )(?:\bsuitable for|\bacceptable|recommended for|'
                    r'reconditi|fit for service|return to service)', cl)
    result_pos = bool(re.search(r'accept|suitable|conform|\bpass\b', rlow)) and 'see comment' not in rlow
    result_neg = bool(re.search(r'reject|not\s+suitable|scrap|unacceptable', rlow))
    if result_pos and neg and not pos:
        findings.append(('warning', 'Comment',
                         f'Result cell says "{result}" but the comment indicates the part is NOT suitable.'))
    elif result_neg and pos and not neg:
        findings.append(('warning', 'Comment',
                         f'Result cell says "{result}" but the comment indicates the part IS suitable.'))
    elif 'see comment' in rlow and bool(neg) != bool(pos):
        findings.append(('info', 'Comment',
                         f'Result defers to the comment; the comment verdict reads '
                         f'{"not suitable / negative" if neg else "suitable / positive"}.'))
    elif 'see comment' in rlow and not neg and not pos:
        findings.append(('critical', 'Disposition',
                         'Result says "See comment", but the comment gives no clear '
                         'accept / reject / repair disposition.'))
    return findings


def _review_hardness(hardness, material):
    findings = []
    if not hardness:
        findings.append(('info', 'Hardness', 'No hardness-results section found.'))
        return findings

    pre = hardness.get('pre', {}).get('value')
    post = hardness.get('post', {}).get('value')
    if pre is None and post is None:
        findings.append(('warning', 'Hardness', 'Hardness section present but no values parsed.'))
        return findings

    # Determine the scale: explicit unit from a cell, else inferred (HRC tops out
    # near 70; a bigger number is HV/HBW). This stops HV microhardness (e.g. 420)
    # being compared to — and flagged against — the HRC reference.
    vals = [v for v in (pre, post) if v is not None]
    units = {hardness.get(k, {}).get('unit') for k in ('pre', 'post')} - {None}
    unit = next(iter(units)) if len(units) == 1 else None
    is_hrc = (unit == 'HRC') or (unit is None and all(v <= 72 for v in vals))
    ustr = unit or ('HRC' if is_hrc else 'HV')

    # Solution treatment should soften the material (post ≤ pre) — scale-agnostic.
    # ±2 guard band so measurement scatter doesn't trip a false warning.
    if pre is not None and post is not None and post > pre + 2:
        findings.append(('warning', 'Hardness',
                         f'Post-solution hardness ({post:g} {ustr}) exceeds pre-solution '
                         f'({pre:g} {ustr}) — solution treatment normally softens the material.'))

    ref = HARDNESS_REF.get(_alloy_key(material))
    if not is_hrc:
        parts = [f'{k}={v["value"]:g}' for k, v in hardness.items() if v.get('value') is not None]
        findings.append(('info', 'Hardness',
                         f'Hardness recorded in {ustr}: {", ".join(parts)} — not compared '
                         f'to the HRC reference.'))
    elif ref:
        lo, hi = ref['hrc']
        note = (' ' + ref['note']) if ref['note'] else ''
        findings.append(('info', 'Hardness',
                         f'{material}: reference aged hardness {lo}–{hi} HRC '
                         f'({ref["base"]}-based, advisory).{note}'))
        for label, val in (('Pre-solution', pre), ('Post-solution', post)):
            if val is None:
                continue
            if val > hi + 2:
                findings.append(('info', 'Hardness',
                                 f'{label} {val:g} HRC is above the aged reference '
                                 f'{lo}–{hi} HRC — verify.'))
            elif val < lo and label == 'Post-solution':
                findings.append(('info', 'Hardness',
                                 f'{label} {val:g} HRC is below the aged reference '
                                 f'{lo}–{hi} HRC — expected for the solution-treated '
                                 f'(pre-aging) condition.'))
    else:
        findings.append(('info', 'Hardness',
                         f'No reference hardness on file for "{material}".'))

    if not any(s == 'warning' for s, _, _ in findings):
        parts = [f'{k}={v["value"]:g}' for k, v in hardness.items() if v.get('value') is not None]
        findings.append(('pass', 'Hardness', f'Hardness values recorded: {", ".join(parts)} {ustr}.'))
    return findings


def _review_evidence_claims(parsed):
    """Flag conclusions that claim more than the recorded evidence establishes."""
    comment = parsed.get('comment') or ''
    claims_mechanical_restoration = re.search(
        r'(?:restor\w*.{0,80}mechanical\s+propert|'
        r'mechanical\s+propert.{0,80}restor\w*)',
        comment,
        re.I | re.DOTALL,
    )
    if not claims_mechanical_restoration:
        return []

    # These templates only expose pre- and post-solution hardness. Neither is a
    # final aged-condition mechanical verification, and micrographs alone cannot
    # demonstrate restored mechanical properties.
    hardness = parsed.get('hardness') or {}
    if not any(key in hardness for key in ('post_age', 'final', 'aged')):
        return [('warning', 'Evidence',
                 'Comment claims that mechanical properties were restored, but the '
                 'report contains no final aged-condition mechanical result; '
                 'microstructure alone does not verify mechanical properties.')]
    return []


def _review_completeness(parsed):
    findings = []
    hdr = parsed['header']
    for key, label in (('customer', 'Customer'), ('job', 'AEG Job No'),
                       ('machine', 'Machine type')):
        if _is_placeholder(hdr.get(key)):
            findings.append(('warning', 'Completeness', f'{label} is blank or a placeholder.'))
    for key, label in (('customer_ref', 'Customer Ref No'), ('qty', 'Quantity per set')):
        if _is_placeholder(hdr.get(key)):
            findings.append(('warning', 'Completeness', f'{label} not provided.'))
    if _is_placeholder(hdr.get('eoh')):
        findings.append(('info', 'Completeness', 'EOH not provided.'))

    if _is_placeholder(parsed['sample'].get('material')):
        findings.append(('warning', 'Completeness', 'Sample material/alloy not stated.'))

    comment = parsed.get('comment') or ''
    if len(comment.strip()) < 40:
        findings.append(('warning', 'Completeness',
                         'Comment / discussion is missing or very short.'))

    pics = parsed.get('pictures', [])
    uncaptioned = [p for p, cap in pics if not cap]
    if not pics:
        findings.append(('warning', 'Micrographs', 'No micrograph captions found.'))
    elif uncaptioned:
        findings.append(('info', 'Micrographs',
                         f'{len(uncaptioned)} of {len(pics)} pictures have no caption.'))
    micrographs = parsed.get('micrograph_count')
    if micrographs == 0:
        findings.append(('warning', 'Micrographs', 'No embedded images found in the workbook.'))
    elif micrographs is not None and pics:
        if micrographs < len(pics):
            findings.append(('critical', 'Micrographs',
                             f'{len(pics)} picture caption(s) but only {micrographs} embedded '
                             f'micrograph(s) — {len(pics) - micrographs} evidence image(s) '
                             f'are missing.'))
        elif micrographs > len(pics):
            findings.append(('warning', 'Micrographs',
                             f'{micrographs} embedded micrograph(s) but only {len(pics)} '
                             f'picture caption(s) — check for uncatalogued or stray images.'))

    so = parsed['signoff']
    missing = [lbl for key, lbl in (('met_lab', 'Met. Lab'), ('mat_eng', 'Mat. Eng'),
                                    ('date', 'Date')) if _is_placeholder(so.get(key))]
    if missing:
        findings.append(('critical', 'Sign-off', f'Missing sign-off field(s): {", ".join(missing)}.'))
    else:
        findings.append(('pass', 'Sign-off', 'Lab, engineer and date all present.'))
    return findings


def _identifier_tokens(text, kind):
    """Normalised identifiers from stacked/space-separated sample or S/N cells."""
    raw = _txt(text).upper()
    if not raw:
        return []
    if kind == 'sample':
        hits = re.findall(r'\bMS\s*[-_]?\s*[A-Z0-9]+\b', raw)
        return [re.sub(r'[\s_-]+', '', hit) for hit in hits]
    tokens = []
    for block in re.split(r'[\n;,]+', raw):
        for part in re.split(r'\s{2,}', block.strip()):
            token = re.sub(r'\s+', '', part)
            if (len(token) >= 4 and re.fullmatch(r'[A-Z0-9/-]+', token)
                    and re.search(r'\d', token)):
                tokens.append(token)
    return tokens


def _review_traceability(parsed):
    findings = []
    sample = parsed.get('sample') or {}
    sample_ids = _identifier_tokens(sample.get('sample_no'), 'sample')
    serial_ids = _identifier_tokens(sample.get('serial'), 'serial')

    if sample_ids and serial_ids and len(sample_ids) != len(serial_ids):
        findings.append(('warning', 'Traceability',
                         f'Sample and serial/part identifiers do not map one-to-one: '
                         f'{len(sample_ids)} sample identifier(s) versus '
                         f'{len(serial_ids)} serial/part identifier(s).'))

    for label, values in (('sample number', sample_ids), ('serial/part number', serial_ids)):
        duplicates = sorted(value for value, count in Counter(values).items() if count > 1)
        if duplicates:
            findings.append(('critical', 'Traceability',
                             f'Duplicate {label}(s): {", ".join(duplicates)}.'))

    result = sample.get('result')
    if _is_placeholder(result):
        findings.append(('warning', 'Disposition', 'Result field is blank or invalid.'))
    return findings


def _review_samples(parsed):
    """A rejection (or other disqualifying result) on any sample past the
    first must not review clean (D3 — release-critical). Every other rule in
    this module still works off the back-compat `sample` (== samples[0]);
    this is the one check that looks at the rest of the table, closing the
    specific gap the defect register calls out: samples 2..N previously never
    reached any of the 20 checks."""
    findings = []
    for sample in (parsed.get('samples') or [])[1:]:
        result = (sample.get('result') or '').strip()
        if re.search(r'reject|scrap|non[-\s]?conform|unacceptable', result, re.I):
            ident = sample.get('serial') or sample.get('sample_no') or '?'
            findings.append((
                'critical', 'Disposition',
                f'Sample {ident}: result is "{result}" and must not be treated '
                f'as passing because it is not the first row of the table.'))
    return findings


def _review_workbook_integrity(parsed):
    errors = parsed.get('excel_errors') or []
    formulas = parsed.get('formula_issues') or []
    findings = []
    if errors:
        shown = ', '.join(f'{e["sheet"]}!{e["cell"]}={e["value"]}' for e in errors[:6])
        extra = f' (+{len(errors) - 6} more)' if len(errors) > 6 else ''
        findings.append(('critical', 'Workbook integrity',
                         f'Excel error value(s) present: {shown}{extra}.'))
    if formulas:
        shown = ', '.join(f'{e["sheet"]}!{e["cell"]} {e["formula"]}' for e in formulas[:4])
        extra = f' (+{len(formulas) - 4} more)' if len(formulas) > 4 else ''
        findings.append(('critical', 'Workbook integrity',
                         f'External-workbook formula reference(s) make the report '
                         f'non-self-contained: {shown}{extra}.'))
    return findings


# Caption / etch / heat-treatment vocabulary now lives in lab_vocab.py and is
# imported at the top of this module (and re-exported for external callers).


def _anchor_order(data):
    """Embedded micrographs in visual (drawing-anchor) order, top-to-bottom /
    left-to-right, excluding logos/thumbnails. Returns a list of image names."""
    if not _PIL_AVAILABLE:
        return []
    try:
        z = zipfile.ZipFile(io.BytesIO(data))
    except Exception:
        return []
    real = set()
    for n in z.namelist():
        if n.startswith('xl/media'):
            try:
                im = Image.open(io.BytesIO(z.read(n)))
                if im.size[0] >= 200 and im.size[1] >= 150:
                    real.add(n.split('/')[-1])
            except Exception:
                pass
    placed = []
    for d in [n for n in z.namelist() if re.match(r'xl/drawings/drawing\d+\.xml$', n)]:
        try:
            rels = z.read(d.replace('drawings/', 'drawings/_rels/') + '.rels').decode('utf-8', 'ignore')
        except Exception:
            continue
        rid2media = {}                       # parse Id/Target independently of order
        for rel in re.findall(r'<Relationship\b[^>]*>', rels):
            rid = re.search(r'Id="([^"]+)"', rel)
            tgt = re.search(r'Target="([^"]+)"', rel)
            if rid and tgt:
                rid2media[rid.group(1)] = tgt.group(1).split('/')[-1]
        xml = z.read(d).decode('utf-8', 'ignore')
        # Tolerate both the xdr:-prefixed (Excel) and bare (openpyxl) namespaces.
        for anc in re.findall(r'<(?:xdr:)?(?:two|one)CellAnchor\b.*?'
                              r'</(?:xdr:)?(?:two|one)CellAnchor>', xml, re.DOTALL):
            fm = re.search(r'<(?:xdr:)?from>.*?<(?:xdr:)?col>(\d+)</(?:xdr:)?col>.*?'
                           r'<(?:xdr:)?row>(\d+)</(?:xdr:)?row>', anc, re.DOTALL)
            em = re.search(r'r:embed="([^"]+)"', anc)
            if fm and em and rid2media.get(em.group(1)) in real:
                placed.append((int(fm.group(2)), int(fm.group(1)), rid2media[em.group(1)]))
    placed.sort()
    return [p[2] for p in placed]


def image_captions(data, pictures):
    """Map each embedded micrograph to its picture caption via anchor order.

    Returns {image_name: caption}; empty when the image count doesn't match the
    caption count (caller falls back to magnification matching).
    """
    order = _anchor_order(data)
    numbered = [(int(m.group(1)), c) for l, c in (pictures or [])
                for m in [_PICNUM.search(l or '')] if m]
    # None-safe key: duplicate picture numbers with a missing caption would make
    # Python compare None to str and raise (crashing the whole review).
    caps = [c for _, c in sorted(numbered, key=lambda t: (t[0], t[1] or ''))]
    if not order or len(order) != len(caps):
        return {}
    return dict(zip(order, caps))


def _picture_image_pairs(data, pictures, images):
    """Align captions to embedded micrographs by anchor order.

    Returns [(label, caption, image_entry), …] when the picture count matches the
    embedded-micrograph count (so the pairing is trustworthy), else [].
    """
    order = _anchor_order(data)
    pics = sorted(((int(m.group(1)), l, c) for l, c in (pictures or [])
                   for m in [_PICNUM.search(l or '')] if m),
                  key=lambda t: (t[0], t[1] or '', t[2] or ''))   # None-safe
    if not order or len(order) != len(pics):
        return []
    by_name = {im.get('image'): im for im in (images or [])}
    pairs = []
    for (_, label, cap), name in zip(pics, order):
        im = by_name.get(name)
        if im is not None:
            pairs.append((label, cap, im))
    return pairs


def _requires_etch_status(label, caption):
    """Etch status is required for microstructure evidence, not every photo.

    Sampling-location photos and coating/thickness measurement images are valid
    without etching and previously caused false caption warnings.
    """
    text = f'{label or ""} {caption or ""}'
    return bool(re.search(r'\bmicrograph|microstructure\b', text, re.I))


def _review_captions(parsed):
    """Caption integrity: numbering, etch status, and comment picture references."""
    findings = []
    pics = parsed.get('pictures') or []
    if not pics:
        return findings
    comment = parsed.get('comment') or ''

    nums = []
    for label, _ in pics:
        m = _PICNUM.search(label or '')
        if m:
            nums.append(int(m.group(1)))

    dups = sorted({n for n in nums if nums.count(n) > 1})
    if dups:
        findings.append(('warning', 'Captions',
                         f'Caption picture number(s) repeated: {", ".join(map(str, dups))}.'))
    if nums:
        missing = sorted(set(range(1, max(nums) + 1)) - set(nums))
        if missing:
            findings.append(('info', 'Captions',
                             f'Picture numbering gap — missing {", ".join(map(str, missing))}.'))

    no_etch = [(label or '?').rstrip(':') for label, cap in pics
               if _requires_etch_status(label, cap)
               and not _ETCH_PAT.search(f"{label} {cap or ''}")]
    if no_etch:
        findings.append(('warning', 'Captions',
                         f'No etch status in caption(s): {", ".join(no_etch)}.'))
    else:
        findings.append(('pass', 'Captions',
                         'Every microstructure caption states an etch status.'))

    # Surface captions that explicitly state unetched / as-polished — legitimate
    # for thickness / crack work, but worth confirming for a microstructure report.
    for label, cap in pics:
        if _UNETCHED_PAT.search(f"{label} {cap or ''}"):
            findings.append(('info', 'Captions',
                             f'{(label or "?").rstrip(":")} caption states unetched / '
                             f'as-polished — confirm intended (a microstructure '
                             f'assessment is normally etched).'))

    refs = [int(m.group(1)) for m in
            re.finditer(r'pics?(?:ture)?s?\.?\s*(?:no\.?\s*)?(\d+)', comment, re.I)]
    for match in re.finditer(
            r'pics?(?:ture)?s?\.?\s*(?:no\.?\s*)?(\d+)\s*(?:-|–|to)\s*(\d+)',
            comment, re.I):
        refs.extend((int(match.group(1)), int(match.group(2))))
    # A duplicated caption number is already reported above. It must not also
    # make a valid reference to the fourth physical picture look out of range.
    max_present = max(max(nums) if nums else 0, len(pics))
    if refs and max(refs) > max_present:
        findings.append(('warning', 'Captions',
                         f'Comment refers to Picture {max(refs)} but the highest present '
                         f'caption is Picture {max_present}.'))

    return findings


def review_metallurgical(parsed):
    findings = []
    findings += _review_workbook_integrity(parsed)
    findings += _review_completeness(parsed)
    findings += _review_traceability(parsed)
    findings += _review_samples(parsed)
    findings += _review_sampling_basis(parsed)
    findings += _review_acceptance_and_methods(parsed)
    findings += _review_hardness(parsed['hardness'], parsed['sample'].get('material'))
    findings += _review_composition(
        parsed['nominal'], parsed['actual'], parsed.get('composition_meta'))
    findings += _review_alloy_identity(parsed)
    findings += _review_comment(parsed)
    findings += _review_evidence_claims(parsed)
    findings += _review_phase_evidence(parsed)
    findings += _review_document_control(parsed)
    findings += _review_captions(parsed)
    return findings


# ════════════════════════════════════════════════════════════════════════
# COATING REPORTS
# ════════════════════════════════════════════════════════════════════════
def _coating_signoff(wb):
    out, loc = {}, {}
    for ws in wb.worksheets:
        for key, pat in (('prepared', r'Prepared\s*by'),
                         ('approved', r'Approved\s*by'),
                         ('date',     r'^Date\s*:')):
            if key in out:
                continue
            lbl = _find(ws, pat)
            if lbl:
                out[key], vloc = _value_right_loc(ws, *lbl)
                loc[key] = {'sheet': ws.title, 'label': lbl, 'value': vloc}
    return out, loc


def parse_coating(wb, media=0):
    # The assessment sheet is the one carrying the actual MIN/MAX design
    # limits — not the Cover sheet, whose table-of-contents also mentions
    # "Coating Coverage Assessment".
    aws = None
    for ws in wb.worksheets:
        if _find(ws, r'Design\s*limit') and _find(ws, r'Measurements'):
            aws = ws
            break

    signoff, signoff_loc = _coating_signoff(wb)
    data = {'title': None, 'report_no': None, 'component': None, 'rows': [],
            'signoff': signoff, 'media': media,
            'loc': {'sheet': aws.title if aws is not None else None,
                    'signoff': signoff_loc}}

    cover = wb.worksheets[0]
    t = _find(cover, r'Coating')
    if t:
        data['title'] = _txt(cover.cell(row=t[0], column=t[1]).value)
    rn = _find(cover, r'Report\s*No')
    if rn:
        data['report_no'] = _value_right(cover, *rn)
    # Component (e.g. "2nd Stage Bucket") sits in the cover header text.
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=25):
            for cell in row:
                comp = _canon_component(_txt(cell.value))
                if comp:
                    data['component'] = comp
                    break
            if data['component']:
                break
        if data['component']:
            break

    if aws is None:
        return data

    meas_loc = _find(aws, r'Measurements')
    avg_loc  = _find(aws, r'Average\s*Values')
    min_loc  = _find(aws, r'^MIN$')
    max_loc  = _find(aws, r'^MAX$')
    if not (meas_loc and avg_loc and min_loc and max_loc):
        return data

    hrow = meas_loc[0]
    min_col, max_col = min_loc[1], max_loc[1]
    # Measurement value columns run from 'Measurements' up to the first summary
    # column (Average / MIN / MAX), and never include MIN/MAX/Average themselves
    # — otherwise a MIN/MAX in that span gets range-checked as a measurement.
    rights = [c for c in (avg_loc[1], min_col, max_col) if c > meas_loc[1]]
    right_bound = min(rights) if rights else aws.max_column + 1
    meas_cols = [c for c in range(meas_loc[1], right_bound) if c not in (min_col, max_col)]

    cur_min = cur_max = None
    for r in range(hrow + 1, aws.max_row + 1):
        m = _num(aws.cell(row=r, column=min_col).value)
        x = _num(aws.cell(row=r, column=max_col).value)
        if m is not None:
            cur_min = m
        if x is not None:
            cur_max = x
        cells = [(c, _num(aws.cell(row=r, column=c).value)) for c in meas_cols]
        cells = [(c, v) for c, v in cells if v is not None]
        if not cells:
            continue
        data['rows'].append({'row': r, 'values': [v for _, v in cells],
                             'cells': cells, 'min': cur_min, 'max': cur_max})
    return data


def review_coating(parsed):
    findings = _review_workbook_integrity(parsed)
    rows = parsed.get('rows', [])
    if not rows:
        findings.append(('warning', 'Coating', 'Could not read the coating-coverage assessment table.'))
        return findings

    out_of_range = 0
    total = 0
    limits_seen = False
    for entry in rows:
        lo, hi = entry['min'], entry['max']
        if lo is None or hi is None:
            continue
        limits_seen = True
        for v in entry['values']:
            total += 1
            if not (lo <= v <= hi):
                out_of_range += 1
                findings.append(('critical', 'Coating',
                                 f'Row {entry["row"]}: thickness {v:g} mm outside '
                                 f'design limit {lo:g}–{hi:g} mm.'))

    if not limits_seen:
        findings.append(('warning', 'Coating', 'No design MIN/MAX limits found to check against.'))
    elif out_of_range == 0:
        findings.append(('pass', 'Coating',
                         f'All {total} thickness measurements within design limits.'))

    so = parsed['signoff']
    missing = [lbl for key, lbl in (('prepared', 'Prepared by'), ('approved', 'Approved by'),
                                    ('date', 'Date')) if _is_placeholder(so.get(key))]
    if missing:
        findings.append(('warning', 'Sign-off', f'Missing sign-off field(s): {", ".join(missing)}.'))
    else:
        findings.append(('pass', 'Sign-off', 'Prepared-by, approved-by and date all present.'))

    micrographs = parsed.get('micrograph_count')
    if micrographs == 0:
        findings.append(('warning', 'Micrographs', 'No embedded reference micrographs found.'))
    return findings


# ════════════════════════════════════════════════════════════════════════
# MICROGRAPH LEGEND OCR  (light "read the legend in the photo" support)
# ════════════════════════════════════════════════════════════════════════
# Burned-in legends follow the AEG convention "<job>_E_<mag>x-<n>" at the
# bottom-left and a scale bar ("10 µm") at the bottom-right. OCR of such small,
# speckle-surrounded text is best-effort: values are correct when read, but not
# every image yields one. Findings are therefore advisory.
_MAG_PATS = [
    re.compile(r'(\d{1,5})\s*[xX%]\s*[-_]\s*(\d)'),   # 500x-1  (magnification + index)
    re.compile(r'E\s*[_ €F]?\s*(\d{1,5})\s*[xX%]'),   # E_500x
    re.compile(r'(?<![\d.])(\d{1,5})\s*[xX%]'),       # 500x
]
# Digit-bounded, not \b-bounded: "7420_E_500x-4" has no \w/\W boundary after
# the job number (underscore is a word char too), so \b(\d{4})\b never matches
# the documented convention above.
_JOB_PAT   = re.compile(r'(?<!\d)(\d{4})(?!\d)')
_SCALE_PAT = re.compile(r'(\d{1,3})\s*[µuμyptwb]+m', re.I)
_CAP_MAG   = re.compile(r'(\d{1,5})\s*[xX]\b')


def _magnification_from_ocr(text):
    """One plausible magnification from a single OCR preprocessing pass."""
    for pattern in _MAG_PATS:
        match = pattern.search(text or '')
        if match:
            value = int(match.group(1))
            if 1 <= value <= 50000:
                return value
    return None


def _select_magnification(ocr_reads):
    """Select a stable OCR magnification by voting across preprocessing reads.

    A number is accepted only when at least two independent threshold passes
    agree and there is no tie. One isolated read such as 600x is retained in
    the vote metadata but never promoted to a reported magnification.
    """
    candidates = [
        value for value in (_magnification_from_ocr(text) for text in ocr_reads)
        if value is not None
    ]
    votes = Counter(candidates)
    if not votes:
        return None, {}
    ranked = votes.most_common()
    value, count = ranked[0]
    tied = len(ranked) > 1 and ranked[1][1] == count
    return (value if count >= 2 and not tied else None), dict(votes)


def _caption_magnification(caption):
    """Single written caption magnification (the report's source of record)."""
    values = {int(match.group(1)) for match in _CAP_MAG.finditer(caption or '')}
    return f'{next(iter(values))}x' if len(values) == 1 else None


def _safe_ocr(im, cfg='--psm 7'):
    try:
        if pytesseract is not None:
            return pytesseract.image_to_string(im, config=cfg) or ''
        if not _TESSERACT_BIN:
            return ''
        payload = io.BytesIO()
        im.save(payload, format='PNG')
        result = subprocess.run(
            [_TESSERACT_BIN, 'stdin', 'stdout', *shlex.split(cfg)],
            input=payload.getvalue(), stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL, check=False, timeout=20,
        )
        return result.stdout.decode('utf-8', errors='replace') if result.returncode == 0 else ''
    except Exception:
        return ''


def _safe_ocr_tsv(im, cfg='--psm 11'):
    """Return Tesseract word boxes as TSV, with or without pytesseract."""
    try:
        if pytesseract is not None:
            return pytesseract.image_to_data(im, config=cfg) or ''
        if not _TESSERACT_BIN:
            return ''
        payload = io.BytesIO()
        im.save(payload, format='PNG')
        result = subprocess.run(
            [_TESSERACT_BIN, 'stdin', 'stdout', *shlex.split(cfg), 'tsv'],
            input=payload.getvalue(), stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL, check=False, timeout=20,
        )
        return result.stdout.decode('utf-8', errors='replace') if result.returncode == 0 else ''
    except Exception:
        return ''


def _binarize(im, thr, scale=4):
    """Keep bright text (white-on-dark legend bar) and upscale small fonts."""
    return im.point(lambda p: 255 if p > thr else 0).resize(
        (max(1, im.width * scale), max(1, im.height * scale)))


_ETCH_THR = 0.05   # edge-density below this ⇒ image looks unetched / very low contrast


def _edge_density(im):
    """Fraction of strong edges in the image body — high ⇒ etched, low ⇒ unetched."""
    if not _PIL_AVAILABLE:
        return None
    w, h = im.size
    c = im.crop((int(w * 0.15), int(h * 0.15), int(w * 0.85), int(h * 0.80)))
    try:
        px = list(c.filter(ImageFilter.FIND_EDGES).get_flattened_data())
    except Exception:
        return None
    return sum(1 for p in px if p > 40) / len(px) if px else None


def _read_legend_im(im):
    """OCR the burned-in legend (ID / magnification / scale-bar) of one micrograph."""
    if not _OCR_AVAILABLE:
        return {}
    w, h = im.size
    lc = im.crop((0, int(h * 0.90), int(w * 0.55), h))           # ID + magnification
    rc = im.crop((int(w * 0.72), int(h * 0.88), w, h))           # scale bar
    lreads = [_safe_ocr(_binarize(lc, t)) for t in (110, 130, 150)]
    lblob = ' '.join(lreads)
    rblob = ' '.join(_safe_ocr(_binarize(rc, t)) for t in (110, 140))

    out = {}
    job_m = _JOB_PAT.search(lblob)
    mag_val, votes = _select_magnification(lreads)
    if votes:
        out['mag_votes'] = {f'{value}x': count for value, count in sorted(votes.items())}
        out['mag_read_count'] = len(lreads)
    if mag_val is not None:
        mag = f'{mag_val}x'
        out['ocr_mag'] = mag
        out['mag'] = mag
        out['mag_source'] = 'ocr'
        out['mag_confidence'] = votes[mag_val] / len(lreads)
        idx = None
        for read in lreads:
            match = _MAG_PATS[0].search(read or '')
            if match and int(match.group(1)) == mag_val:
                idx = match.group(2)
                break
        out['id'] = (f'{job_m.group(1)}_' if job_m else '') + f'E_{mag_val}x' + \
                    (f'-{idx}' if idx else '')
    if job_m:
        out['job'] = job_m.group(1)
    s = _SCALE_PAT.search(rblob) or _SCALE_PAT.search(lblob)
    if s:
        out['scale'] = f'{s.group(1)} µm'
    return out


def _measurement_color(image, bbox, pad=80):
    """Nearest saturated annotation color to one OCR measurement label."""
    left, top, right, bottom = bbox
    x0, y0 = max(0, left - pad), max(0, top - pad)
    x1, y1 = min(image.width, right + pad), min(image.height, bottom + pad)
    crop = image.crop((x0, y0, x1, y1)).convert('RGB')
    distances = {name: float('inf') for name in
                 ('red', 'blue', 'magenta', 'yellow', 'green', 'cyan')}
    counts = Counter()
    for y in range(crop.height):
        for x in range(crop.width):
            red, green, blue = crop.getpixel((x, y))
            color = None
            if red > 120 and blue > 80 and min(red, blue) > green + 40:
                color = 'magenta'
            elif red > 130 and red > green + 55 and red > blue + 45:
                color = 'red'
            elif blue > 110 and blue > red + 45 and blue > green + 20:
                color = 'blue'
            elif red > 140 and green > 110 and blue < 90:
                color = 'yellow'
            elif green > 120 and green > red + 40 and green > blue + 30:
                color = 'green'
            elif green > 110 and blue > 110 and min(green, blue) > red + 35:
                color = 'cyan'
            if color is None:
                continue
            counts[color] += 1
            absolute_x, absolute_y = x0 + x, y0 + y
            dx = max(left - absolute_x, 0, absolute_x - right)
            dy = max(top - absolute_y, 0, absolute_y - bottom)
            distances[color] = min(distances[color], (dx * dx + dy * dy) ** 0.5)
    plausible = [name for name, distance in distances.items()
                 if distance <= 60 and counts[name] >= 5]
    return min(plausible, key=distances.get) if plausible else 'unclassified'


def _read_measurement_data_im(im):
    """Read thickness values and retain their colored annotation series."""
    if not _OCR_AVAILABLE:
        return [], {}
    source = im.convert('RGB')
    gray = source.convert('L')
    scale = 3
    # Include low-placed callouts while excluding the bottom legend/scale bar.
    body = gray.crop((0, 0, gray.width, int(gray.height * 0.91)))
    big = body.resize((body.width * scale, body.height * scale))
    bright = big.point(lambda pixel: 255 if pixel > 180 else 0)
    tsv = _safe_ocr_tsv(bright)
    lines = {}
    try:
        rows = csv.DictReader(io.StringIO(tsv), delimiter='\t')
        for row in rows:
            if not (row.get('text') or '').strip():
                continue
            key = (row.get('block_num'), row.get('par_num'), row.get('line_num'))
            lines.setdefault(key, []).append(row)
    except Exception:
        lines = {}

    values = []
    groups = {}
    for words in lines.values():
        text = ' '.join(row.get('text') or '' for row in words)
        matches = list(re.finditer(r'(?<!\d)(\d{1,3})\s*[µuμpy]m\b', text, re.I))
        if len(matches) != 1:
            continue
        value = int(matches[0].group(1))
        try:
            bbox = (
                min(int(row['left']) for row in words) // scale,
                min(int(row['top']) for row in words) // scale,
                max(int(row['left']) + int(row['width']) for row in words) // scale,
                max(int(row['top']) + int(row['height']) for row in words) // scale,
            )
        except (KeyError, TypeError, ValueError):
            continue
        color = _measurement_color(source, bbox)
        values.append(value)
        groups.setdefault(color, []).append(value)

    if not values:
        # String-only fallback for unusual TSV output. The accepted p/y glyphs
        # are common micro-symbol OCR errors and remain constrained to the body.
        text = _safe_ocr(bright, '--psm 11')
        values = [int(value) for value in re.findall(
            r'(\d{1,3})\s*[µuμpy]m\b', text, re.I)]
        if values:
            groups['unclassified'] = list(values)
    return sorted(values), {color: sorted(series) for color, series in groups.items()}


def _read_measurements_im(im):
    """Back-compatible list-only measurement reader."""
    values, _groups = _read_measurement_data_im(im)
    return values


def analyze_images(data, want_bytes=False, max_images=40):
    """Single pass over embedded micrographs.

    Returns (images, ocr_used) where each image dict carries:
      'image', 'strong', 'etched', 'measurements', optional 'mag'/'scale'/'id'/'job',
      and 'bytes'/'ext' when want_bytes is set.
    """
    images = []
    if not _PIL_AVAILABLE:
        return images, False
    try:
        z = zipfile.ZipFile(io.BytesIO(data))
        names = sorted(n for n in z.namelist() if n.startswith('xl/media'))
    except Exception:
        return images, _OCR_AVAILABLE
    for n in names[:max_images]:
        raw = z.read(n)
        try:
            source_image = Image.open(io.BytesIO(raw)).convert('RGB')
            im = source_image.convert('L')
        except Exception:
            continue
        w, h = im.size
        if w < 200 or h < 150:           # skip logos / thumbnails
            continue
        strong = _edge_density(im)
        measurements, measurement_groups = _read_measurement_data_im(source_image)
        entry = {'image': n.split('/')[-1],
                 'strong': strong,
                 'etched': (strong is None) or (strong >= _ETCH_THR),
                 'measurements': measurements,
                 'measurement_groups': measurement_groups}
        entry.update(_read_legend_im(im))
        if want_bytes:
            entry['bytes'] = raw
            entry['ext'] = n.rsplit('.', 1)[-1].lower()
        images.append(entry)
    return images, _OCR_AVAILABLE


def read_image_legends(data, max_images=40):
    """Back-compat: the legend subset of analyze_images()."""
    images, ocr_used = analyze_images(data, max_images=max_images)
    legends = [im for im in images
               if im.get('ocr_mag') or im.get('scale') or im.get('job') or im.get('id')]
    return legends, ocr_used


def _comment_thickness_um(comment):
    """Thickness values in the comment text, normalised to µm."""
    return {claim['value_um'] for claim in _comment_thickness_claims(comment)}


def _sentence_context(text, position):
    """Sentence containing a position, without splitting decimal numbers."""
    boundaries = [0]
    boundaries.extend(match.end() for match in re.finditer(
        r'(?<!\d)[.!?]\s+(?=[A-Z])|\n+', text))
    boundaries.append(len(text))
    for left, right in zip(boundaries, boundaries[1:]):
        if left <= position < right:
            return text[left:right].strip()
    return text


def _picture_references(text):
    refs = set()
    pattern = re.compile(
        r'\bpic(?:ture)?s?\.?\s*(?:no\.?\s*)?'
        r'(\d+(?:\s*(?:,\s*(?:and\s*)?|&|and|to|-|–)\s*\d+)*)',
        re.I,
    )
    for match in pattern.finditer(text or ''):
        body = match.group(1)
        refs.update(int(value) for value in re.findall(r'\d+', body))
        for range_match in re.finditer(r'(\d+)\s*(?:to|-|–)\s*(\d+)', body, re.I):
            first, last = map(int, range_match.groups())
            if 0 < first <= last <= 100:
                refs.update(range(first, last + 1))
    return sorted(refs)


def _comment_thickness_claims(comment):
    """Quantitative thickness statements with their local wording/context."""
    text = comment or ''
    claims = []
    for match in re.finditer(
            r'(?P<qual><\s*|≤\s*|up\s+to\s+|maximum\s+(?:of\s+)?|max\.?\s*)?'
            r'(?P<value>\d+(?:\.\d+)?)\s*(?P<unit>mm|µm|um|μm)\b',
            text, re.I):
        value = float(match.group('value'))
        value_um = round(value * 1000) if match.group('unit').lower() == 'mm' else round(value)
        context = _sentence_context(text, match.start())
        qual = (match.group('qual') or '').strip().lower()
        refs = _picture_references(context)
        claims.append({
            'value_um': value_um,
            'raw': match.group(0).strip(),
            'context': context,
            'picture_refs': refs,
            'is_limit': qual.startswith('<') or qual.startswith('≤'),
            'is_max': 'up to' in qual or 'max' in qual,
        })
    # The same value is often repeated in a summary sentence. Preserve the
    # first, most local evidence statement rather than issuing duplicates.
    unique = {}
    for claim in claims:
        key = (claim['value_um'], claim['is_limit'], claim['is_max'])
        unique.setdefault(key, claim)
    return list(unique.values())


def picture_etch_verdicts(images, pictures, data):
    """Per-picture caption↔contrast verdicts via 1:1 caption/micrograph pairing.

    Returns a list of {'index', 'label', 'severity', 'note'} (possibly empty) when
    captions pair to micrographs, else None (caller falls back to the aggregate
    count). `index` is the position in `pictures`, so a caller can find the caption
    cell. Contrast is advisory, so verdicts read "verify".
    """
    if not data:
        return None
    pairs = _picture_image_pairs(data, pictures, images)
    if not pairs:
        return None
    idx_of = {}
    for i, (label, _) in enumerate(pictures or []):
        idx_of.setdefault(label, i)
    out = []
    for label, cap, im in pairs:
        if im.get('strong') is None:
            continue
        pic = (label or 'Picture').rstrip(':')
        et = caption_etchant(f"{label} {cap or ''}")
        named = et and et not in ('Unetched', 'Etched (unspecified)')
        note = None
        if named and not im.get('etched'):
            note = (f'{pic}: caption names {et} but the micrograph reads low-contrast — '
                    f'the etch may not have developed; verify (contrast is advisory).')
        elif et == 'Unetched' and im.get('etched'):
            note = (f'{pic}: caption says unetched but the micrograph reads etched-type '
                    f'contrast — verify (contrast is advisory).')
        if note:
            out.append({'index': idx_of.get(label), 'label': label,
                        'severity': 'warning', 'note': note})
    return out


def picture_magnification_verdicts(images, pictures, data):
    """Reconcile each written caption with its paired micrograph's stable OCR.

    The written caption is the report's source of record. OCR is accepted only
    after multi-pass consensus and is retained as secondary evidence. A
    disagreement is therefore a warning to inspect the burned-in legend, never
    a rejection of either number.
    """
    if not data:
        return None
    pairs = _picture_image_pairs(data, pictures, images)
    if not pairs:
        return None
    idx_of = {}
    for i, (label, _) in enumerate(pictures or []):
        idx_of.setdefault(label, i)
    out = []
    for label, cap, im in pairs:
        caption_mag = _caption_magnification(cap)
        ocr_mag = im.get('ocr_mag') or (
            im.get('mag') if im.get('mag_source') == 'ocr' else None)
        if caption_mag:
            im['caption_mag'] = caption_mag
            im['mag'] = caption_mag
            im['mag_source'] = 'caption'
        if not (caption_mag and ocr_mag and caption_mag != ocr_mag):
            continue
        # An ID built from the disputed OCR magnitude is not safe to use for
        # duplicate-ID checks or downstream metadata.
        im.pop('id', None)
        votes = im.get('mag_votes') or {}
        count = votes.get(ocr_mag)
        total = im.get('mag_read_count')
        consensus = (
            f' in {count} of {total} preprocessing passes'
            if count is not None and total else ''
        )
        pic = (label or 'Picture').rstrip(':')
        note = (
            f'{pic}: written caption states {caption_mag}, while legend OCR '
            f'consistently read {ocr_mag}{consensus}. Treat the caption as the '
            f'recorded value and verify the burned-in legend; OCR may be wrong.'
        )
        out.append({
            'index': idx_of.get(label),
            'label': label,
            'severity': 'warning',
            'note': note,
        })
    return out


def _review_etch(images, pictures, verdicts):
    """Aggregate contrast summary plus the per-picture caption↔contrast verdicts.

    `verdicts` is picture_etch_verdicts(...): a list (per-picture findings) or
    None (couldn't pair 1:1 → fall back to the aggregate unetched-vs-low count).
    """
    findings = []
    scored = [im for im in images if im.get('strong') is not None]
    if not scored:
        return findings
    n_low = sum(1 for im in scored if not im.get('etched'))
    findings.append(('info', 'Photo etch',
                     f'{len(scored) - n_low} of {len(scored)} micrograph(s) show etched-type '
                     f'contrast; {n_low} low-contrast (unetched / faint post-HT).'))
    if verdicts is not None:
        for v in verdicts:
            findings.append((v['severity'], 'Photo etch', v['note']))
    else:
        n_cap = sum(1 for label, cap in (pictures or [])
                    if _UNETCHED_PAT.search(f"{label} {cap or ''}"))
        if n_low != n_cap:
            findings.append(('info', 'Photo etch',
                             f'{n_low} micrograph(s) read as low-contrast vs {n_cap} caption(s) '
                             f'marked "unetched" — worth a glance (faint post-HT etch reads low).'))
    return findings


def _review_thickness(parsed, images, ocr_used=True, data=None):
    """Reconcile reported thickness with burned-in photo measurements.

    A reported average is accepted only when it reproduces an image-group (or
    cited-image) mean within a practical OCR/rounding tolerance. Merely falling
    somewhere between the minimum and maximum is not evidence of an average.
    """
    findings = []
    claims = _comment_thickness_claims(parsed.get('comment'))
    if not claims:
        return findings
    if not ocr_used:
        return findings

    pairs = _picture_image_pairs(data, parsed.get('pictures') or [], images)
    numbered_images = {}
    for label, _caption, image in pairs:
        number = _PICNUM.search(label or '')
        if number:
            numbered_images.setdefault(int(number.group(1)), []).append(image)

    for claim in claims:
        selected_images = [image for number in claim['picture_refs']
                           for image in numbered_images.get(number, [])]
        if not selected_images:
            selected_images = list(images)

        individual_groups = []
        aggregate_by_color = {}
        values = []
        for image in selected_images:
            image_values = list(image.get('measurements') or [])
            values.extend(image_values)
            color_groups = image.get('measurement_groups') or {}
            if color_groups:
                for color, series in color_groups.items():
                    if not series:
                        continue
                    individual_groups.append(list(series))
                    aggregate_by_color.setdefault(color, []).extend(series)
            elif image_values:
                individual_groups.append(image_values)
                aggregate_by_color.setdefault('unclassified', []).extend(image_values)

        # When the report explicitly cites several pictures, one colored series
        # represents the same feature across those pictures. Compare the claim
        # with that combined series; otherwise keep image-level series separate.
        if len(set(claim['picture_refs'])) > 1 and aggregate_by_color:
            groups = list(aggregate_by_color.values())
        else:
            groups = individual_groups
        if not values:
            findings.append(('warning', 'Thickness evidence',
                             f'The report states {claim["raw"]}, but no burned-in '
                             'measurement was read from the cited micrograph(s); verify the '
                             'measurement method and source data.'))
            continue

        target = claim['value_um']
        tolerance = max(3.0, target * 0.12)
        means = []
        for group in groups:
            if not group:
                continue
            ordered = sorted(group)
            means.append(sum(ordered) / len(ordered))
            # One bad leading digit is common in tiny callouts (18 → 48). For a
            # sufficiently populated series, retain a one-at-each-end trimmed
            # mean as a secondary candidate; the raw range remains disclosed.
            if len(ordered) >= 5:
                trimmed = ordered[1:-1]
                means.append(sum(trimmed) / len(trimmed))

        if claim['is_limit']:
            supported = max(values) < target + tolerance
        elif claim['is_max']:
            supported = abs(max(values) - target) <= tolerance
        else:
            supported = any(abs(mean - target) <= tolerance for mean in means)
        if supported:
            continue

        evidence = sorted({round(mean, 1) for mean in means})
        finding = (
            f'Reported thickness {claim["raw"]} is not reproduced by the burned-in '
            f'measurements (range {min(values):g}–{max(values):g} µm'
            + (f'; candidate mean(s) {", ".join(f"{value:g}" for value in evidence)} µm'
               if evidence else '')
            + '). Verify the stated calculation and cited image.'
        )
        findings.append(('warning', 'Thickness evidence', finding))
    return findings


def _caption_mags(pictures):
    """Magnifications mentioned in the written picture captions, e.g. {'200x'}."""
    return {mag for _, cap in pictures or []
            for mag in [_caption_magnification(cap)] if mag}


def _digit_dist(a, b):
    """Positional digit difference between two same-length strings; len-gap otherwise."""
    if len(a) != len(b):
        return max(len(a), len(b))
    return sum(x != y for x, y in zip(a, b))


def _review_legends(legends, ocr_used, caption_mags, report_job=None):
    findings = []
    if not ocr_used:
        findings.append(('info', 'Photo legends',
                         'Legend OCR unavailable (Tesseract not installed) — skipped.'))
        return findings
    if not legends:
        findings.append(('info', 'Photo legends',
                         'Could not read a legend from any embedded micrograph.'))
        return findings

    img_mags = sorted({
        l.get('ocr_mag') or l.get('mag')
        for l in legends if l.get('ocr_mag') or l.get('mag')
    },
                      key=lambda s: int(s[:-1]))
    findings.append(('info', 'Photo legends',
                     f'Read legends from {len(legends)} micrograph(s); '
                     f'stable OCR magnifications: '
                     f'{", ".join(img_mags) if img_mags else "n/a"}.'))

    # Two visibly different micrographs sharing one burned-in ID means at
    # least one was captured/labelled wrong — the AEG "-n" suffix exists to
    # keep these unique, so a repeat is never expected.
    ids = [l['id'] for l in legends if l.get('id')]
    dupes = sorted({i for i in ids if ids.count(i) > 1})
    if dupes:
        findings.append(('warning', 'Photo legends',
                         f'Multiple micrographs share the identical burned-in ID '
                         f'{", ".join(dupes)} — verify they aren\'t the same image mislabeled.'))

    # Cross-check the job number burned into every legend against the report.
    # A previous implementation passed the whole set as soon as ONE image
    # matched, which hid a stray image from another job. Mixed exact/non-exact
    # readings are therefore always surfaced. A uniform one-digit mismatch is
    # still informational because it can be a repeatable OCR error.
    legend_jobs = [l['job'] for l in legends if l.get('job')]
    if report_job and report_job.isdigit() and legend_jobs:
        distinct = sorted(set(legend_jobs))
        mismatched = sorted({job for job in legend_jobs if job != report_job})
        if not mismatched:
            findings.append(('pass', 'Photo legends',
                             f'Micrograph legends carry the report job number ({report_job}).'))
        elif report_job in legend_jobs:
            findings.append(('warning', 'Photo legends',
                             f'Mixed micrograph job numbers: report {report_job}, but some '
                             f'images read {", ".join(mismatched)} — verify that no image was '
                             f'copied from another report.'))
        else:
            best = min(legend_jobs, key=lambda j: _digit_dist(j, report_job))
            if len(distinct) > 1 or _digit_dist(best, report_job) >= 2:
                seen = ", ".join(distinct)
                findings.append(('warning', 'Photo legends',
                                 f'Legend job number(s) [{seen}] do not match the report job '
                                 f'{report_job} — verify the micrographs belong to this report '
                                 f'(or an OCR misread).'))
            else:
                findings.append(('info', 'Photo legends',
                                 f'Legend job numbers are within one digit of the report job '
                                 f'({report_job}) — likely OCR variance.'))
    return findings


# ════════════════════════════════════════════════════════════════════════
# REPORT TITLE / FILENAME vs CONTENT  (catch a misidentified workbook)
# ════════════════════════════════════════════════════════════════════════
# Component synonyms (GE terminology): bucket≡blade (rotating), vane≡nozzle
# (stationary). Order matters — multi-word parts first.
_PART_SYNONYMS = [
    (r'transition\s*piece',  'transition piece'),
    (r'combustion\s*liner',  'combustion liner'),
    (r'\bliner\b',           'combustion liner'),
    (r'\bbucket\b|\bblade\b', 'bucket'),
    (r'\bvane\b|\bnozzle\b',  'vane'),
    (r'\bshroud\b',          'shroud'),
    (r'\bdiaphragm\b',       'diaphragm'),
    (r'\bseal\b',            'seal'),
]


def _component_identity(text):
    """Return (stage_number, canonical_part) from free report-title text."""
    raw = re.sub(r'[_-]+', ' ', text or '')
    lowered = raw.lower()
    part = next((name for pat, name in _PART_SYNONYMS if re.search(pat, lowered)), None)
    stage_match = (
        re.search(r'\b(\d{1,2})\s*(?:st|nd|rd|th)?\s*(?:stage|stg)\b', lowered)
        or re.search(r'\b(?:stage|stg)\s*(\d{1,2})\b', lowered)
    )
    stage = int(stage_match.group(1)) if stage_match else None
    return stage, part


def _canon_component(text):
    """Canonical 'stage + part' from free text, e.g. '2nd Stage Bucket' → '2 bucket'."""
    stage, part = _component_identity(text)
    if part is None:
        return None
    return (f'{stage} ' if stage is not None else '') + part


def _canon_machine(text):
    """Canonical machine/set designation from a filename or report header.

    The aliases below reflect the forms used by the supplied reports:
    FS.7 ≡ MS7001 and 7FA ≡ MS7001FA. V-series names remain explicit.
    """
    raw = (text or '').upper()

    v_model = re.search(r'(?<![A-Z0-9])V\s*(\d{2})\s*[.\-]?\s*(\d)(?!\d)', raw)
    if v_model:
        return f'V{v_model.group(1)}.{v_model.group(2)}'

    ms_model = re.search(r'(?<![A-Z0-9])MS\s*([679])\s*001\s*([A-Z]{1,3})?', raw)
    if ms_model:
        frame, suffix = ms_model.group(1), ms_model.group(2) or ''
        return f'{frame}{suffix}' if suffix else f'MS{frame}001'

    fs_model = re.search(r'(?<![A-Z0-9])FS\s*[.\-]?\s*([679])\s*(FA)?(?![A-Z0-9])', raw)
    if fs_model:
        if fs_model.group(2):
            return f'{fs_model.group(1)}FA'
        return f'MS{fs_model.group(1)}001'

    short_f = re.search(r'(?<![A-Z0-9])([679])\s*F\s*([A-Z]?)(?![A-Z0-9])', raw)
    if short_f:
        return f'{short_f.group(1)}F{short_f.group(2)}'

    return None


def _content_job(parsed, rtype):
    """4-digit AEG job number from the report content, for either report family."""
    if rtype == 'metallurgical':
        m = re.search(r'\d{4}', parsed.get('header', {}).get('job') or '')
    else:
        m = re.search(r'\d{4}', parsed.get('report_no') or '')
    return m.group() if m else ''


def review_filename(filename, parsed, rtype):
    """Check that the report title/filename agrees with its contents."""
    findings = []
    name = re.sub(r'\.xlsx?$', '', os.path.basename(filename or ''), flags=re.I)
    if not name:
        return findings
    low = name.lower()
    matched = []
    category = 'Title identity'

    # Job number (title vs content). Digit-bounded, not \b-bounded — an
    # underscore right after the digits (e.g. "7504_AEN_Saudi...") is a word
    # char too, so \b(\d{4})\b would never match this very common convention.
    fjob = re.search(r'(?<!\d)(\d{4})(?!\d)', name)
    cjob = _content_job(parsed, rtype)
    if fjob and cjob:
        if fjob.group(1) == cjob:
            matched.append('job')
        else:
            findings.append(('critical', category,
                             f'Report title job number {fjob.group(1)} does not match '
                             f'the internal AEG job number {cjob}.'))
    elif cjob and not fjob:
        findings.append(('warning', category,
                         f'Report title does not state the internal AEG job number {cjob}.'))
    elif fjob and not cjob:
        findings.append(('warning', category,
                         f'Report title states job {fjob.group(1)}, but no internal '
                         f'AEG job number was found.'))

    # Report type.
    if 'coating' in low and rtype == 'metallurgical':
        findings.append(('critical', category,
                         'Report title says "Coating" but the content is metallurgical.'))
    elif re.search(r'metallurg', low) and rtype == 'coating':
        findings.append(('critical', category,
                         'Report title says "Metallurgical" but the content is a coating report.'))
    elif ('coating' in low and rtype == 'coating') or \
         (re.search(r'metallurg', low) and rtype == 'metallurgical'):
        matched.append('type')

    content_description = (
        parsed.get('sample', {}).get('description')
        if rtype == 'metallurgical'
        else parsed.get('component')
    ) or ''
    fstage, fpart = _component_identity(name)
    cstage, cpart = _component_identity(content_description)

    # Stage and component are checked independently so a title cannot pass by
    # getting one of the two right.
    if fstage is not None and cstage is not None:
        if fstage == cstage:
            matched.append('stage')
        else:
            findings.append(('critical', category,
                             f'Report title says Stage {fstage}, but the internal '
                             f'component description says Stage {cstage} '
                             f'("{content_description}").'))
    elif cstage is not None and fstage is None:
        findings.append(('warning', category,
                         f'Report title does not state Stage {cstage} from the internal '
                         f'component description "{content_description}".'))
    elif fstage is not None and cstage is None:
        findings.append(('warning', category,
                         f'Report title says Stage {fstage}, but the internal component '
                         f'description does not state a stage.'))

    if fpart and cpart:
        if fpart == cpart:
            matched.append('component')
        else:
            findings.append(('critical', category,
                             f'Report title component "{fpart}" does not match the '
                             f'internal component "{cpart}" ("{content_description}").'))
    elif cpart and not fpart:
        findings.append(('warning', category,
                         f'Report title does not state the internal component "{cpart}".'))
    elif fpart and not cpart:
        findings.append(('warning', category,
                         f'Report title states component "{fpart}", but no internal '
                         f'component name was found.'))

    # Machine/set designation: cover both the explicit V-series wording and GE
    # aliases used in these files (FS.7/MS7001 and 7FA/MS7001FA).
    content_machine = (
        parsed.get('header', {}).get('machine') if rtype == 'metallurgical' else None
    ) or ''
    fmachine = _canon_machine(name)
    cmachine = _canon_machine(content_machine)
    if rtype == 'metallurgical':
        if fmachine and cmachine:
            if fmachine == cmachine:
                matched.append('machine/set')
            else:
                findings.append(('critical', category,
                                 f'Report title machine/set "{fmachine}" does not match '
                                 f'the internal Machine Type "{content_machine}" '
                                 f'("{cmachine}").'))
        elif cmachine and not fmachine:
            findings.append(('warning', category,
                             f'Report title does not state the internal machine/set '
                             f'"{content_machine}".'))
        elif fmachine and not cmachine:
            findings.append(('warning', category,
                             f'Report title states machine/set "{fmachine}", but the '
                             f'internal Machine Type is blank or unrecognised.'))

    # Customer (advisory, lenient — pass on any shared word ≥3 chars).
    ccust = parsed.get('header', {}).get('customer') if rtype == 'metallurgical' else None
    if ccust:
        ctoks = set(re.findall(r'[a-z]{3,}', ccust.lower()))
        if ctoks and not (ctoks & set(re.findall(r'[a-z]{3,}', low))):
            findings.append(('info', category,
                             f'Report title customer does not obviously match the '
                             f'internal customer "{ccust}".'))

    if matched and not any(
            c == category and s in ('critical', 'warning') for s, c, _ in findings):
        findings.append(('pass', category,
                         f'Report title agrees with the content ({", ".join(matched)}).'))
    return findings


def _revision_score(name):
    """Conservative filename revision rank used only for co-uploaded versions."""
    stem = os.path.splitext(os.path.basename(name or ''))[0].lower().strip()
    if re.search(r'(?:^|[\s_-])-?r$', stem):
        return 3
    if any(token in stem for token in ('updated', 'revised', 'revision')):
        return 2
    if any(token in stem for token in ('final', 'done')):
        return 1
    return 0


def add_version_findings(reports):
    """Flag persistent blockers in explicit revised versions of one job."""
    groups = {}
    for report in reports:
        if 'error' in report:
            continue
        job = ((report.get('parsed') or {}).get('header') or {}).get('job')
        if job:
            groups.setdefault(str(job).strip(), []).append(report)

    for job, group in groups.items():
        if len(group) < 2:
            continue
        ranked = sorted(group, key=lambda report: _revision_score(report.get('name')))
        earlier, revised = ranked[0], ranked[-1]
        if _revision_score(revised.get('name')) <= _revision_score(earlier.get('name')):
            continue

        earlier_blockers = {
            (category, message)
            for severity, category, message in earlier.get('findings') or []
            if severity in ('critical', 'warning')
        }
        persistent = [
            (severity, category, message)
            for severity, category, message in revised.get('findings') or []
            if severity in ('critical', 'warning')
            and (category, message) in earlier_blockers
        ]
        if not persistent:
            continue
        fail_count = sum(severity == 'critical' for severity, _cat, _msg in persistent)
        warning_count = sum(severity == 'warning' for severity, _cat, _msg in persistent)
        severity = 'critical' if fail_count else 'warning'
        message = (
            f'Revised job {job} workbook still contains {fail_count} fail(s) and '
            f'{warning_count} warning(s) already present in the earlier uploaded '
            f'version ({earlier["name"]}); the revision did not close the recorded gaps.'
        )
        finding = (severity, 'Revision control', message)
        revised.setdefault('version_findings', []).append(finding)
        revised['findings'] = list(revised.get('findings') or []) + [finding]


# ════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ════════════════════════════════════════════════════════════════════════
def _media_count(data):
    try:
        z = zipfile.ZipFile(io.BytesIO(data))
        return sum(1 for n in z.namelist() if n.startswith('xl/media'))
    except Exception:
        return 0


def review_report(filename, data, ocr=True):
    """Review one report. Returns (report_type, parsed, findings).

    ocr : when True (and the OCR stack is available) the burned-in legend of
          each embedded micrograph is read and cross-checked against captions.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    rtype = detect_type(wb)
    media = _media_count(data)
    excel_errors = _excel_errors(wb)
    formula_issues = _formula_issues(wb_formulas)
    micrograph_names = _anchor_order(data) if _PIL_AVAILABLE else None
    micrograph_count = len(micrograph_names) if micrograph_names is not None else None

    if rtype == 'coating':
        parsed = parse_coating(wb, media)
        parsed['excel_errors'] = excel_errors
        parsed['formula_issues'] = formula_issues
        parsed['micrograph_count'] = micrograph_count
        findings = review_coating(parsed)
    elif rtype == 'metallurgical':
        parsed = parse_metallurgical(wb, media, micrograph_count)
        parsed['excel_errors'] = excel_errors
        parsed['formula_issues'] = formula_issues
        parsed['document_text'] = _workbook_text(wb)
        parsed['page_control'] = _page_control(wb_formulas)
        findings = review_metallurgical(parsed)
    else:
        parsed = {'excel_errors': excel_errors, 'formula_issues': formula_issues,
                  'micrograph_count': micrograph_count}
        findings = [('warning', 'Format',
                     'Unrecognised layout — not classified as a metallurgical or coating report.')]
        findings += _review_workbook_integrity(parsed)

    title_findings = review_filename(filename, parsed, rtype)
    parsed['title_findings'] = title_findings
    findings += title_findings

    images = []
    ocr_used = False
    if ocr:
        images, ocr_used = analyze_images(data)
        mag_verdicts = picture_magnification_verdicts(
            images, parsed.get('pictures', []), data)
        for verdict in mag_verdicts or []:
            findings.append((verdict['severity'], 'Magnification', verdict['note']))
        parsed['photo_magnification'] = mag_verdicts or []
        legends = [im for im in images
                   if im.get('ocr_mag') or im.get('scale')
                   or im.get('job') or im.get('id')]
        cap_mags = _caption_mags(parsed.get('pictures', []))
        report_job = parsed.get('header', {}).get('job')
        findings += _review_legends(legends, ocr_used, cap_mags, report_job)
        etch_verdicts = picture_etch_verdicts(images, parsed.get('pictures', []), data)
        findings += _review_etch(images, parsed.get('pictures', []), etch_verdicts)
        thickness_findings = _review_thickness(
            parsed, images, ocr_used=ocr_used, data=data)
        findings += thickness_findings
        parsed['thickness_findings'] = thickness_findings
        parsed['photo_etch'] = etch_verdicts or []
    parsed['ocr_used'] = ocr_used
    parsed['images'] = images
    parsed['legends'] = [im for im in images
                         if im.get('ocr_mag') or im.get('scale')
                         or im.get('job') or im.get('id')]
    return rtype, parsed, findings


def summarize(findings):
    """Return counts per severity."""
    out = {'critical': 0, 'warning': 0, 'info': 0, 'pass': 0}
    for sev, _, _ in findings:
        out[sev] = out.get(sev, 0) + 1
    return out


def collect_highlights(parsed):
    """Map the cell-anchored findings to worksheet cells, for the annotated view.

    Returns a list of {'cell': (row, col), 'severity', 'category', 'tag', 'note'}
    on the sheet named in parsed['loc']['sheet']. `tag` is a short label drawn on
    the image; `note` is the full sentence shown in the legend. Findings with no
    single cell to point at (e.g. "no embedded images") are intentionally omitted
    here — they still appear in the textual findings list.
    """
    loc = parsed.get('loc') or {}
    out = []

    def add(cell, severity, category, tag, note):
        if cell:
            out.append({'cell': tuple(cell), 'severity': severity,
                        'category': category, 'tag': tag, 'note': note})

    def anchor(entry):
        """Prefer the value cell; fall back to the label cell."""
        entry = entry or {}
        return entry.get('value') or entry.get('label')

    def first_cell(mapping):
        return next((value for value in (mapping or {}).values() if value), None)

    # ── Workbook integrity — visible errors / external formula references ──
    for error in parsed.get('excel_errors') or []:
        add((error.get('row'), error.get('col')), 'critical', 'Workbook integrity',
            error.get('value') or 'Excel error',
            f'{error.get("sheet")}!{error.get("cell")} contains '
            f'{error.get("value")}.')
    for issue in parsed.get('formula_issues') or []:
        add((issue.get('row'), issue.get('col')), 'critical', 'Workbook integrity',
            'external formula',
            f'{issue.get("sheet")}!{issue.get("cell")} uses external-workbook '
            f'formula {issue.get("formula")}.')

    # ── Report title identity — point at the conflicting internal field ──
    title_header_loc = loc.get('header') or {}
    title_sample_loc = loc.get('sample') or {}
    for severity, category, note in parsed.get('title_findings') or []:
        if severity == 'pass':
            continue
        lowered = note.lower()
        cell, tag = None, 'title mismatch'
        if 'job number' in lowered or 'states job' in lowered:
            cell, tag = anchor(title_header_loc.get('job')), 'job mismatch'
        elif 'machine/set' in lowered or 'machine type' in lowered:
            cell, tag = anchor(title_header_loc.get('machine')), 'machine mismatch'
        elif 'stage' in lowered or 'component' in lowered:
            cell, tag = anchor(title_sample_loc.get('description')), 'component mismatch'
        elif 'customer' in lowered:
            cell, tag = anchor(title_header_loc.get('customer')), 'customer?'
        add(cell, severity, category, tag, note)

    # ── Composition — Actual cells out of tolerance vs Nominal ──
    deviations, systemic = _composition_deviations(
        parsed.get('nominal') or {}, parsed.get('actual') or {})
    nominal_values = parsed.get('nominal') or {}
    actual_values = parsed.get('actual') or {}
    common_count = len(set(nominal_values) & set(actual_values))
    unexpected_major = {
        el: value for el, value in actual_values.items()
        if el not in nominal_values and abs(value) >= 0.5
    }
    gross_conflict = _gross_composition_conflict(
        deviations, common_count, unexpected_major)
    aloc = loc.get('actual') or {}
    if gross_conflict:
        note = _gross_composition_note(
            deviations, common_count, unexpected_major)
        add(first_cell(aloc), 'critical', 'Composition', 'table conflict', note)
    elif systemic:
        worst = sorted(deviations, key=lambda item: -abs(item[3]))[:4]
        detail = ", ".join(f"{el} {rel:+.0f}%" for el, _nom, _act, rel, _sev in worst)
        note = (f'{len(deviations)} of {common_count} comparable elements differ materially '
                f'from the nominal values ({detail}); verify against the governing alloy '
                f'specification and analytical method.')
        add(first_cell(aloc), 'warning', 'Composition', 'multiple deviations', note)
    else:
        for el, nom, act, rel, sev in deviations:
            add(aloc.get(el), sev, 'Composition', f'{el} {rel:+.0f}%',
                f'{el}: actual {act:g} vs nominal {nom:g} wt% '
                f'({rel:+.0f}%) — verify.')
    comp_meta = (parsed.get('composition_meta') or {}).get('actual') or {}
    duplicate_headers = set(comp_meta.get('duplicate_headers') or [])
    for entry in comp_meta.get('entries') or []:
        el, raw = entry.get('element'), entry.get('raw') or ''
        cell = (entry.get('row'), entry.get('col'))
        if el in duplicate_headers:
            add(cell, 'critical', 'Composition', f'duplicate {el}',
                f'Actual composition table repeats element header {el}.')
        if entry.get('value') is None and _NOT_QUANTIFIED.match(raw):
            nom = (parsed.get('nominal') or {}).get(el)
            sev = 'critical' if nom is not None and nom >= 0.5 else 'warning'
            note = (f'{el}: actual result is "{raw}"'
                    + (f' while nominal is {nom:g} wt%.' if nom is not None else '.'))
            add(cell, sev, 'Composition', f'{el} {raw}', note)

    # ── Hardness — post-solution should not exceed pre-solution ──
    hd = parsed.get('hardness') or {}
    pre = (hd.get('pre') or {}).get('value')
    post = (hd.get('post') or {}).get('value')
    hloc = loc.get('hardness') or {}
    if pre is not None and post is not None and post > pre + 2:
        note = (f'Post-solution hardness ({post:g} HRC) exceeds pre-solution '
                f'({pre:g} HRC) — solution treatment normally softens the material.')
        for key in ('pre', 'post'):
            add(anchor(hloc.get(key)), 'warning', 'Hardness', 'post > pre', note)

    # ── Completeness — blank header fields / material ──
    hdr = parsed.get('header') or {}
    hdr_loc = loc.get('header') or {}
    for key, label in (('customer', 'Customer'), ('job', 'AEG Job No'),
                       ('machine', 'Machine type')):
        if _is_placeholder(hdr.get(key)):
            add(anchor(hdr_loc.get(key)), 'warning', 'Completeness',
                f'{label} blank', f'{label} is blank or a placeholder.')
    smp = parsed.get('sample') or {}
    if _is_placeholder(smp.get('material')):
        add(anchor((loc.get('sample') or {}).get('material')), 'warning',
            'Completeness', 'Material blank', 'Sample material/alloy not stated.')

    for key, label in (('customer_ref', 'Customer Ref No'), ('qty', 'Quantity per set')):
        if _is_placeholder(hdr.get(key)):
            add(anchor(hdr_loc.get(key)), 'warning', 'Completeness',
                f'{label} blank', f'{label} not provided.')

    # ── Release basis — criteria, method evidence and representative sampling ──
    governance = _review_acceptance_and_methods(parsed)
    for severity, category, note in governance:
        if category == 'Acceptance criteria':
            add(anchor(loc.get('comment')), severity, category, 'no criteria', note)
        elif category == 'Chemistry method':
            for cell in (first_cell(loc.get('nominal')), first_cell(loc.get('actual'))):
                add(cell, severity, category, 'method missing', note)
        elif category == 'Hardness evidence':
            for key in ('pre', 'post'):
                add(anchor((loc.get('hardness') or {}).get(key)), severity,
                    category, 'evidence missing', note)

    for severity, category, note in _review_sampling_basis(parsed):
        add(anchor(hdr_loc.get('qty')), severity, category, 'no sampling plan', note)
        add(anchor((loc.get('sample') or {}).get('sample_no')), severity,
            category, 'no sampling plan', note)

    for severity, category, note in _review_phase_evidence(parsed):
        add(anchor(loc.get('comment')), severity, category, 'basis?', note)

    # ── Traceability — one sample must map to one serial/part identifier ──
    sample_ids = _identifier_tokens(smp.get('sample_no'), 'sample')
    serial_ids = _identifier_tokens(smp.get('serial'), 'serial')
    if sample_ids and serial_ids and len(sample_ids) != len(serial_ids):
        note = (f'Sample and serial/part identifiers do not map one-to-one: '
                f'{len(sample_ids)} sample identifier(s) versus '
                f'{len(serial_ids)} serial/part identifier(s).')
        for key in ('sample_no', 'serial'):
            add(anchor((loc.get('sample') or {}).get(key)), 'warning',
                'Traceability', 'count mismatch', note)
    for label, values, key in (
            ('sample number', sample_ids, 'sample_no'),
            ('serial/part number', serial_ids, 'serial')):
        duplicates = sorted(value for value, count in Counter(values).items() if count > 1)
        if duplicates:
            note = f'Duplicate {label}(s): {", ".join(duplicates)}.'
            add(anchor((loc.get('sample') or {}).get(key)), 'critical',
                'Traceability', 'duplicate ID', note)

    # ── Disposition — "See comment" must lead to an actual verdict ──
    result = smp.get('result') or ''
    comment = parsed.get('comment') or ''
    if 'see comment' in result.lower():
        has_negative = bool(re.search(
            r'not\s+suitable|unsuitable|not\s+recommend|\breject|\bscrap|'
            r'beyond\s+repair|non[-\s]?conform|unacceptable', comment, re.I))
        has_positive = bool(re.search(
            r'(?<!not )(?:\bsuitable for|\bacceptable|recommended for|'
            r'reconditi|fit for service|return to service)', comment, re.I))
        if not has_negative and not has_positive:
            note = ('Result says "See comment", but the comment gives no clear '
                    'accept / reject / repair disposition.')
            sloc = loc.get('sample') or {}
            add(anchor(sloc.get('result')), 'critical', 'Disposition', 'no verdict', note)
            add(anchor(loc.get('comment')), 'critical', 'Disposition', 'no verdict', note)

    # ── Completeness — missing or very short comment ──
    if len((parsed.get('comment') or '').strip()) < 40:
        add(anchor(loc.get('comment')), 'warning', 'Completeness',
            'Short comment', 'Comment / discussion is missing or very short.')

    # ── Sign-off — missing fields (point at the label) ──
    # One combined note (matching the textual finding) shared across the cells,
    # so the box and the findings list don't double-report it.
    so = parsed.get('signoff') or {}
    so_loc = loc.get('signoff') or {}
    so_fields = (('met_lab', 'Met. Lab'), ('mat_eng', 'Mat. Eng'), ('date', 'Date'))
    so_missing = [label for key, label in so_fields if _is_placeholder(so.get(key))]
    if so_missing:
        so_note = f'Missing sign-off field(s): {", ".join(so_missing)}.'
        for key, label in so_fields:
            if _is_placeholder(so.get(key)):
                entry = so_loc.get(key) or {}
                add(entry.get('label') or entry.get('value'), 'critical',
                    'Sign-off', f'{label} missing', so_note)

    # ── Captions — no etch status, or explicitly unetched ──
    pics = parsed.get('pictures') or []
    ploc = loc.get('pictures') or []
    micrographs = parsed.get('micrograph_count')
    if micrographs is not None and micrographs < len(pics):
        note = (f'{len(pics)} picture caption(s) but only {micrographs} embedded '
                f'micrograph(s) — {len(pics) - micrographs} evidence image(s) are missing.')
        for entry in ploc[micrographs:]:
            add(anchor(entry), 'critical', 'Micrographs', 'image missing', note)
    no_etch = [(label or '?').rstrip(':') for label, cap in pics
               if _requires_etch_status(label, cap)
               and not _ETCH_PAT.search(f"{label} {cap or ''}")]
    no_etch_note = f'No etch status in caption(s): {", ".join(no_etch)}.'
    for i, (label, cap) in enumerate(pics):
        text = f"{label} {cap or ''}"
        entry = ploc[i] if i < len(ploc) else {}
        if _requires_etch_status(label, cap) and not _ETCH_PAT.search(text):
            add(anchor(entry), 'warning', 'Captions', 'No etch status', no_etch_note)
        elif _UNETCHED_PAT.search(text):
            add(anchor(entry), 'info', 'Captions', 'Unetched',
                f'{(label or "?").rstrip(":")} caption states unetched / as-polished — '
                f'confirm intended (a microstructure assessment is normally etched).')

    # ── Magnification — stable legend OCR disagrees with paired caption ──
    for verdict in parsed.get('photo_magnification') or []:
        idx = verdict.get('index')
        entry = ploc[idx] if (idx is not None and 0 <= idx < len(ploc)) else {}
        add(anchor(entry), verdict.get('severity', 'warning'), 'Magnification',
            'legend OCR?', verdict['note'])

    # ── Photo etch — per-picture caption↔contrast mismatch (anchor to caption) ──
    for v in parsed.get('photo_etch') or []:
        idx = v.get('index')
        entry = ploc[idx] if (idx is not None and 0 <= idx < len(ploc)) else {}
        add(anchor(entry), v.get('severity', 'warning'), 'Photo etch', 'etch?', v['note'])

    # ── Thickness — calculation must reconcile, not merely fall in the range ──
    for severity, category, note in parsed.get('thickness_findings') or []:
        if severity in ('critical', 'warning'):
            add(anchor(loc.get('comment')), severity, category, 'calculation?', note)

    # ── Coating — thickness measurements outside design limits ──
    for entry in parsed.get('rows') or []:
        lo, hi = entry.get('min'), entry.get('max')
        if lo is None or hi is None:
            continue
        for col, v in entry.get('cells', []):
            if not (lo <= v <= hi):
                add((entry['row'], col), 'critical', 'Coating', f'{v:g} mm',
                    f'Row {entry["row"]}: thickness {v:g} mm outside design '
                    f'limit {lo:g}–{hi:g} mm.')

    return out


# ── CLI ───────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    for path in sys.argv[1:]:
        with open(path, 'rb') as f:
            data = f.read()
        rtype, parsed, findings = review_report(path, data)
        counts = summarize(findings)
        print('=' * 78)
        print(f'{path}')
        print(f'  type: {rtype}   '
              f'critical={counts["critical"]} warning={counts["warning"]} '
              f'info={counts["info"]} pass={counts["pass"]}')
        for sev, cat, msg in findings:
            tag = {'critical': 'FAIL', 'warning': 'WARN', 'info': 'INFO', 'pass': 'OK  '}[sev]
            print(f'   [{tag}] {cat}: {msg}')
        for lg in parsed.get('legends', []):
            bits = [lg[k] for k in ('id', 'mag', 'scale') if lg.get(k)]
            print(f'     · {lg["image"]}: {"  ".join(bits)}')


if __name__ == '__main__':
    main()
