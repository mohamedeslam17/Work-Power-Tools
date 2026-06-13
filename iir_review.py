#!/usr/bin/env python3
"""
IIR Review Tool - Ansaldo Energia
Automated quality/consistency review of Incoming Inspection Reports (IIR /
"Detailed Assessment Customer Report") delivered as Excel workbooks.

It parses one report, runs a battery of consistency + completeness checks and
emits a severity-tagged findings checklist (.xlsx).

Usage: python3 iir_review.py report.xlsx [findings.xlsx]
"""
import sys, os, re, zipfile
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── severity model ─────────────────────────────────────────────────────────
FAIL, WARN, INFO, PASS = "FAIL", "WARN", "INFO", "PASS"
SEV_RANK  = {FAIL: 0, WARN: 1, INFO: 2, PASS: 3}
SEV_FILL  = {FAIL: "FFC7CE", WARN: "FFEB9C", INFO: "D9E1F2", PASS: "C6EFCE"}
SEV_FONT  = {FAIL: "9C0006", WARN: "9C6500", INFO: "1F4E78", PASS: "006100"}
SEV_ICON  = {FAIL: "🔴", WARN: "🟠", INFO: "🔵", PASS: "🟢"}

# ── low-level helpers ──────────────────────────────────────────────────────
def _norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''

def _num(v):
    """Best-effort numeric extraction from a cell value (int/float/str)."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        m = re.search(r'-?\d+(?:\.\d+)?', v.replace(',', ''))
        if m:
            g = m.group()
            return float(g) if '.' in g else int(g)
    return None

def _canon(s):
    """Canonicalise a defect/finding label for fuzzy comparison: lower-case,
    drop punctuation, singularise each word."""
    s = re.sub(r'[^a-z0-9 ]+', ' ', str(s).lower())
    s = re.sub(r'\s+', ' ', s).strip()
    return ' '.join(re.sub(r's$', '', w) for w in s.split())

def _cells(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                yield c

def _find_label(ws, pattern, flags=re.I):
    rx = re.compile(pattern, flags)
    for c in _cells(ws):
        if isinstance(c.value, str) and rx.search(c.value):
            return c
    return None

def _value_right_of(ws, cell, max_gap=14):
    """First non-empty cell value to the right of `cell` in the same row."""
    if cell is None:
        return None
    for col in range(cell.column + 1, cell.column + 1 + max_gap):
        v = ws.cell(cell.row, col).value
        if _norm(v):
            return _norm(v)
    return None

def _sheet_by(wb, *needles):
    """Return the first sheet whose name contains any needle (case-insensitive)."""
    for name in wb.sheetnames:
        low = name.lower()
        if any(n.lower() in low for n in needles):
            return wb[name]
    return None

# ── drawing / image inspection (openpyxl drops images without Pillow, so we
#    read the package directly) ──────────────────────────────────────────────
def _image_anchors_per_sheet(path, sheetnames):
    """Map sheet name -> number of embedded <xdr:pic> drawings.

    Walks workbook.xml + sheet rels + drawing rels straight from the .xlsx zip,
    so it does not depend on Pillow being installed.
    """
    result = {n: 0 for n in sheetnames}
    try:
        z = zipfile.ZipFile(path)
        names = set(z.namelist())

        # sheetId/name (workbook order) -> sheetN.xml file via rels
        wb_xml = z.read('xl/workbook.xml').decode('utf-8', 'ignore')
        rels   = z.read('xl/_rels/workbook.xml.rels').decode('utf-8', 'ignore')
        rid_to_target = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
        ordered = []  # (name, sheetfile)
        for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wb_xml):
            name, rid = m.group(1), m.group(2)
            tgt = rid_to_target.get(rid, '')
            sf = 'xl/' + tgt.lstrip('/').replace('../', '')
            ordered.append((name, sf))

        for name, sf in ordered:
            base = os.path.basename(sf)
            relpath = f'xl/worksheets/_rels/{base}.rels'
            if relpath not in names:
                continue
            srel = z.read(relpath).decode('utf-8', 'ignore')
            dm = re.search(r'Target="([^"]*drawing\d+\.xml)"', srel)
            if not dm:
                continue
            draw = 'xl/drawings/' + os.path.basename(dm.group(1))
            if draw not in names:
                continue
            dxml = z.read(draw).decode('utf-8', 'ignore')
            result[name] = len(re.findall(r'<xdr:pic\b', dxml))
        z.close()
    except Exception:
        pass
    return result

# ════════════════════════════════════════════════════════════════════════════
# PARSE
# ════════════════════════════════════════════════════════════════════════════
def parse_iir(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {'path': str(path), 'file': os.path.basename(str(path)),
            'sheets': list(wb.sheetnames)}

    cover    = _sheet_by(wb, 'Cover')
    contents = _sheet_by(wb, 'Contents')
    es       = _sheet_by(wb, 'Executive', 'Summary of Re')  # exec summary
    recv     = _sheet_by(wb, 'Received')
    damages  = _sheet_by(wb, 'Damages')

    # ── identity: prefer the Contents header block (clean label/value cells) ──
    ident = {}
    if contents is not None:
        for key, label in [('customer', r'Customer'), ('component', r'Component'),
                           ('author', r'Author'), ('doc_no', r'Doc\.?\s*No')]:
            ident[key] = _value_right_of(contents, _find_label(contents, label))

    # ── cover holds PO#, reviewer, approver, title, date as combined strings ──
    if cover is not None:
        joined = [(_norm(c.value)) for c in _cells(cover) if isinstance(c.value, str)]
        for line in joined:
            low = line.lower()
            if low.startswith('po#') or low.startswith('po #'):
                ident['po'] = _norm(re.sub(r'(?i)^po\s*#\s*:?', '', line))
            elif 'reviewed by' in low:
                ident['reviewer'] = _norm(re.sub(r'(?i)^reviewed by\s*:?', '', line))
            elif 'approved by' in low:
                ident['approver'] = _norm(re.sub(r'(?i)^approved by\s*:?', '', line))
            elif low.startswith('doc. no') or low.startswith('doc no'):
                ident.setdefault('doc_no_cover', _norm(re.sub(r'(?i)^doc\.?\s*no\.?\s*', '', line)))
            elif ' / ' in line and re.search(r'(?i)(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', line) \
                    and 'ansaldo' not in low and not low.startswith('doc'):
                parts = line.split('/')
                ident.setdefault('preparer', _norm(parts[0]))
                ident.setdefault('prep_date', _norm(parts[-1]))
            elif ('g0' in low or '(g' in low) and ('gt' in low or '–' in line or '-' in line) \
                    and 'reviewed' not in low and 'approved' not in low and 'po#' not in low:
                ident.setdefault('title', line)
    data['ident'] = ident

    # ── executive-summary narrative ──────────────────────────────────────────
    es_text = ''
    if es is not None:
        body = _find_label(es, r'incoming\s+(assessment|inspection)')
        if body is None:
            # longest string cell on the sheet is the narrative
            strs = [(_norm(c.value), c) for c in _cells(es) if isinstance(c.value, str)]
            if strs:
                es_text = max(strs, key=lambda t: len(t[0]))[0]
        else:
            es_text = _norm(body.value)
    data['exec_text'] = es_text
    # received count claimed in the narrative ("a total of 48 ...")
    m = re.search(r'total of\s+(\d+)', es_text, re.I)
    data['exec_received'] = int(m.group(1)) if m else None
    # scrap positions named in the narrative
    pos = []
    for grp in re.findall(r'POS\s*#?\s*([\d ,and&]+)', es_text, re.I):
        pos += [int(x) for x in re.findall(r'\d+', grp)]
    data['exec_scrap_pos'] = sorted(set(pos))

    # ── received-parts table (sum across all part-number rows) ───────────────
    rp = {'rows': 0, 'required': 0, 'received': 0, 'scrap': 0, 'reconditionable': 0,
          'found': False}
    if recv is not None:
        hr = None
        cols = {}
        for r in range(1, 20):
            rowmap = {_norm(c.value).lower(): c.column for c in recv[r] if c.value is not None}
            if any(k == 'received' for k in rowmap) and any('scrap' == k for k in rowmap):
                hr, cols = r, rowmap
                break
        if hr:
            def cidx(name):
                for k, v in cols.items():
                    if name in k:
                        return v
                return None
            idx = {k: cidx(k) for k in ['required', 'received', 'scrap', 'reconditionable']}
            for r in range(hr + 1, hr + 30):
                first = recv.cell(r, 1).value
                if isinstance(first, str) and first.strip().lower().startswith('table'):
                    break
                rec = _num(recv.cell(r, idx['received']).value) if idx['received'] else None
                if rec is None:
                    continue
                rp['found'] = True
                rp['rows'] += 1
                for k in ('required', 'received', 'scrap', 'reconditionable'):
                    val = _num(recv.cell(r, idx[k]).value) if idx[k] else None
                    if val:
                        rp[k] += val
    data['received_parts'] = rp

    # ── serial-number protocol (all "Serial Number*" sheets) ─────────────────
    sn_rows, sumrow = [], {}
    for name in wb.sheetnames:
        if not name.lower().startswith('serial number'):
            continue
        ws = wb[name]
        hr = next((r for r in range(1, 14)
                   if any(isinstance(c.value, str) and 'position' in c.value.lower()
                          for c in ws[r])), None)
        colmap = {}
        if hr:
            colmap = {_norm(c.value).lower(): c.column for c in ws[hr] if c.value is not None}
        def col(needle):
            for k, v in colmap.items():
                if needle in k:
                    return v
            return None
        c_pos = col('position'); c_pn = col('part number'); c_sn = col('serial')
        c_scope = col('repair scope') or col('scope'); c_scrap = col('scrap')
        defect_cols = {k: v for k, v in colmap.items()
                       if k not in ('position no', 'part number', 'serial number',
                                    'repair scope', 'repair scope5', 'scrap', 'adder')
                       and v not in (c_pos, c_pn, c_sn, c_scope, c_scrap)}
        if hr:
            for r in range(hr + 1, ws.max_row + 1):
                pv = ws.cell(r, c_pos).value if c_pos else None
                if not isinstance(pv, (int, float)):
                    continue
                scope = _norm(ws.cell(r, c_scope).value) if c_scope else ''
                scrap = bool(_norm(ws.cell(r, c_scrap).value)) if c_scrap else False
                defects = [name_ for name_, cc in defect_cols.items()
                           if _norm(ws.cell(r, cc).value)]
                sn_rows.append({
                    'pos': int(pv),
                    'pn': _norm(ws.cell(r, c_pn).value) if c_pn else '',
                    'sn': _norm(ws.cell(r, c_sn).value) if c_sn else '',
                    'scope': scope.upper(), 'scrap': scrap,
                    'defects': defects, 'sheet': name,
                })
        # sum row anywhere on the sheet
        for r in range(1, ws.max_row + 1):
            cells = [(c.column, c.value) for c in ws[r] if c.value is not None]
            for col_i, val in cells:
                if isinstance(val, str) and re.match(r'(?i)(sum\s+\w+|total parts)', val.strip()):
                    for col_j, val_j in cells:
                        if col_j > col_i and isinstance(val_j, (int, float)):
                            sumrow[_norm(val).lower()] = val_j
                            break
    data['sn_rows'] = sn_rows
    data['sn_sumrow'] = sumrow

    # ── findings table (Summary of Damages) ──────────────────────────────────
    findings_tbl = {}
    if damages is not None:
        hdr = _find_label(damages, r'^\s*Finding\s*$')
        if hdr:
            for r in range(hdr.row + 1, hdr.row + 25):
                name_c = damages.cell(r, hdr.column).value
                if isinstance(name_c, str) and name_c.strip() and 'table' not in name_c.lower():
                    cnt = None
                    for c in damages[r]:
                        if isinstance(c.value, (int, float)):
                            cnt = c.value
                    findings_tbl[_norm(name_c)] = cnt
                elif isinstance(name_c, str) and 'table' in name_c.lower():
                    break
    data['findings_tbl'] = findings_tbl

    # ── operating data (executive summary) ───────────────────────────────────
    op = {}
    if es is not None:
        for label in ['EOH', 'Total  EOH', 'Total EOH', 'Total OH', 'OH',
                      'Total Starts', 'Starts', 'Total Trips', 'Trips']:
            c = _find_label(es, rf'^\s*{re.escape(label)}\s*\d*\s*$')
            if c:
                below = es.cell(c.row + 1, c.column).value
                op.setdefault(label.strip(), below)
    data['operating'] = op

    # ── page footers + photos ────────────────────────────────────────────────
    # Footers appear either as "Page X of N" or just "Page X" (template drift).
    footers = []
    for name in wb.sheetnames:
        ws = wb[name]
        for c in _cells(ws):
            if isinstance(c.value, str):
                mm = re.search(r'Page\s+(\d+)(?:\s+of\s+(\d+))?', c.value, re.I)
                if mm:
                    n = int(mm.group(2)) if mm.group(2) else None
                    footers.append((name, int(mm.group(1)), n))
                    break
    data['footers'] = footers

    anchors = _image_anchors_per_sheet(path, wb.sheetnames)
    photos = []
    for name in wb.sheetnames:
        if 'photo' not in name.lower():
            continue
        ws = wb[name]
        caps = [_norm(c.value) for c in _cells(ws)
                if isinstance(c.value, str) and re.match(r'(?i)fig\.?\s*\d+', c.value.strip())]
        photos.append({'sheet': name, 'captions': caps, 'images': anchors.get(name, 0)})
    data['photos'] = photos

    wb.close()
    return data

# ════════════════════════════════════════════════════════════════════════════
# CHECKS
# ════════════════════════════════════════════════════════════════════════════
def _f(check, severity, sheet, detail, category="General"):
    return {'check': check, 'category': category, 'severity': severity,
            'sheet': sheet, 'detail': detail}

def run_checks(d):
    out = []
    ident = d['ident']
    rp = d['received_parts']
    sn = d['sn_rows']
    sumrow = d['sn_sumrow']

    # ── identity / metadata completeness ─────────────────────────────────────
    req_fields = [('doc_no', 'Document number'), ('customer', 'Customer'),
                  ('component', 'Component'), ('author', 'Author'),
                  ('reviewer', 'Reviewer'), ('approver', 'Approver'), ('po', 'PO#')]
    missing = [lbl for k, lbl in req_fields if not _norm(ident.get(k))]
    if missing:
        out.append(_f("Header metadata complete", FAIL, "Cover/Contents",
                      "Missing: " + ", ".join(missing), "Identity"))
    else:
        out.append(_f("Header metadata complete", PASS, "Cover/Contents",
                      f"Doc {ident.get('doc_no')} · {ident.get('customer')} · {ident.get('component')}",
                      "Identity"))

    po = _norm(ident.get('po'))
    if po and re.search(r'(?i)follow|tbd|pending|n/?a', po):
        out.append(_f("PO number assigned", WARN, "Cover",
                      f"PO# is provisional: '{po}'", "Identity"))

    doc = _norm(ident.get('doc_no'))
    if doc and not re.fullmatch(r'\d+\s*Rev\s*-?\s*\d+', doc, re.I):
        out.append(_f("Doc-number format", WARN, "Contents",
                      f"Unexpected doc-number format: '{doc}' (expected 'NNNN Rev-0')", "Identity"))
    doc_cover = _norm(ident.get('doc_no_cover'))
    if doc and doc_cover:
        if re.sub(r'\s+', '', doc.lower()) != re.sub(r'\s+', '', doc_cover.lower()):
            out.append(_f("Doc-number cover vs contents", WARN, "Cover/Contents",
                          f"Cover '{doc_cover}' != Contents '{doc}'", "Identity"))

    prep = _norm(ident.get('preparer'))
    if prep and len(prep.split()) < 2:
        out.append(_f("Preparer name complete", WARN, "Cover",
                      f"Preparer appears to be a first name only: '{prep}'", "Identity"))

    for k, lbl in [('reviewer', 'Reviewer'), ('approver', 'Approver')]:
        v = ident.get(k)
        if isinstance(v, str) and v != v.strip():
            out.append(_f(f"{lbl} field tidy", INFO, "Cover",
                          f"{lbl} value has stray whitespace: '{v}'", "Identity"))

    # ── received-parts internal reconciliation ───────────────────────────────
    if rp['found']:
        recv, scrap, recon, reqd = rp['received'], rp['scrap'], rp['reconditionable'], rp['required']
        if scrap + recon == recv:
            out.append(_f("Received = Scrap + Reconditionable", PASS, "Summary of Received Parts",
                          f"{recv} received = {scrap} scrap + {recon} reconditionable", "Quantities"))
        else:
            out.append(_f("Received = Scrap + Reconditionable", FAIL, "Summary of Received Parts",
                          f"{scrap} scrap + {recon} reconditionable = {scrap + recon} ≠ {recv} received "
                          f"(off by {recv - scrap - recon})", "Quantities"))
        if reqd and recv and recv > reqd:
            out.append(_f("Received ≤ Required", WARN, "Summary of Received Parts",
                          f"Received {recv} exceeds Required {reqd}", "Quantities"))
        # Required column sanity (catches stray strings like '333' for 33)
        if reqd and recv and reqd > 3 * recv and recv > 0:
            out.append(_f("Required quantity plausible", WARN, "Summary of Received Parts",
                          f"Required total ({reqd}) is far larger than Received ({recv}) — "
                          f"check the 'Required' column for data-entry errors", "Quantities"))
    else:
        out.append(_f("Received-parts table found", FAIL, "Summary of Received Parts",
                      "Could not locate the Required/Received/Scrap/Reconditionable table", "Quantities"))

    # ── serial-number protocol ───────────────────────────────────────────────
    positions = [r['pos'] for r in sn]
    npos = len(positions)
    if rp['found'] and npos:
        if npos == rp['received']:
            out.append(_f("Positions listed = Received", PASS, "Serial Number",
                          f"{npos} positions listed = {rp['received']} received", "Quantities"))
        else:
            out.append(_f("Positions listed = Received", FAIL, "Serial Number",
                          f"{npos} positions in protocol ≠ {rp['received']} received", "Quantities"))

    if positions:
        dupes = sorted({p for p in positions if positions.count(p) > 1})
        if dupes:
            out.append(_f("Position numbers unique", FAIL, "Serial Number",
                          f"Duplicate position numbers: {dupes}", "Integrity"))
        expected = set(range(1, max(positions) + 1))
        gaps = sorted(expected - set(positions))
        if gaps:
            out.append(_f("Position numbering contiguous", WARN, "Serial Number",
                          f"Missing position numbers: {gaps}", "Integrity"))
        # serial numbers present for non-scrap rows
        no_sn = [r['pos'] for r in sn if not r['sn']]
        if no_sn:
            out.append(_f("Serial number per position", WARN, "Serial Number",
                          f"No serial number for position(s): {no_sn}", "Integrity"))
        dup_sn = {}
        for r in sn:
            if r['sn']:
                dup_sn.setdefault(r['sn'], []).append(r['pos'])
        dups = {k: v for k, v in dup_sn.items() if len(v) > 1}
        if dups:
            out.append(_f("Serial numbers unique", WARN, "Serial Number",
                          "Repeated serial numbers: " +
                          "; ".join(f"{k}→pos {v}" for k, v in dups.items()), "Integrity"))
        # every row has a repair scope or is scrap
        no_scope = [r['pos'] for r in sn if not r['scope'] and not r['scrap']]
        if no_scope:
            out.append(_f("Repair scope per position", WARN, "Serial Number",
                          f"No repair scope and not scrapped: position(s) {no_scope}", "Integrity"))
        else:
            out.append(_f("Repair scope per position", PASS, "Serial Number",
                          "Every position has a repair scope or scrap mark", "Integrity"))
        # repair-scope values must be one of L / M / H / S
        bad_vals = sorted({r['scope'] for r in sn
                           if r['scope'] and r['scope'] not in ('L', 'M', 'H', 'S')})
        if bad_vals:
            out.append(_f("Repair-scope values valid", WARN, "Serial Number",
                          f"Unexpected repair-scope value(s): {bad_vals} (expected L/M/H/S)", "Integrity"))
        # scrap marks and scope 'S' must agree both ways
        scrap_not_s = [r['pos'] for r in sn if r['scrap'] and r['scope'] not in ('S', '')]
        s_not_scrap = [r['pos'] for r in sn if r['scope'] == 'S' and not r['scrap']]
        if scrap_not_s:
            out.append(_f("Scrap mark ↔ scope 'S'", WARN, "Serial Number",
                          f"Scrap-marked but scope is not 'S': position(s) {scrap_not_s}", "Integrity"))
        if s_not_scrap:
            out.append(_f("Scrap mark ↔ scope 'S'", WARN, "Serial Number",
                          f"Scope 'S' but not scrap-marked: position(s) {s_not_scrap}", "Integrity"))

    # ── serial-number sum row reconciliation ─────────────────────────────────
    if sumrow:
        scrap = sumrow.get('sum scrap', 0); light = sumrow.get('sum light', 0)
        med = sumrow.get('sum medium', 0);  heavy = sumrow.get('sum heavy', 0)
        total = sumrow.get('total parts received')
        ssum = scrap + light + med + heavy
        if total is not None:
            if ssum == total:
                out.append(_f("Scope totals = Total received", PASS, "Serial Number",
                              f"Scrap {scrap} + L {light} + M {med} + H {heavy} = {total}", "Quantities"))
            else:
                out.append(_f("Scope totals = Total received", FAIL, "Serial Number",
                              f"Scrap+L+M+H = {ssum} ≠ Total {total}", "Quantities"))
            if npos and total != npos:
                out.append(_f("Sum-row total = positions listed", WARN, "Serial Number",
                              f"Sum-row total {total} ≠ {npos} positions listed", "Quantities"))
        # cross-check reconditionable count vs received-parts table
        if rp['found']:
            recond_sn = light + med + heavy
            if recond_sn and recond_sn != rp['reconditionable']:
                out.append(_f("Reconditionable: table vs protocol", FAIL,
                              "Received Parts / Serial Number",
                              f"Received-Parts table says {rp['reconditionable']} reconditionable, "
                              f"but protocol sums {recond_sn} (L {light}+M {med}+H {heavy})", "Quantities"))
            if sumrow.get('sum scrap') is not None and sumrow['sum scrap'] != rp['scrap']:
                out.append(_f("Scrap count: table vs protocol", FAIL,
                              "Received Parts / Serial Number",
                              f"Received-Parts scrap {rp['scrap']} ≠ protocol scrap {sumrow['sum scrap']}",
                              "Quantities"))
    elif sn:
        out.append(_f("Serial-number summary totals present", WARN, "Serial Number",
                      "No 'Sum Scrap/Light/Medium/Heavy / Total Parts Received' row found "
                      "(present in the standard template)", "Completeness"))

    # ── stated sum row vs the scopes actually marked on each row ─────────────
    if sumrow and sn:
        scope_ct = {c: sum(1 for r in sn if r['scope'] == c) for c in ('S', 'L', 'M', 'H')}
        diffs = []
        for word, key, code in [('Scrap', 'sum scrap', 'S'), ('Light', 'sum light', 'L'),
                                 ('Medium', 'sum medium', 'M'), ('Heavy', 'sum heavy', 'H')]:
            if key in sumrow and sumrow[key] != scope_ct[code]:
                diffs.append(f"{word} row = {sumrow[key]} but {scope_ct[code]} part(s) scoped '{code}'")
        if diffs:
            out.append(_f("Sum row = scopes marked", FAIL, "Serial Number",
                          "; ".join(diffs) + " — the totals row may be stale", "Quantities"))
        else:
            out.append(_f("Sum row = scopes marked", PASS, "Serial Number",
                          f"Totals row agrees with marked scopes "
                          f"(S {scope_ct['S']} / L {scope_ct['L']} / M {scope_ct['M']} / H {scope_ct['H']})",
                          "Quantities"))

    # ── Summary-of-Damages finding counts vs defect marks in the protocol ────
    if d['findings_tbl'] and sn:
        tally = {}
        for r in sn:
            for dd in r['defects']:
                tally[_canon(dd)] = tally.get(_canon(dd), 0) + 1
        compared, mism = 0, []
        for name, cnt in d['findings_tbl'].items():
            if not isinstance(cnt, (int, float)):
                continue
            key = _canon(name)
            if key in tally:
                compared += 1
                if tally[key] != cnt:
                    mism.append((name, int(cnt), tally[key]))
        for name, cnt, act in mism:
            out.append(_f("Finding count vs protocol", WARN, "Summary of Damages / Serial Number",
                          f"'{name}' summary shows {cnt} but {act} part(s) marked in the protocol "
                          f"(off by {cnt - act:+d})", "Consistency"))
        if compared and not mism:
            out.append(_f("Finding counts vs protocol", PASS, "Summary of Damages",
                          f"{compared} finding categor{'y' if compared == 1 else 'ies'} reconcile "
                          f"with the protocol defect marks", "Consistency"))

    # ── executive summary cross-checks ───────────────────────────────────────
    if d['exec_received'] is not None and rp['found'] and d['exec_received'] != rp['received']:
        out.append(_f("Exec-summary received count", WARN, "Executive Summary",
                      f"Narrative says 'total of {d['exec_received']}' but table received = {rp['received']}",
                      "Consistency"))
    if d['exec_scrap_pos'] and sn:
        sn_scrap = {r['pos'] for r in sn if r['scrap']}
        narrative = set(d['exec_scrap_pos'])
        only_text = sorted(narrative - sn_scrap)   # claimed scrap that isn't marked
        only_tbl = sorted(sn_scrap - narrative)    # marked scrap not enumerated
        if only_text:
            out.append(_f("Scrap positions named are marked", WARN, "Executive Summary",
                          f"Position(s) named as scrap in the summary but not scrap-marked "
                          f"in the protocol: {only_text}", "Consistency"))
        if only_tbl:
            out.append(_f("Scrap positions enumerated", INFO, "Executive Summary",
                          f"Scrap-marked in the protocol but not enumerated in the summary: "
                          f"{only_tbl} (often qualification/sample parts)", "Consistency"))

    # ── findings table sanity ────────────────────────────────────────────────
    if d['findings_tbl'] and rp['found'] and rp['received']:
        over = {k: v for k, v in d['findings_tbl'].items()
                if isinstance(v, (int, float)) and v > rp['received']}
        if over:
            out.append(_f("Finding counts ≤ received", WARN, "Summary of Damages",
                          "Counts exceed received qty: " +
                          ", ".join(f"{k}={v}" for k, v in over.items()), "Consistency"))

    # ── page numbering ───────────────────────────────────────────────────────
    footers = d['footers']
    if footers:
        xs = [x for _, x, _ in footers]
        ns = {n for _, _, n in footers if n is not None}
        nsheets = len(d['sheets'])
        page_ok = True
        # page-number sequence should be contiguous & unique
        dupes = sorted({x for x in xs if xs.count(x) > 1})
        if dupes:
            page_ok = False
            out.append(_f("Page numbers unique", WARN, "All pages",
                          f"Repeated footer page numbers: {dupes}", "Completeness"))
        gaps = sorted(set(range(min(xs), max(xs) + 1)) - set(xs))
        if gaps:
            page_ok = False
            out.append(_f("Page sequence contiguous", WARN, "All pages",
                          f"Missing footer page numbers: {gaps}", "Completeness"))
        if len(ns) > 1:
            page_ok = False
            out.append(_f("Page 'of N' consistent", WARN, "All pages",
                          f"Footer total page count differs across sheets: {sorted(ns)}", "Completeness"))
        elif not ns:
            page_ok = False
            out.append(_f("Page footers show total", INFO, "All pages",
                          "Footers use 'Page X' without an 'of N' total (template drift)", "Completeness"))
        elif next(iter(ns)) != nsheets:
            page_ok = False
            out.append(_f("Page total vs sheet count", INFO, "All pages",
                          f"Footers say 'of {next(iter(ns))}' but workbook has {nsheets} sheets "
                          f"(a page total may not have been updated)", "Completeness"))
        if page_ok:
            out.append(_f("Page numbering consistent", PASS, "All pages",
                          f"All footers agree: {nsheets} pages", "Completeness"))

    # ── photos ───────────────────────────────────────────────────────────────
    for ph in d['photos']:
        if ph['captions'] and ph['images'] == 0:
            out.append(_f("Photos embedded", FAIL, ph['sheet'],
                          f"{len(ph['captions'])} captions but no embedded images found", "Completeness"))
        elif ph['captions'] and ph['images'] < len(ph['captions']):
            out.append(_f("Photo per caption", WARN, ph['sheet'],
                          f"{len(ph['captions'])} captions but only {ph['images']} images", "Completeness"))
    if d['photos'] and not any(f['severity'] in (FAIL, WARN) and 'photo' in f['sheet'].lower()
                               for f in out):
        total_imgs = sum(p['images'] for p in d['photos'])
        out.append(_f("Incoming photos present", PASS, "Incoming photos",
                      f"{len(d['photos'])} photo sheet(s), {total_imgs} images embedded", "Completeness"))

    # stable sort: severity first, then category
    out.sort(key=lambda x: (SEV_RANK[x['severity']], x['category']))
    return out

def count_severities(findings):
    return {s: sum(1 for f in findings if f['severity'] == s) for s in (FAIL, WARN, INFO, PASS)}

def verdict_of(counts):
    """Return (severity, label) summarising a set of findings."""
    if counts[FAIL]:
        return FAIL, "FAIL — issues require correction"
    if counts[WARN]:
        return WARN, "REVIEW — warnings to confirm"
    return PASS, "PASS — no blocking issues"

def top_issue(findings):
    """The single most important finding's detail (first FAIL, else WARN)."""
    for sev in (FAIL, WARN):
        for f in findings:
            if f['severity'] == sev:
                return f"[{sev}] {f['check']}: {f['detail']}"
    return "—"

# ════════════════════════════════════════════════════════════════════════════
# EXCEL CHECKLIST
# ════════════════════════════════════════════════════════════════════════════
def _set_border(cell, color="BFBFBF"):
    side = Side(style="thin", color=color)
    cell.border = Border(left=side, right=side, top=side, bottom=side)

def build_checklist(data, findings, out_path):
    wb = openpyxl.Workbook()
    NAVY = "1A1A2E"
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    label_font = Font(bold=True, name="Calibri", size=10)

    counts = count_severities(findings)
    _, verdict = verdict_of(counts)

    # ── Sheet 1: Review Summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Review Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 70

    ws['A1'] = "IIR Quality Review"
    ws['A1'].font = Font(bold=True, size=15, color=NAVY, name="Calibri")
    ws['A2'] = data['file']
    ws['A2'].font = Font(italic=True, color="555555", name="Calibri")

    ident = data['ident']
    rp = data['received_parts']
    rows = [
        ("Document No.", ident.get('doc_no')),
        ("Customer", ident.get('customer')),
        ("Component", ident.get('component')),
        ("Prepared by", ident.get('preparer') or ident.get('author')),
        ("Reviewed by", ident.get('reviewer')),
        ("Approved by", ident.get('approver')),
        ("PO #", ident.get('po')),
        ("Received / Scrap / Recond.",
         f"{rp.get('received')} / {rp.get('scrap')} / {rp.get('reconditionable')}" if rp.get('found') else "—"),
        ("Positions in protocol", len(data['sn_rows'])),
        ("", ""),
        ("Overall verdict", verdict),
        ("Findings", f"{counts[FAIL]} FAIL · {counts[WARN]} WARN · {counts[INFO]} INFO · {counts[PASS]} PASS"),
    ]
    r = 4
    for label, val in rows:
        ws.cell(r, 1, label).font = label_font
        ws.cell(r, 2, "" if val is None else str(val)).font = Font(name="Calibri", size=10)
        if label == "Overall verdict":
            sev = FAIL if counts[FAIL] else (WARN if counts[WARN] else PASS)
            ws.cell(r, 2).fill = PatternFill("solid", fgColor=SEV_FILL[sev])
            ws.cell(r, 2).font = Font(bold=True, color=SEV_FONT[sev], name="Calibri", size=10)
        r += 1

    # ── Sheet 2: Findings checklist ──────────────────────────────────────────
    fs = wb.create_sheet("Findings")
    fs.sheet_view.showGridLines = False
    headers = ["#", "Severity", "Category", "Check", "Sheet", "Detail"]
    widths = [5, 11, 14, 34, 26, 80]
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = fs.cell(1, j, h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        fs.column_dimensions[chr(64 + j)].width = w
    for i, f in enumerate(findings, start=1):
        row = i + 1
        vals = [i, f"{SEV_ICON[f['severity']]} {f['severity']}", f['category'],
                f['check'], f['sheet'], f['detail']]
        for j, v in enumerate(vals, start=1):
            c = fs.cell(row, j, v)
            c.alignment = Alignment(vertical="top", wrap_text=(j == 6),
                                    horizontal="center" if j in (1, 2) else "left")
            c.font = Font(name="Calibri", size=10)
            _set_border(c)
            if j == 2:
                c.fill = PatternFill("solid", fgColor=SEV_FILL[f['severity']])
                c.font = Font(bold=True, color=SEV_FONT[f['severity']], name="Calibri", size=10)
    fs.freeze_panes = "A2"
    fs.auto_filter.ref = f"A1:F{len(findings) + 1}"

    # ── Sheet 3: Extracted Data (traceability) ───────────────────────────────
    ds = wb.create_sheet("Extracted Data")
    ds.sheet_view.showGridLines = False
    ds_headers = ["Pos", "Part Number", "Serial Number", "Scope", "Scrap", "Defects", "Sheet"]
    ds_widths = [6, 22, 18, 8, 8, 40, 18]
    for j, (h, w) in enumerate(zip(ds_headers, ds_widths), start=1):
        c = ds.cell(1, j, h); c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
        ds.column_dimensions[chr(64 + j)].width = w
    for i, rr in enumerate(data['sn_rows'], start=2):
        vals = [rr['pos'], rr['pn'], rr['sn'], rr['scope'], "X" if rr['scrap'] else "",
                ", ".join(rr['defects']), rr['sheet']]
        for j, v in enumerate(vals, start=1):
            c = ds.cell(i, j, v); c.font = Font(name="Calibri", size=9)
            _set_border(c)
            if rr['scrap']:
                c.fill = PatternFill("solid", fgColor="FCE4E4")
    ds.freeze_panes = "A2"

    wb.save(out_path)
    return counts, verdict

# ════════════════════════════════════════════════════════════════════════════
# BATCH SUMMARY  (many reports → one cross-report workbook)
# ════════════════════════════════════════════════════════════════════════════
def record_of(data, findings):
    """Compact per-report record consumed by the batch summary and the app."""
    return {'file': data['file'], 'ident': data['ident'],
            'rp': data['received_parts'], 'npos': len(data['sn_rows']),
            'findings': findings, 'counts': count_severities(findings)}

def build_batch_summary(records, out_path):
    """records: list of record_of(...) dicts. Writes a cross-report workbook
    with a one-row-per-report 'Batch Summary' and a pooled 'All Findings' tab."""
    NAVY = "1A1A2E"
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    wb = openpyxl.Workbook()

    # ── Sheet 1: one row per report ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Batch Summary"
    ws.sheet_view.showGridLines = False
    ws['A1'] = "IIR Batch Review"
    ws['A1'].font = Font(bold=True, size=15, color=NAVY, name="Calibri")
    ws['A2'] = f"{len(records)} report(s) reviewed"
    ws['A2'].font = Font(italic=True, color="555555", name="Calibri")

    headers = ["Report", "Doc No", "Customer", "Component", "Recv", "Scrap", "Recond",
               "Pos", "Fail", "Warn", "Info", "Pass", "Verdict", "Top issue"]
    widths  = [30, 13, 26, 24, 7, 7, 8, 6, 6, 6, 6, 6, 10, 75]
    hrow = 4
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(hrow, j, h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[chr(64 + j)].width = w

    tot = {FAIL: 0, WARN: 0, INFO: 0, PASS: 0}
    for i, rec in enumerate(records):
        counts = rec.get('counts') or count_severities(rec['findings'])
        for k in tot:
            tot[k] += counts[k]
        sev, label = verdict_of(counts)
        ident, rp = rec['ident'], rec['rp']
        gv = (lambda k: rp.get(k) if rp.get('found') else "")
        vals = [rec['file'], ident.get('doc_no', ''), ident.get('customer', ''),
                ident.get('component', ''), gv('received'), gv('scrap'),
                gv('reconditionable'), rec['npos'], counts[FAIL], counts[WARN],
                counts[INFO], counts[PASS], label.split(' — ')[0], top_issue(rec['findings'])]
        r = hrow + 1 + i
        for j, v in enumerate(vals, start=1):
            c = ws.cell(r, j, v); c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="top", wrap_text=(j == 14),
                                    horizontal="center" if 5 <= j <= 13 else "left")
            _set_border(c)
            if j == 13:
                c.fill = PatternFill("solid", fgColor=SEV_FILL[sev])
                c.font = Font(bold=True, color=SEV_FONT[sev], name="Calibri", size=10)
            elif j == 9 and counts[FAIL]:
                c.fill = PatternFill("solid", fgColor=SEV_FILL[FAIL])
            elif j == 10 and counts[WARN]:
                c.fill = PatternFill("solid", fgColor=SEV_FILL[WARN])
    # totals row
    r = hrow + 1 + len(records)
    tc = ws.cell(r, 8, "TOTAL"); tc.font = Font(bold=True, name="Calibri", size=10)
    tc.alignment = Alignment(horizontal="right")
    for j, key in [(9, FAIL), (10, WARN), (11, INFO), (12, PASS)]:
        c = ws.cell(r, j, tot[key]); c.font = Font(bold=True, name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(hrow + 1, 1)
    ws.auto_filter.ref = f"A{hrow}:N{hrow + len(records)}"

    # ── Sheet 2: pooled findings across all reports ──────────────────────────
    fs = wb.create_sheet("All Findings")
    fs.sheet_view.showGridLines = False
    fheaders = ["Report", "Doc No", "Severity", "Category", "Check", "Sheet", "Detail"]
    fwidths  = [28, 12, 11, 14, 32, 24, 80]
    for j, (h, w) in enumerate(zip(fheaders, fwidths), start=1):
        c = fs.cell(1, j, h); c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        fs.column_dimensions[chr(64 + j)].width = w
    pooled = [(rec['file'], rec['ident'].get('doc_no', ''), f)
              for rec in records for f in rec['findings']]
    pooled.sort(key=lambda t: (SEV_RANK[t[2]['severity']], t[0]))
    for i, (fname, doc, f) in enumerate(pooled, start=2):
        vals = [fname, doc, f"{SEV_ICON[f['severity']]} {f['severity']}", f['category'],
                f['check'], f['sheet'], f['detail']]
        for j, v in enumerate(vals, start=1):
            c = fs.cell(i, j, v); c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(vertical="top", wrap_text=(j == 7),
                                    horizontal="center" if j == 3 else "left")
            _set_border(c)
            if j == 3:
                c.fill = PatternFill("solid", fgColor=SEV_FILL[f['severity']])
                c.font = Font(bold=True, color=SEV_FONT[f['severity']], name="Calibri", size=9)
    fs.freeze_panes = "A2"
    if pooled:
        fs.auto_filter.ref = f"A1:G{len(pooled) + 1}"

    wb.save(out_path)
    return tot

# ════════════════════════════════════════════════════════════════════════════
def review(path, out_path=None):
    """Parse + check one IIR workbook and write the checklist. Returns
    (data, findings, out_path)."""
    data = parse_iir(path)
    findings = run_checks(data)
    if out_path is None:
        out_path = f"IIR_Review_{Path(path).stem}.xlsx"
    build_checklist(data, findings, out_path)
    return data, findings, out_path

def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(1)

    # Back-compatible: "report.xlsx out.xlsx" (single input + explicit output).
    explicit_out = None
    if len(args) == 2 and os.path.exists(args[0]) \
            and not os.path.exists(args[1]) and args[1].lower().endswith('.xlsx'):
        inputs, explicit_out = [args[0]], args[1]
    else:
        inputs = [a for a in args if a.lower().endswith('.xlsx') and os.path.exists(a)]
    for a in args:
        if a.lower().endswith('.xlsx') and a not in inputs and a != explicit_out:
            print(f"  (skipped, not found: {a})")
    if not inputs:
        sys.exit("No existing .xlsx input file(s) given.")

    records = []
    for src in inputs:
        out = explicit_out or f"IIR_Review_{Path(src).stem}.xlsx"
        data, findings, out = review(src, out)
        counts = count_severities(findings)
        sev, _ = verdict_of(counts)
        print(f"\n{SEV_ICON[sev]} {data['file']}")
        print(f"   {data['ident'].get('doc_no')} · {data['ident'].get('customer')} · "
              f"{data['ident'].get('component')}")
        for f in findings:
            if f['severity'] in (FAIL, WARN):
                print(f"     {SEV_ICON[f['severity']]} {f['check']}: {f['detail']}")
        print(f"   → {counts[FAIL]} FAIL · {counts[WARN]} WARN · {counts[INFO]} INFO · "
              f"{counts[PASS]} PASS   ({out})")
        records.append(record_of(data, findings))

    if len(records) > 1:
        batch = "IIR_Batch_Summary.xlsx"
        tot = build_batch_summary(records, batch)
        print(f"\n📊 Batch summary of {len(records)} reports → {batch}")
        print(f"   {tot[FAIL]} FAIL · {tot[WARN]} WARN · {tot[INFO]} INFO · {tot[PASS]} PASS")

if __name__ == "__main__":
    main()
