#!/usr/bin/env python3
"""
Annotated review images for the Lab Report Reviewer.

Two products, both pure-Pillow (degrade to None / [] when Pillow is absent):

  * render_report_image(...) — draws the report's data region as a
    spreadsheet-like grid, boxes and numbers the cells that triggered findings,
    and bakes a numbered legend underneath. The "issue areas, highlighted and
    annotated" view.
  * annotate_micrographs(...) — boxes each embedded micrograph's burned-in
    legend / scale-bar regions, flags low contrast and surfaces any thickness
    measurements read from it.

Cell anchoring comes from lab_review.collect_highlights(); this module only
draws. It never raises into the caller — on any trouble it returns None / [].
"""
import io
import os
import re
import shutil
import signal
import subprocess
import tempfile
import textwrap
import threading
import zipfile

# Serialize LibreOffice conversions: concurrent headless soffice on a small host
# collides on its profile and can OOM. One render at a time.
_LO_LOCK = threading.Lock()

try:
    from PIL import Image, ImageChops, ImageDraw, ImageFont
    _PIL = True
except Exception:                       # pragma: no cover - exercised by guard
    _PIL = False

try:
    import fitz                          # PyMuPDF — rasterises the LibreOffice PDF
    _FITZ = True
except Exception:                       # pragma: no cover - exercised by guard
    _FITZ = False

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, PatternFill, Side

from lab_review import collect_highlights, analyze_images

# ── Palette ───────────────────────────────────────────────────────────────
_SEV_RGB = {
    'critical': (214, 45, 56),     # red
    'warning':  (236, 134, 18),    # orange
    'info':     (24, 110, 214),    # blue
    'pass':     (32, 160, 80),     # green
}
_SEV_RANK = {'critical': 3, 'warning': 2, 'info': 1, 'pass': 0}
_WHITE   = (255, 255, 255)
_GRID    = (203, 207, 214)
_HDR_BG  = (240, 242, 246)
_TEXT    = (28, 32, 38)
_MUTED   = (118, 124, 132)
_TITLE_BG = (37, 47, 62)

_FONT_DIRS = ('/usr/share/fonts/truetype/dejavu/', '/usr/share/fonts/dejavu/')


def _font(size, bold=False):
    name = 'DejaVuSans-Bold.ttf' if bold else 'DejaVuSans.ttf'
    for d in _FONT_DIRS:
        try:
            return ImageFont.truetype(d + name, size)
        except Exception:
            continue
    try:
        return ImageFont.truetype(name, size)
    except Exception:
        return ImageFont.load_default()


def _textw(draw, text, font):
    try:
        return draw.textlength(text, font=font)
    except Exception:
        return len(text) * (font.size * 0.6 if hasattr(font, 'size') else 7)


def _fit(draw, text, font, maxw):
    """Truncate `text` with an ellipsis so it fits within maxw pixels."""
    if _textw(draw, text, font) <= maxw:
        return text
    ell = '…'
    while text and _textw(draw, text + ell, font) > maxw:
        text = text[:-1]
    return (text + ell) if text else ''


# ════════════════════════════════════════════════════════════════════════
# THE ANNOTATED REPORT GRID
# ════════════════════════════════════════════════════════════════════════
def render_report_image(data, parsed, findings=None, rtype=None, filename=None,
                        max_rows=48, max_cols=18, scale=2):
    """Return PNG bytes of the annotated report grid, or None.

    `data` is the workbook bytes; `parsed` is lab_review's parsed dict (it must
    carry the 'loc' map). `scale` super-samples the canvas for crisp text.
    """
    if not _PIL:
        return None
    try:
        return _render(data, parsed, findings, filename, max_rows, max_cols, scale)
    except Exception:
        return None


def _grid_layout(ws, highlights, max_rows, max_cols):
    """Column/row pixel geometry for the annotated grid `_render` draws.

    Factored out of `_render` so `cell_pixel_rect` can map a single cell to
    its pixel rect without redrawing anything — used to crop a zoomed detail
    around one finding's cell (the UI's "click a finding, see its cell").
    Returns None if the sheet has no content and no highlight to anchor on.
    """
    spans = {}            # (r, c) of every covered cell → (r0, c0, r1, c1)
    for rng in ws.merged_cells.ranges:
        box = (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                spans[(r, c)] = box

    def cell_text(r, c):
        v = ws.cell(row=r, column=c).value
        return '' if v is None else str(v).strip()

    # Content bounding box (cells with text), unioned with every highlight cell.
    rmin = cmin = 10**9
    rmax = cmax = 0
    seen = False
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200),
                            max_col=min(ws.max_column, 60)):
        for cell in row:
            if cell_text(cell.row, cell.column):
                seen = True
                rmin, cmin = min(rmin, cell.row), min(cmin, cell.column)
                rmax, cmax = max(rmax, cell.row), max(cmax, cell.column)
    for h in highlights:
        r, c = h['cell']
        rmin, cmin = min(rmin, r), min(cmin, c)
        rmax, cmax = max(rmax, r), max(cmax, c)
        seen = True
    if not seen:
        return None
    r1 = min(rmax, rmin + max_rows - 1)
    c1 = min(cmax, cmin + max_cols - 1)
    rows = list(range(rmin, r1 + 1))
    cols = list(range(cmin, c1 + 1))

    # Measure column widths from their content.
    probe = Image.new('RGB', (8, 8))
    pd = ImageDraw.Draw(probe)
    f_cell = _font(13)
    f_hdr  = _font(12, bold=True)
    PAD = 8
    cw = {}
    for c in cols:
        w = _textw(pd, get_column_letter(c), f_hdr) + 10
        for r in rows:
            if spans.get((r, c), (r, c, r, c))[:2] != (r, c) and (r, c) in spans:
                continue                       # measured at the anchor instead
            w = max(w, _textw(pd, cell_text(r, c), f_cell) + 2 * PAD + 6)
        cw[c] = int(max(52, min(240, w)))
    RH = 30
    GUT = 46                                   # row-number gutter
    TH = 24                                    # column-letter header band
    TITLE_H = 58

    # x/y geometry.
    col_x = {}
    x = GUT
    for c in cols:
        col_x[c] = x
        x += cw[c]
    grid_right = x
    row_y = {}
    y = TITLE_H + TH
    for r in rows:
        row_y[r] = y
        y += RH
    grid_bottom = y

    def x_right(c):
        return col_x[c] + cw[c] if c in col_x else grid_right

    def rect_for(r, c):
        """Pixel rect (x0, y0, x1, y1) for a cell, honouring merges + clipping."""
        r0, c0, r2, c2 = spans.get((r, c), (r, c, r, c))
        r0, r2 = max(r0, rmin), min(r2, r1)
        c0, c2 = max(c0, cmin), min(c2, c1)
        x0 = col_x.get(c0, GUT)
        x1 = x_right(c2)
        y0 = row_y.get(r0, TITLE_H + TH)
        y1 = row_y.get(r2, grid_bottom - RH) + RH
        return x0, y0, x1, y1

    return {
        'rmin': rmin, 'cmin': cmin, 'r1': r1, 'c1': c1, 'spans': spans,
        'rows': rows, 'cols': cols, 'cell_text': cell_text, 'rect_for': rect_for,
        'col_x': col_x, 'cw': cw, 'row_y': row_y, 'RH': RH, 'GUT': GUT, 'TH': TH,
        'TITLE_H': TITLE_H, 'grid_right': grid_right, 'grid_bottom': grid_bottom,
    }


def cell_pixel_rect(data, parsed, cell, max_rows=48, max_cols=18, scale=2):
    """Pixel rect (x0, y0, x1, y1) for one spreadsheet `cell` (row, col), in
    the same coordinate space as `render_report_image`'s PNG when called with
    matching max_rows/max_cols/scale (the defaults match). None if Pillow is
    unavailable, the workbook can't be read, or the cell falls outside the
    rendered window (e.g. beyond max_rows/max_cols from the content origin).
    """
    if not _PIL:
        return None
    try:
        loc = parsed.get('loc') or {}
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        sheet_name = loc.get('sheet')
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        highlights = [h for h in collect_highlights(parsed) if h.get('cell')]
        layout = _grid_layout(ws, highlights, max_rows, max_cols)
        if not layout:
            return None
        r, c = cell
        if not (layout['rmin'] <= r <= layout['r1']
                and layout['cmin'] <= c <= layout['c1']):
            return None
        x0, y0, x1, y1 = layout['rect_for'](r, c)
        return tuple(int(v * scale) for v in (x0, y0, x1, y1))
    except Exception:
        return None


def _render(data, parsed, findings, filename, max_rows, max_cols, scale):
    loc = parsed.get('loc') or {}
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheet_name = loc.get('sheet')
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    highlights = [h for h in collect_highlights(parsed) if h.get('cell')]

    layout = _grid_layout(ws, highlights, max_rows, max_cols)
    if not layout:
        return None
    rmin, cmin, r1, c1 = layout['rmin'], layout['cmin'], layout['r1'], layout['c1']
    rows, cols, spans = layout['rows'], layout['cols'], layout['spans']
    cell_text = layout['cell_text']
    rect_for = layout['rect_for']
    col_x, cw, row_y = layout['col_x'], layout['cw'], layout['row_y']
    RH, GUT, TH, TITLE_H = layout['RH'], layout['GUT'], layout['TH'], layout['TITLE_H']
    grid_right, grid_bottom = layout['grid_right'], layout['grid_bottom']
    PAD = 8

    # ── Legend: one entry per distinct finding note (a finding may box >1 cell) ──
    # Numbered via build_issue_index, the SAME numbering the UI's issue cards
    # use — a badge drawn on this image must mean the same finding as the
    # matching number in the findings list beside it, not a coincidentally
    # different one from a second, independently-ordered numbering scheme.
    canonical_issues, _canonical_extras = build_issue_index(parsed, findings)
    order = [issue['note'] for issue in canonical_issues]
    meta = {issue['note']: (issue['severity'], issue['category'])
            for issue in canonical_issues}
    num = {issue['note']: issue['num'] for issue in canonical_issues}

    content_w = max(grid_right, 760)
    wrap_chars = max(40, int((content_w - 70) / 8))
    leg_lines = []
    for note in order:
        sev, cat = meta[note]
        wrapped = textwrap.wrap(f"{cat} — {note}", wrap_chars) or ['']
        leg_lines.append((num[note], sev, wrapped))

    # Warning/critical findings not represented by a box (filename, photo
    # legends, caption numbering, …) — listed so nothing is silently missing.
    boxed_notes = {h['note'] for h in highlights}
    extra_lines = []
    for sev, cat, msg in (findings or []):
        if sev in ('critical', 'warning') and msg not in boxed_notes:
            extra_lines.append((sev, textwrap.wrap(f"{cat} — {msg}", wrap_chars) or ['']))

    def _block_h(lines):
        return 28 + sum(22 + 18 * (len(w) - 1) + 8 for *_, w in lines)

    if order or extra_lines:
        leg_h = 16
        if order:
            leg_h += _block_h(leg_lines)
        if extra_lines:
            leg_h += _block_h(extra_lines)
    else:
        leg_h = 44
    W = int(content_w + 16)
    H = int(grid_bottom + 18 + leg_h + 12)

    img = Image.new('RGB', (W * scale, H * scale), _WHITE)
    d = ImageDraw.Draw(img)
    f_title = _font(20 * scale, bold=True)
    f_sub   = _font(13 * scale)
    f_cellS = _font(13 * scale)
    f_hdrS  = _font(12 * scale, bold=True)
    f_legS  = _font(14 * scale)
    f_legbS = _font(14 * scale, bold=True)
    f_badge = _font(13 * scale, bold=True)

    def S(v):                                   # scale a coordinate
        return int(v * scale)

    # Title band.
    d.rectangle([0, 0, W * scale, S(TITLE_H)], fill=_TITLE_BG)
    title = filename or sheet_name or 'Lab report'
    d.text((S(14), S(11)), _fit(d, title, f_title, (W - 28) * scale),
           font=f_title, fill=_WHITE)
    counts = {}
    for h in highlights:
        counts[h['severity']] = counts.get(h['severity'], 0) + 1
    sub = "Annotated review — " + (
        ", ".join(f"{counts[s]} {s}" for s in ('critical', 'warning', 'info')
                  if counts.get(s)) or "no cell-level issues flagged")
    d.text((S(15), S(38)), sub, font=f_sub, fill=(200, 208, 220))

    # Column-letter header band.
    d.rectangle([S(GUT), S(TITLE_H), S(grid_right), S(TITLE_H + TH)], fill=_HDR_BG)
    for c in cols:
        cx = S(col_x[c] + cw[c] / 2)
        d.text((cx, S(TITLE_H + 5)), get_column_letter(c), font=f_hdrS,
               fill=_MUTED, anchor='ma')
    # Row-number gutter.
    d.rectangle([0, S(TITLE_H + TH), S(GUT), S(grid_bottom)], fill=_HDR_BG)
    for r in rows:
        d.text((S(GUT - 6), S(row_y[r] + RH / 2)), str(r), font=f_hdrS,
               fill=_MUTED, anchor='rm')

    # Cells (skip merge-covered non-anchors).
    drawn = set()
    for r in rows:
        for c in cols:
            box = spans.get((r, c))
            if box and box[:2] != (r, c):
                continue
            x0, y0, x1, y1 = rect_for(r, c)
            if (x0, y0, x1, y1) in drawn:
                continue
            drawn.add((x0, y0, x1, y1))
            d.rectangle([S(x0), S(y0), S(x1), S(y1)], outline=_GRID, width=1)
            txt = cell_text(r, c)
            if txt:
                d.text((S(x0 + PAD), S((y0 + y1) / 2)),
                       _fit(d, txt, f_cellS, (x1 - x0 - 2 * PAD) * scale),
                       font=f_cellS, fill=_TEXT, anchor='lm')

    # Highlight boxes + badges. A finding may box several cells (so the badge
    # number ties back to one legend line); a cell may carry several numbers.
    # The badge sits in the cell's top-right corner — usually clear of the value.
    by_cell = {}
    for h in highlights:
        by_cell.setdefault(h['cell'], set()).add((num[h['note']], h['severity']))
    for cell, items in by_cell.items():
        sev = max((s for _, s in items), key=lambda s: _SEV_RANK[s])
        color = _SEV_RGB[sev]
        x0, y0, x1, y1 = rect_for(*cell)
        d.rounded_rectangle([S(x0) + 1, S(y0) + 1, S(x1) - 1, S(y1) - 1],
                            radius=4 * scale, outline=color, width=max(2, scale + 1))
        label = ",".join(str(n) for n, _ in sorted(items))
        bh = 16 * scale
        bw = _textw(d, label, f_badge) + 9 * scale
        bx = max(S(x0) + 1, S(x1) - bw - 1)
        by = S(y0) + 1
        d.rounded_rectangle([bx, by, bx + bw, by + bh], radius=3 * scale, fill=color)
        d.text((bx + bw / 2, by + bh / 2), label, font=f_badge, fill=_WHITE, anchor='mm')

    # Legend panel.
    ly = grid_bottom + 16
    if order or extra_lines:
        if order:
            d.text((S(14), S(ly)), "Findings", font=f_legbS, fill=_TEXT)
            ly += 26
            for i, sev, wrapped in leg_lines:
                color = _SEV_RGB[sev]
                d.rounded_rectangle([S(14), S(ly), S(34), S(ly + 18)],
                                    radius=4 * scale, fill=color)
                d.text((S(24), S(ly + 9)), str(i), font=f_badge, fill=_WHITE, anchor='mm')
                d.text((S(44), S(ly)), wrapped[0], font=f_legS, fill=_TEXT)
                for ex in wrapped[1:]:
                    ly += 18
                    d.text((S(44), S(ly)), ex, font=f_legS, fill=_MUTED)
                ly += 26
        if extra_lines:
            d.text((S(14), S(ly)), "Also flagged (not tied to a cell)",
                   font=f_legbS, fill=_TEXT)
            ly += 26
            for sev, wrapped in extra_lines:
                d.ellipse([S(17), S(ly + 4), S(31), S(ly + 18)], fill=_SEV_RGB[sev])
                d.text((S(44), S(ly)), wrapped[0], font=f_legS, fill=_TEXT)
                for ex in wrapped[1:]:
                    ly += 18
                    d.text((S(44), S(ly)), ex, font=f_legS, fill=_MUTED)
                ly += 26
    else:
        d.text((S(14), S(ly)), "No cell-level issues to highlight on this sheet.",
               font=f_legS, fill=_MUTED)

    out = io.BytesIO()
    img.save(out, format='PNG')
    return out.getvalue()


# ════════════════════════════════════════════════════════════════════════
# ANNOTATED MICROGRAPHS
# ════════════════════════════════════════════════════════════════════════
def annotate_micrographs(data, parsed, max_images=12):
    """Return [(image_name, png_bytes, caption)] with issue regions boxed.

    Uses the per-image analysis already on parsed['images'] when present, else
    runs analyze_images() here. Boxes the burned-in legend and scale-bar regions,
    flags low contrast, and surfaces any thickness measurements read.
    """
    if not _PIL:
        return []
    try:
        z = zipfile.ZipFile(io.BytesIO(data))
    except Exception:
        return []
    media = {n.split('/')[-1]: n for n in z.namelist() if n.startswith('xl/media')}

    images = parsed.get('images')
    if not images:
        images, _ = analyze_images(data, max_images=max_images)

    out = []
    for entry in images[:max_images]:
        name = entry.get('image')
        if name not in media:
            continue
        try:
            im = Image.open(io.BytesIO(z.read(media[name]))).convert('RGB')
        except Exception:
            continue
        out.append((name, _annotate_one(im, entry), _micro_caption(entry)))
    return out


def _micro_caption(entry):
    bits = []
    if entry.get('mag'):
        bits.append(entry['mag'])
    if entry.get('id'):
        bits.append(entry['id'])
    if entry.get('scale'):
        bits.append(f"scale {entry['scale']}")
    bits.append("etched" if entry.get('etched') else "low-contrast")
    meas = entry.get('measurements') or []
    if meas:
        bits.append("thickness " + ", ".join(f"{m} µm" for m in meas))
    return "  ·  ".join(bits)


def _annotate_one(im, entry):
    w, h = im.size
    big = max(1, 900 // max(1, w))                 # upsample small micrographs a touch
    if big > 1:
        im = im.resize((w * big, h * big))
        w, h = im.size
    d = ImageDraw.Draw(im, 'RGBA')
    f = _font(max(13, w // 42), bold=True)
    fs = _font(max(12, w // 48))

    def banner(xy, text, color):
        tw = _textw(d, text, fs) + 14
        d.rectangle([xy[0], xy[1], xy[0] + tw, xy[1] + fs.size + 10], fill=color + (235,))
        d.text((xy[0] + 7, xy[1] + 5), text, font=fs, fill=_WHITE)

    # Legend region (bottom-left) — where the ID / magnification is read.
    lr = (2, int(h * 0.90), int(w * 0.55), h - 2)
    d.rectangle(lr, outline=_SEV_RGB['warning'], width=max(2, w // 360))
    tag = "legend"
    if entry.get('mag') or entry.get('id'):
        tag += ": " + (entry.get('id') or entry.get('mag'))
    banner((lr[0], max(0, lr[1] - fs.size - 12)), tag, _SEV_RGB['warning'])

    # Scale-bar region (bottom-right).
    sr = (int(w * 0.72), int(h * 0.88), w - 2, h - 2)
    d.rectangle(sr, outline=_SEV_RGB['info'], width=max(2, w // 360))
    banner((sr[0], max(0, sr[1] - fs.size - 12)),
           "scale" + (f": {entry['scale']}" if entry.get('scale') else ""),
           _SEV_RGB['info'])

    # Contrast / thickness call-outs along the top.
    y = 4
    if entry.get('strong') is not None and not entry.get('etched'):
        banner((4, y), "low contrast — verify etch state", _SEV_RGB['critical'])
        y += fs.size + 14
    meas = entry.get('measurements') or []
    if meas:
        banner((4, y), "thickness read: " + ", ".join(f"{m} µm" for m in meas),
               _SEV_RGB['pass'])

    out = io.BytesIO()
    im.save(out, format='PNG')
    return out.getvalue()


# ════════════════════════════════════════════════════════════════════════
# PIXEL-FAITHFUL VIEW  (LibreOffice render of the real workbook + overlays)
# ════════════════════════════════════════════════════════════════════════
# How it works: the flagged cells are filled in the workbook with an
# exactly-known colour (a severity tint, perturbed per finding so each is
# unique) and given a medium border in the severity colour. LibreOffice then
# renders the *real* workbook — original fonts, column widths, borders, merged
# cells and embedded micrographs intact — to PDF, which PyMuPDF rasterises.
# Each finding's cell is then found back by its unique fill colour (no fragile
# cell→pixel maths), so a numbered badge lands exactly on it and a legend ties
# the numbers to the findings. Falls back (caller's job) to the drawn grid when
# LibreOffice or PyMuPDF is unavailable.

_TINT = {                       # light fills (kept readable over the cell text)
    'critical': (255, 214, 217),
    'warning':  (255, 232, 200),
    'info':     (210, 228, 250),
    'pass':     (212, 240, 222),
}
_BORDER_HEX = {'critical': 'C62D38', 'warning': 'D9821A',
               'info': '1A6ED6', 'pass': '1F9E50'}


def libreoffice_available():
    """True when a pixel-faithful render is possible here."""
    return bool(_PIL and _FITZ and _find_soffice())


def _find_soffice():
    for name in ('soffice', 'libreoffice'):
        p = shutil.which(name)
        if p:
            return p
    for p in ('/usr/bin/soffice', '/usr/bin/libreoffice',
              '/opt/libreoffice/program/soffice'):
        if os.path.exists(p):
            return p
    return None


def _unique_fill(i, sev):
    """A severity-tinted fill that is unique per finding index (so it can be
    found back in the raster by exact colour). Steps of 6 keep masks (±2) apart."""
    r, g, b = _TINT.get(sev, _TINT['warning'])
    b = max(140, b - (i % 8) * 6)
    g = max(140, g - (i // 8) * 6)
    return (r, g, b)


def build_issue_index(parsed, findings=None):
    """Return unique, report-centric issues and unanchored findings.

    A single finding may intentionally point to several cells. It must still be
    presented and counted once, with every affected cell retained as an anchor.
    """
    highlights = sorted(
        (h for h in collect_highlights(parsed) if h.get('cell')),
        key=lambda h: (h['cell'][0], h['cell'][1]),
    )
    issues = []
    by_finding = {}
    for highlight in highlights:
        key = (highlight.get('category') or 'Review', highlight['note'])
        issue = by_finding.get(key)
        if issue is None:
            issue = {
                'num': len(issues) + 1,
                'severity': highlight['severity'],
                'category': key[0],
                'note': key[1],
                'cells': [],
                'refs': [],
                'pages': [],
            }
            by_finding[key] = issue
            issues.append(issue)
        elif _SEV_RANK[highlight['severity']] > _SEV_RANK[issue['severity']]:
            issue['severity'] = highlight['severity']
        cell = tuple(highlight['cell'])
        if cell not in issue['cells']:
            issue['cells'].append(cell)
            issue['refs'].append(f'{get_column_letter(cell[1])}{cell[0]}')

    boxed_notes = {issue['note'] for issue in issues}

    def finding_stem(text):
        """Comparable leading clause for a finding and its longer explanation."""
        leading = re.split(r'\s+[—–]\s+', text or '', maxsplit=1)[0]
        return re.sub(r'[\s.]+$', '', leading).casefold()

    boxed_stems = {finding_stem(note) for note in boxed_notes}
    extras = []
    by_extra = {}
    for severity, category, message in findings or []:
        if (severity not in ('critical', 'warning')
                or message in boxed_notes
                or finding_stem(message) in boxed_stems):
            continue
        key = (category, message)
        extra = by_extra.get(key)
        if extra is None:
            extra = {
                'severity': severity,
                'category': category,
                'note': message,
            }
            by_extra[key] = extra
            extras.append(extra)
        elif _SEV_RANK[severity] > _SEV_RANK[extra['severity']]:
            extra['severity'] = severity
    return issues, extras


def render_report_faithful_view(
        data, parsed, findings=None, filename=None, dpi=150, timeout=90,
        fit_width=False):
    """Return a page-oriented annotated report package and a status string.

    The package contains individually annotated page PNGs, unique issue
    records, report-level findings, and PDF/PNG downloads. This is the primary
    UI product; it avoids forcing every page and finding into one tall image.

    fit_width : the AEG template's own print setup is narrower than the sheet,
        so a faithful render splits columns across pages — page 1 of report
        6943 cuts off mid-table (Result, Remarks and the outgoing-coating
        columns land on later, otherwise near-empty pages). Passing True adds
        fit-to-one-page-wide scaling so a reviewer sees every column of a row
        together. It is deliberately OFF by default: this function's contract
        is fidelity to the document as it prints, and the callers that want
        readability over fidelity opt in.
    """
    if not _PIL:
        return None, 'Pillow unavailable'
    if not _FITZ:
        return None, 'PyMuPDF unavailable'
    if not _find_soffice():
        return None, 'LibreOffice not installed'
    try:
        return _faithful_view(data, parsed, findings, filename, dpi, timeout,
                              fit_width), 'ok'
    except subprocess.TimeoutExpired:
        return None, 'LibreOffice timed out'
    except Exception as e:
        return None, f'{type(e).__name__}: {e}'


def render_report_faithful(data, parsed, findings=None, filename=None, dpi=130, timeout=90):
    """Return (png_bytes, status). status is 'ok' or a short reason on failure
    (so the caller can fall back to render_report_image). `findings` lets the
    legend also list warning/critical findings that aren't tied to a cell."""
    view, status = render_report_faithful_view(
        data, parsed, findings, filename, dpi, timeout)
    return ((view or {}).get('combined_png'), status)


def _faithful_view(data, parsed, findings, filename, dpi, timeout, fit_width=False):
    loc = parsed.get('loc') or {}
    issues, extras = build_issue_index(parsed, findings)

    wb = openpyxl.load_workbook(io.BytesIO(data))      # keep styles + images
    sheet = loc.get('sheet')
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    if fit_width:
        # Scale to one page wide, unlimited pages tall, and drop the column
        # breaks that would otherwise still split the sheet horizontally.
        # Height is left free so rows are not crushed to fit.
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        try:
            ws.col_breaks.brk = []
        except Exception:
            pass

    span = {}
    for rng in ws.merged_cells.ranges:
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                span[(rr, cc)] = rng

    # One cell can carry several unique issues, while one issue can point to
    # several cells. Each cell gets one unique colour locator and a badge with
    # the relevant issue number(s).
    cells, by_cell = [], {}
    for issue in issues:
        for cell in issue['cells']:
            group = by_cell.get(cell)
            if group is None:
                group = {
                    'cell': cell,
                    'severity': issue['severity'],
                    'issue_nums': [issue['num']],
                }
                by_cell[cell] = group
                cells.append(group)
            else:
                if issue['num'] not in group['issue_nums']:
                    group['issue_nums'].append(issue['num'])
                if _SEV_RANK[issue['severity']] > _SEV_RANK[group['severity']]:
                    group['severity'] = issue['severity']

    anchors = []
    for i, group in enumerate(cells):
        r, c = group['cell']
        sev = group['severity']
        rgb = _unique_fill(i, sev)
        fill = PatternFill('solid', fgColor='%02X%02X%02X' % rgb)
        side = Side(style='medium', color=_BORDER_HEX.get(sev, 'D9821A'))
        border = Border(left=side, right=side, top=side, bottom=side)
        rng = span.get((r, c))
        cells = ([(rr, cc) for rr in range(rng.min_row, rng.max_row + 1)
                  for cc in range(rng.min_col, rng.max_col + 1)]
                 if rng else [(r, c)])
        for rr, cc in cells:
            cell = ws.cell(row=rr, column=cc)
            cell.fill = fill
            cell.border = border
        anchors.append({
            'severity': sev,
            'issue_nums': sorted(group['issue_nums']),
            'rgb': rgb,
            'ref': f'{get_column_letter(c)}{r}',
        })

    with tempfile.TemporaryDirectory() as tmp:
        xpath = os.path.join(tmp, 'annotated.xlsx')
        wb.save(xpath)
        pdf = _xlsx_to_pdf(xpath, tmp, timeout)
        if not pdf:
            raise RuntimeError('LibreOffice conversion failed')
        pages = _raster_pdf(pdf, dpi)
    if not pages:
        raise RuntimeError('no pages rendered')
    source_page_count = len(pages)
    pages, omitted_blank_pages = _trim_blank_trailing_pages(pages)
    if omitted_blank_pages:
        extras.append({
            'severity': 'warning',
            'category': 'Pagination',
            'note': (
                f'The workbook print setup produced {source_page_count} pages, but '
                f'{len(omitted_blank_pages)} trailing page(s) contained no report content '
                f'and were omitted from the effective report view: '
                f'{", ".join(map(str, omitted_blank_pages))}.'
            ),
        })
    view = _compose_faithful_view(pages, anchors, issues, extras, filename, dpi)
    view['source_page_count'] = source_page_count
    view['effective_page_count'] = len(pages)
    view['omitted_blank_pages'] = omitted_blank_pages
    return view


def _xlsx_to_pdf(xlsx_path, outdir, timeout):
    exe = _find_soffice()
    profile = 'file://' + os.path.join(outdir, 'lo_profile')
    cmd = [exe, '--headless', '--norestore', '--invisible', '--nologo',
           '-env:UserInstallation=' + profile,
           '--convert-to', 'pdf:calc_pdf_Export', '--outdir', outdir, xlsx_path]
    with _LO_LOCK:
        # start_new_session so a timeout can kill the whole soffice process group
        # (soffice forks soffice.bin — killing just the launcher leaves a zombie).
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                                start_new_session=True)
        try:
            proc.communicate(timeout=timeout)
        except subprocess.TimeoutExpired:
            try:
                os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
            except Exception:
                proc.kill()
            proc.communicate()
            raise
    pdf = os.path.join(outdir, os.path.splitext(os.path.basename(xlsx_path))[0] + '.pdf')
    return pdf if os.path.exists(pdf) else None


def _raster_pdf(pdf, dpi):
    try:                       # quieten benign 'structure tree' notices on tagged PDFs
        fitz.TOOLS.mupdf_display_errors(False)
    except Exception:
        pass
    doc = fitz.open(pdf)
    mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pages = []
    for pg in doc:
        pix = pg.get_pixmap(matrix=mat, alpha=False)
        pages.append(Image.frombytes('RGB', (pix.width, pix.height), pix.samples).copy())
    doc.close()
    return pages


def _visually_blank_report_page(page):
    """True for a page containing only margins/footer furniture.

    Trailing pages in the legacy workbooks often contain only the hard-coded
    footer. A low whole-page ink ratio distinguishes that footer furniture
    from a short but valid sign-off/remarks page.
    """
    gray = page.convert('L')
    width, height = gray.size
    central = gray.crop((int(width * 0.03), 0, int(width * 0.97), height))
    histogram = central.histogram()
    dark = sum(histogram[:245])
    pixels = max(1, central.width * central.height)
    return dark / pixels < 0.004


def _trim_blank_trailing_pages(pages):
    """Return (effective pages, omitted 1-based source page numbers)."""
    kept = list(pages)
    omitted = []
    while len(kept) > 1 and _visually_blank_report_page(kept[-1]):
        omitted.append(len(kept))
        kept.pop()
    omitted.reverse()
    return kept, omitted


def _color_bbox(img, rgb, tol=2):
    """Bounding box of pixels matching rgb within tol on every channel, or None."""
    bands = []
    for ch, v in zip(img.split()[:3], rgb):
        lo, hi = v - tol, v + tol
        bands.append(ch.point(lambda p: 255 if lo <= p <= hi else 0))
    mask = ImageChops.multiply(ImageChops.multiply(bands[0], bands[1]), bands[2])
    return mask.getbbox()


def _annotate_faithful_pages(pages, anchors, issues, dpi):
    """Draw issue markers on individual report pages.

    Cell fills are unique locator colours. Several cells may carry the same
    issue number, and a cell may carry more than one issue number.
    """
    located = {}
    for index, anchor in enumerate(anchors):
        for page_index, page in enumerate(pages):
            bbox = _color_bbox(page, anchor['rgb'])
            if bbox:
                located[index] = (page_index, bbox)
                break

    issue_pages = {issue['num']: set() for issue in issues}
    by_page = {index: [] for index in range(len(pages))}
    for index, (page_index, bbox) in located.items():
        anchor = anchors[index]
        by_page[page_index].append((anchor, bbox))
        for number in anchor['issue_nums']:
            issue_pages[number].add(page_index + 1)

    radius = max(11, int(dpi * 0.085))
    font = _font(max(12, int(radius * 1.05)), bold=True)
    page_entries = []
    annotated_images = []
    for page_index, source in enumerate(pages):
        page = source.copy().convert('RGB')
        draw = ImageDraw.Draw(page)
        page_issue_nums = set()
        placed = []          # badge rects already drawn on this page
        for anchor, bbox in by_page[page_index]:
            numbers = anchor['issue_nums']
            page_issue_nums.update(numbers)
            label = ','.join(str(number) for number in numbers)
            color = _SEV_RGB[anchor['severity']]

            # Reinforce the affected region after rasterisation.
            draw.rounded_rectangle(
                [bbox[0], bbox[1], max(bbox[0] + 2, bbox[2] - 1),
                 max(bbox[1] + 2, bbox[3] - 1)],
                radius=3,
                outline=color,
                width=max(2, int(dpi / 72)),
            )

            # The badge must never sit ON a value — not the flagged one (the
            # whole point is to read it) and not its neighbour's either. Park
            # it straddling the cell's top-right CORNER: cell text is
            # vertically centred in its row and horizontally inset from the
            # column border, so the corner junction is the one spot that is
            # reliably rule-and-whitespace. Nudge down when it would land on a
            # badge already drawn, so markers never stack.
            label_w = max(2 * radius, int(_textw(draw, label, font) + radius))
            x0 = bbox[2] - label_w / 2
            x0 = max(3, min(x0, page.width - 3 - label_w))
            y0 = max(3, min(bbox[1] - radius, page.height - 3 - 2 * radius))

            def _hits(ax0, ay0):
                return any(not (ax0 + label_w < px0 or ax0 > px1
                                or ay0 + 2 * radius < py0 or ay0 > py1)
                           for px0, py0, px1, py1 in placed)

            step = radius
            for _ in range(6):
                if not _hits(x0, y0):
                    break
                y0 = min(y0 + step, page.height - 3 - 2 * radius)
            x1, y1 = x0 + label_w, y0 + 2 * radius
            placed.append((x0, y0, x1, y1))

            draw.rounded_rectangle(
                [x0, y0, x1, y1],
                radius=radius,
                fill=color,
                outline=_WHITE,
                width=max(2, int(dpi / 72)),
            )
            draw.text(
                ((x0 + x1) / 2, (y0 + y1) / 2),
                label,
                font=font,
                fill=_WHITE,
                anchor='mm',
            )

        output = io.BytesIO()
        page.save(output, format='PNG', optimize=True)
        page_entries.append({
            'number': page_index + 1,
            'png': output.getvalue(),
            'issue_nums': sorted(page_issue_nums),
        })
        annotated_images.append(page)

    for issue in issues:
        issue['pages'] = sorted(issue_pages.get(issue['num'], set()))
    return page_entries, annotated_images


def _pages_to_pdf(pages, dpi):
    if not pages:
        return None
    output = io.BytesIO()
    first, rest = pages[0].convert('RGB'), [page.convert('RGB') for page in pages[1:]]
    first.save(
        output,
        format='PDF',
        save_all=True,
        append_images=rest,
        resolution=dpi,
    )
    return output.getvalue()


def _wrap_download_text(draw, text, font, max_width):
    """Wrap text by rendered pixel width without losing long tokens."""
    paragraphs = str(text or '').splitlines() or ['']
    lines = []
    for paragraph in paragraphs:
        words = paragraph.split()
        if not words:
            lines.append('')
            continue
        line = ''
        for word in words:
            candidate = f'{line} {word}'.strip()
            if not line or _textw(draw, candidate, font) <= max_width:
                line = candidate
                continue

            lines.append(line)
            line = ''
            while word and _textw(draw, word, font) > max_width:
                cut = len(word)
                while cut > 1 and _textw(draw, word[:cut], font) > max_width:
                    cut -= 1
                lines.append(word[:cut])
                word = word[cut:]
            line = word
        if line:
            lines.append(line)
    return lines or ['']


def _download_issue_cards(records, panel_width, panel_height, dpi):
    """Prepare fully wrapped cards and split long comments into continuations."""
    scale = max(0.8, dpi / 150.0)
    pad = max(18, int(24 * scale))
    card_pad = max(12, int(15 * scale))
    title_font = _font(max(15, int(18 * scale)), bold=True)
    meta_font = _font(max(11, int(13 * scale)), bold=True)
    note_font = _font(max(13, int(16 * scale)))
    title_lh = max(19, int(24 * scale))
    meta_lh = max(16, int(19 * scale))
    note_lh = max(18, int(23 * scale))
    card_width = panel_width - 2 * pad
    text_width = card_width - 2 * card_pad
    usable_height = panel_height - max(94, int(116 * scale)) - pad
    max_note_lines = max(
        1,
        int((usable_height - 2 * card_pad - title_lh - meta_lh - 12) / note_lh),
    )

    probe = Image.new('RGB', (8, 8), _WHITE)
    draw = ImageDraw.Draw(probe)
    cards = []
    for record in records:
        number = record.get('num')
        severity = record.get('severity', 'warning')
        label = {
            'critical': 'FAIL',
            'warning': 'WARNING',
            'info': 'NOTE',
            'pass': 'PASS',
        }.get(severity, 'REVIEW')
        title = f'{number}. {label}' if number is not None else f'REPORT CHECK - {label}'
        refs = record.get('refs') or []
        meta = record.get('category') or 'Review'
        if refs:
            meta += ' - ' + ', '.join(refs)
        meta_lines = _wrap_download_text(draw, meta, meta_font, text_width)
        note_lines = _wrap_download_text(
            draw, record.get('note') or '', note_font, text_width)

        # Long comments remain complete by continuing them in another card/page.
        chunk_size = max(1, max_note_lines - max(0, len(meta_lines) - 1))
        chunks = [
            note_lines[index:index + chunk_size]
            for index in range(0, len(note_lines), chunk_size)
        ] or [['']]
        for index, chunk in enumerate(chunks):
            chunk_title = title + (' (continued)' if index else '')
            height = (
                2 * card_pad + title_lh + 5
                + len(meta_lines) * meta_lh + 6
                + len(chunk) * note_lh
            )
            cards.append({
                'severity': severity,
                'title': chunk_title,
                'title_font': title_font,
                'title_lh': title_lh,
                'meta_lines': meta_lines,
                'meta_font': meta_font,
                'meta_lh': meta_lh,
                'note_lines': chunk,
                'note_font': note_font,
                'note_lh': note_lh,
                'padding': card_pad,
                'height': height,
            })
    return cards, pad, usable_height


def _paginate_download_cards(cards, usable_height, gap):
    """Pack issue cards onto panels without dropping overflow comments."""
    if not cards:
        return [[]]
    groups, current, used = [], [], 0
    for card in cards:
        needed = card['height'] + (gap if current else 0)
        if current and used + needed > usable_height:
            groups.append(current)
            current, used = [], 0
            needed = card['height']
        current.append(card)
        used += needed
    if current:
        groups.append(current)
    return groups


def _download_page_records(page_entry, issues, extras, first_page):
    numbers = set(page_entry.get('issue_nums') or [])
    records = [issue for issue in issues if issue.get('num') in numbers]
    if first_page:
        records.extend(issue for issue in issues if not issue.get('pages'))
        records.extend(extras or [])
    return records


def _draw_download_panel(
        canvas, panel_x, panel_width, page_number, report_count, filename,
        cards, panel_index, panel_count, pad, dpi):
    """Draw comments beside one annotated report page."""
    draw = ImageDraw.Draw(canvas)
    scale = max(0.8, dpi / 150.0)
    draw.rectangle(
        [panel_x, 0, panel_x + panel_width, canvas.height],
        fill=(246, 248, 251),
    )
    draw.line(
        [(panel_x, 0), (panel_x, canvas.height)],
        fill=(205, 211, 220),
        width=max(2, int(2 * scale)),
    )

    heading_font = _font(max(20, int(27 * scale)), bold=True)
    sub_font = _font(max(11, int(13 * scale)), bold=True)
    file_font = _font(max(10, int(12 * scale)))
    x = panel_x + pad
    y = pad
    draw.text((x, y), 'Review comments', font=heading_font, fill=_TEXT)
    y += max(31, int(38 * scale))
    page_label = f'REPORT PAGE {page_number} OF {report_count}'
    if panel_count > 1:
        page_label += f'  |  COMMENTS {panel_index} OF {panel_count}'
    draw.text((x, y), page_label, font=sub_font, fill=(83, 96, 116))
    y += max(22, int(27 * scale))
    draw.text(
        (x, y),
        _fit(draw, filename or 'Lab report', file_font, panel_width - 2 * pad),
        font=file_font,
        fill=(104, 115, 132),
    )
    y += max(30, int(39 * scale))

    if not cards:
        clear_font = _font(max(15, int(18 * scale)), bold=True)
        body_font = _font(max(12, int(15 * scale)))
        body_lines = _wrap_download_text(
            draw,
            'The report page is included unchanged except for any numbered markers.',
            body_font,
            panel_width - 4 * pad,
        )
        body_lh = max(17, int(21 * scale))
        clear_height = max(
            96,
            2 * pad + max(25, int(30 * scale)) + len(body_lines) * body_lh,
        )
        draw.rounded_rectangle(
            [x, y, panel_x + panel_width - pad, y + clear_height],
            radius=max(8, int(10 * scale)),
            fill=(244, 252, 247),
            outline=(87, 174, 116),
            width=max(1, int(2 * scale)),
        )
        draw.text(
            (x + pad, y + pad),
            'No marked issues on this page',
            font=clear_font,
            fill=(32, 122, 67),
        )
        body_y = y + pad + max(28, int(33 * scale))
        for line in body_lines:
            draw.text(
                (x + pad, body_y),
                line,
                font=body_font,
                fill=(79, 98, 87),
            )
            body_y += body_lh
        return

    gap = max(10, int(13 * scale))
    for card in cards:
        color = _SEV_RGB.get(card['severity'], _SEV_RGB['warning'])
        fill = {
            'critical': (255, 247, 247),
            'warning': (255, 250, 243),
            'info': (247, 250, 255),
            'pass': (246, 252, 248),
        }.get(card['severity'], _WHITE)
        x1 = panel_x + panel_width - pad
        y1 = min(canvas.height - pad, y + card['height'])
        draw.rounded_rectangle(
            [x, y, x1, y1],
            radius=max(8, int(10 * scale)),
            fill=fill,
            outline=(213, 219, 228),
            width=max(1, int(1.5 * scale)),
        )
        draw.rounded_rectangle(
            [x, y, x + max(5, int(6 * scale)), y1],
            radius=max(2, int(3 * scale)),
            fill=color,
        )
        tx = x + card['padding']
        ty = y + card['padding']
        draw.text(
            (tx, ty),
            card['title'],
            font=card['title_font'],
            fill=color,
        )
        ty += card['title_lh'] + 5
        for line in card['meta_lines']:
            draw.text((tx, ty), line, font=card['meta_font'], fill=(96, 108, 126))
            ty += card['meta_lh']
        ty += 6
        for line in card['note_lines']:
            draw.text((tx, ty), line, font=card['note_font'], fill=(45, 55, 72))
            ty += card['note_lh']
        y = y1 + gap


def _compose_download_pages(
        annotated_pages, page_entries, issues, extras, filename, dpi):
    """Create PDF-ready pages with every issue comment visibly embedded."""
    download_pages = []
    report_count = len(annotated_pages)
    for index, (report, entry) in enumerate(zip(annotated_pages, page_entries)):
        report = report.convert('RGB')
        scale = max(0.8, dpi / 150.0)
        panel_width = max(int(report.width * 0.42), int(500 * scale))
        records = _download_page_records(entry, issues, extras, index == 0)
        cards, pad, usable_height = _download_issue_cards(
            records, panel_width, report.height, dpi)
        gap = max(10, int(13 * scale))
        card_groups = _paginate_download_cards(cards, usable_height, gap)
        # Overflow comments use additional side columns on the same PDF page.
        # The annotated download therefore keeps exactly the report's source
        # page count while guaranteeing that every comment remains visible.
        canvas = Image.new(
            'RGB',
            (report.width + panel_width * len(card_groups), report.height),
            _WHITE,
        )
        canvas.paste(report, (0, 0))
        for panel_index, group in enumerate(card_groups, start=1):
            panel_x = report.width + panel_width * (panel_index - 1)
            _draw_download_panel(
                canvas,
                panel_x,
                panel_width,
                entry.get('number', index + 1),
                report_count,
                filename,
                group,
                panel_index,
                len(card_groups),
                pad,
                dpi,
            )
        download_pages.append(canvas)
    return download_pages


def _stack_pages(pages):
    """Compatibility/download image; the interactive UI uses one page at a time."""
    if not pages:
        return None
    margin, gap = 12, 12
    width = max(page.width for page in pages) + 2 * margin
    height = sum(page.height for page in pages) + gap * (len(pages) - 1) + 2 * margin
    canvas = Image.new('RGB', (width, height), (245, 247, 250))
    y = margin
    for page in pages:
        x = (width - page.width) // 2
        canvas.paste(page, (x, y))
        y += page.height + gap
    output = io.BytesIO()
    canvas.save(output, format='PNG', optimize=True)
    return output.getvalue()


def _compose_faithful_view(pages, anchors, issues, extras, filename, dpi):
    page_entries, annotated_images = _annotate_faithful_pages(
        pages, anchors, issues, dpi)
    download_pages = _compose_download_pages(
        annotated_images, page_entries, issues, extras, filename, dpi)
    return {
        'filename': filename or 'Lab report',
        'pages': page_entries,
        'issues': issues,
        'extras': extras,
        'annotated_pdf': _pages_to_pdf(download_pages, dpi),
        'combined_png': _stack_pages(annotated_images),
    }
