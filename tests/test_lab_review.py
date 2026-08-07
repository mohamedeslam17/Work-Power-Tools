import unittest
from unittest.mock import patch

from openpyxl import Workbook

import lab_review
from lab_review import (
    _canon_machine,
    _comment,
    _formula_issues,
    _review_comment,
    _review_completeness,
    _review_composition,
    _review_hardness,
    _review_captions,
    _review_acceptance_and_methods,
    _review_document_control,
    _review_evidence_claims,
    _review_legends,
    _review_phase_evidence,
    _review_sampling_basis,
    _review_thickness,
    _review_traceability,
    _select_magnification,
    _vote_job,
    add_version_findings,
    collect_highlights,
    find_duplicate_compositions,
    picture_magnification_verdicts,
    review_filename,
)


def messages(findings, severity=None):
    return [
        message for sev, _category, message in findings
        if severity is None or sev == severity
    ]


class LabReviewRegressionTests(unittest.TestCase):
    def test_revised_upload_is_flagged_when_blockers_persist(self):
        blocker = ("critical", "Acceptance criteria", "No controlled criterion.")
        reports = [
            {
                "name": "7080 Metallurgical Report.xlsx",
                "parsed": {"header": {"job": "7080"}},
                "findings": [blocker],
            },
            {
                "name": "7080 Metallurgical Report - R.xlsx",
                "parsed": {"header": {"job": "7080"}},
                "findings": [blocker],
            },
        ]

        add_version_findings(reports)

        self.assertNotIn("version_findings", reports[0])
        self.assertEqual(reports[1]["version_findings"][0][0], "critical")
        self.assertEqual(reports[1]["version_findings"][0][1], "Revision control")

    def test_plural_comments_label_is_parsed(self):
        ws = Workbook().active
        ws["A1"] = "Comments:"
        ws["A2"] = "A complete metallurgical discussion."

        value, location = _comment(ws)

        self.assertEqual(value, "A complete metallurgical discussion.")
        self.assertEqual(location["value"], (2, 1))

    def test_missing_captioned_micrographs_is_release_blocking(self):
        parsed = {
            "header": {
                "customer": "AEN",
                "job": "7504",
                "machine": "MS7001FA",
                "customer_ref": "4400015000",
                "eoh": "13934",
            },
            "sample": {"material": "GTD-111"},
            "comment": "A sufficiently long metallurgical discussion for the report.",
            "pictures": [(f"Picture {number}:", "Ortho. Acid Etched 100x")
                         for number in range(1, 11)],
            "micrograph_count": 4,
            "signoff": {"met_lab": "Lab", "mat_eng": "Engineer", "date": "2026-07-28"},
        }

        findings = _review_completeness(parsed)

        self.assertTrue(any(
            "6 evidence image(s) are missing" in message
            for message in messages(findings, "critical")
        ))

    def test_duplicate_headers_and_major_lod_are_not_silently_dropped(self):
        nominal = {"Al": 3.0, "Cr": 14.0, "Nb": 2.8}
        actual = {"Cr": 1.7, "Nb": 1.43}
        meta = {
            "actual": {
                "duplicate_headers": ["Nb"],
                "entries": [
                    {"element": "Cr", "raw": "1.7", "value": 1.7},
                    {"element": "Al", "raw": "<LOD", "value": None},
                    {"element": "Nb", "raw": "1.43", "value": 1.43},
                    {"element": "Nb", "raw": "1.43", "value": 1.43},
                ],
            },
        }

        findings = _review_composition(nominal, actual, meta)
        critical = messages(findings, "critical")

        self.assertTrue(any("repeats element header" in message for message in critical))
        self.assertTrue(any("major alloying element was not quantified" in message
                            for message in critical))

    def test_multi_percent_unexpected_vanadium_marks_corrupted_actual_table(self):
        nominal = {
            "Al": 3.4, "Co": 8.5, "Cr": 16.0, "Mo": 1.75,
            "Nb": 0.9, "Ni": 68.49, "Ti": 3.4, "W": 2.6,
        }
        actual = {
            "Al": 4.43, "Co": 8.4, "Cr": 12.64, "Fe": 1.99,
            "Mo": 1.8, "Nb": 0.9, "Ni": 58.7, "Ti": 2.04,
            "V": 5.75,
        }

        findings = _review_composition(nominal, actual)

        self.assertTrue(any(
            "appears corrupted, misaligned or copied" in message
            for message in messages(findings, "critical")
        ))

    def test_see_comment_without_a_verdict_is_not_a_finding(self):
        """A descriptive comment under Result = "See comment" is not a defect.

        This check previously raised a critical. Removed on Mohamed's
        instruction (docs/FEEDBACK.md entry 004): at AEG the comment is a
        record of the metallurgical condition, and the release decision does
        not live in it. It fired on 100% of the real corpus, and its constancy
        was hiding genuine criticals — see D11 in docs/REBUILD.md.

        Inverted rather than deleted so the decision is explicit and the check
        cannot be reintroduced by accident.
        """
        parsed = {
            "comment": (
                "The microstructure contains primary carbides and gamma-prime. "
                "Solution heat treatment dissolved part of the secondary phases."
            ),
            "coating": {},
            "sample": {"material": "Rene80", "result": "See comment"},
        }

        findings = _review_comment(parsed)

        self.assertEqual(
            [m for m in messages(findings) if "disposition" in m.lower()], [])

    def test_comment_contradicting_the_result_is_still_flagged(self):
        """Guard: dropping the disposition check must not weaken the
        contradiction checks, which remain the useful part of this rule."""
        parsed = {
            "comment": "The part is not suitable for further service.",
            "coating": {},
            "sample": {"material": "Rene80", "result": "Acceptable"},
        }

        findings = _review_comment(parsed)

        self.assertTrue(any(
            "NOT suitable" in message for message in messages(findings, "warning")))

    def test_microstructure_does_not_prove_restored_mechanical_properties(self):
        parsed = {
            "comment": (
                "The heat-treatment cycle restores the material's microstructural "
                "stability and mechanical properties."
            ),
            "hardness": {
                "pre": {"value": 35.2, "unit": "HRC"},
                "post": {"value": 32.2, "unit": "HRC"},
            },
        }

        findings = _review_evidence_claims(parsed)

        self.assertTrue(any(
            "no final aged-condition mechanical result" in message
            for message in messages(findings, "warning")
        ))

    def test_sample_and_serial_count_mismatch_is_flagged(self):
        parsed = {
            "sample": {
                "sample_no": "MS 7221C    MS 7217C\nMS 7375C\nMS 7441C",
                "serial": "K3WP045262\nK3WP045203\nK3WP045216",
                "result": "See comment",
            },
        }

        findings = _review_traceability(parsed)

        self.assertTrue(any(
            "do not map one-to-one" in message
            for message in messages(findings, "warning")
        ))

    def test_spaced_serial_prefix_is_one_identifier(self):
        parsed = {
            "sample": {
                "sample_no": "MS 7242C\nMS 7306C\nMS 7431C\nMS 7439C",
                "serial": "H3AR 027331\nH3AR 027329\nH3AR 027537\nH3AR 016178",
                "result": "Acceptable",
            },
        }

        findings = _review_traceability(parsed)

        self.assertFalse(any(category == "Traceability" for _sev, category, _msg in findings))

    def test_slash_serial_is_one_identifier(self):
        parsed = {
            "sample": {
                "sample_no": "MS 6827C",
                "serial": "W007745/1-08",
                "result": "Acceptable",
            },
        }

        findings = _review_traceability(parsed)

        self.assertFalse(any(category == "Traceability" for _sev, category, _msg in findings))

    def test_missing_release_basis_is_not_a_finding_but_methods_still_are(self):
        """A report that does not restate its controlling specification is fine.

        The "Acceptance criteria" critical was removed on Mohamed's instruction
        (docs/FEEDBACK.md entry 005): at AEG the governing specification lives
        in the work order and customer contract, not in the report. It was the
        last finding that fired on 100% of the real corpus.

        The analytical-method checks are deliberately kept — how the chemistry
        and hardness work was performed and recorded does belong in the report.
        Inverted rather than deleted so the decision is explicit.
        """
        parsed = {
            "document_text": "Material Composition Nominal Actual Pre-Solution 42 HRC",
            "nominal": {"Ni": 60, "Cr": 16},
            "actual": {"Ni": 59, "Cr": 16},
            "hardness": {"pre": {"value": 42}, "post": {"value": 35}},
        }

        findings = _review_acceptance_and_methods(parsed)

        self.assertEqual(
            [f for f in findings if f[1] == "Acceptance criteria"], [])
        self.assertTrue(any(category == "Chemistry method"
                            for _sev, category, _msg in findings))
        self.assertTrue(any(category == "Hardness evidence"
                            for _sev, category, _msg in findings))

    def test_a_report_stating_its_specification_is_still_not_penalised(self):
        """Guard: the surviving method checks must recognise real evidence."""
        parsed = {
            "document_text": (
                "Chemical analysis by ICP-OES, calibrated against traceable "
                "standards, measurement uncertainty +/- 0.05 wt%."
            ),
            "nominal": {"Ni": 60}, "actual": {"Ni": 59},
        }

        findings = _review_acceptance_and_methods(parsed)

        self.assertEqual(
            [f for f in findings if f[1] == "Chemistry method"], [])

    def test_sampling_location_is_not_a_sampling_plan(self):
        parsed = {
            "document_text": "Sampling Location: trailing edge",
            "header": {"qty": "92"},
            "sample": {"sample_no": "MS 1C\nMS 2C"},
        }

        findings = _review_sampling_basis(parsed)

        self.assertTrue(any("No sampling plan" in message
                            for message in messages(findings, "warning")))

    def test_optical_gamma_prime_claim_requires_stronger_evidence(self):
        parsed = {
            "comment": "Gamma prime dissolution and re-precipitation are observed.",
            "document_text": "Picture 1 Ortho Acid Etched 500x",
        }

        findings = _review_phase_evidence(parsed)

        self.assertEqual(len(findings), 2)
        self.assertTrue(all(severity == "warning" for severity, _cat, _msg in findings))

    def test_static_page_ten_footer_is_flagged(self):
        parsed = {
            "document_text": "Metallurgical report",
            "page_control": [{"sheet": "MET", "kind": "odd",
                              "text": "WBI.9013 | Page 10"}],
        }

        findings = _review_document_control(parsed)

        self.assertTrue(any(category == "Pagination" and "hard-coded" in message
                            for _sev, category, message in findings))

    def test_thickness_must_reproduce_a_measurement_or_mean(self):
        parsed = {
            "comment": "Average oxidation thickness is 0.022 mm.",
            "pictures": [],
        }
        images = [{"measurements": [18, 43]}]

        findings = _review_thickness(parsed, images, ocr_used=True)

        self.assertTrue(any("not reproduced" in message
                            for message in messages(findings, "warning")))

    def test_supported_thickness_average_is_not_flagged(self):
        parsed = {
            "comment": "Average oxide thickness is 0.041 mm.",
            "pictures": [],
        }
        images = [{"measurements": [41, 37, 44, 40, 42]}]

        findings = _review_thickness(parsed, images, ocr_used=True)

        self.assertEqual(findings, [])

    def test_thickness_color_series_are_reconciled_separately(self):
        parsed = {
            "comment": (
                "As shown in pictures 1 to 2, the chromium deficient layer is "
                "0.065 mm and the oxidation layer is 0.033 mm."
            ),
            "pictures": [
                ("Picture 1:", "Micrograph etched 100x"),
                ("Picture 2:", "Micrograph etched 100x"),
            ],
        }
        images = [
            {
                "measurements": [78, 89, 93, 22, 22, 25],
                "measurement_groups": {
                    "blue": [78, 89, 93], "red": [22, 22, 25],
                },
            },
            {
                "measurements": [39, 44, 48, 22, 24, 44],
                "measurement_groups": {
                    "blue": [39, 44, 48], "red": [22, 24, 44],
                },
            },
        ]

        findings = _review_thickness(parsed, images, ocr_used=True)

        self.assertEqual(len(findings), 1)
        self.assertIn("0.033 mm", findings[0][2])

    def test_up_to_thickness_uses_highest_burned_in_reading(self):
        parsed = {
            "comment": "Oxidation reaches up to 0.062 mm in pics. 2 & 3.",
            "pictures": [],
        }
        images = [{"measurements": [50, 59, 62, 71, 80]}]

        findings = _review_thickness(parsed, images, ocr_used=True)

        self.assertTrue(any("range 50–80" in message
                            for message in messages(findings, "warning")))

    def test_one_ocr_outlier_does_not_break_supported_color_average(self):
        parsed = {
            "comment": "Inter-diffusion is 0.019 mm in pics. 2 & 3.",
            "pictures": [],
        }
        images = [
            {
                "measurements": [17, 18, 20],
                "measurement_groups": {"red": [17, 18, 20]},
            },
            {
                "measurements": [18, 20, 23, 48],
                "measurement_groups": {"red": [18, 20, 23, 48]},
            },
        ]

        findings = _review_thickness(parsed, images, ocr_used=True)

        self.assertEqual(findings, [])

    def test_external_formula_reference_is_detected(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "MET"
        ws["L4"] = "=[1]Cover!G44"

        issues = _formula_issues(wb)

        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0]["cell"], "L4")
        self.assertEqual(issues[0]["kind"], "external-reference")

    def test_structured_table_formula_is_not_an_external_reference(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=SUM(Table1[Thickness])"

        self.assertEqual(_formula_issues(wb), [])

    def test_one_correct_photo_does_not_hide_a_wrong_job_photo(self):
        legends = [
            {"image": "image1.png", "job": "7645", "mag": "100x"},
            {"image": "image2.png", "job": "7646", "mag": "25x"},
        ]

        findings = _review_legends(
            legends, ocr_used=True, caption_mags={"25x", "100x"}, report_job="7645")

        self.assertTrue(any(
            "Mixed micrograph job numbers" in message
            for message in messages(findings, "warning")
        ))

    def test_written_600x_is_not_rejected(self):
        parsed = {
            "pictures": [
                ("Picture 1:", "Middle Wall Micrograph\nOrtho. Acid Etched 600x"),
            ],
            "comment": "",
        }

        findings = _review_captions(parsed)

        self.assertFalse(any(
            category == "Magnification"
            for _severity, category, _message in findings
        ))

    def test_non_micrograph_photo_does_not_require_etch_status(self):
        parsed = {
            "pictures": [
                ("Picture 1:", "Sampling Location"),
                ("Picture 2:", "Al slurry coating thickness average 0.041 mm"),
                ("Picture 3:", "Micrograph Middle Wall Ortho Acid Etched 500x"),
            ],
            "comment": "",
        }

        findings = _review_captions(parsed)

        self.assertFalse(any("No etch status" in message for message in messages(findings)))

    def test_picture_range_overreference_is_detected(self):
        parsed = {
            "pictures": [(f"Picture {number}:", "Micrograph Etched 200x")
                         for number in range(1, 7)],
            "comment": "Secondary carbides are shown in pics 4-8.",
        }

        findings = _review_captions(parsed)

        self.assertTrue(any("Picture 8" in message
                            for message in messages(findings, "warning")))

    def test_duplicate_caption_number_still_allows_fourth_picture_reference(self):
        parsed = {
            "comment": "Gamma prime is shown in pic.no.4.",
            "pictures": [
                ("Picture 1:", "Micrograph etched 25x"),
                ("Picture 2:", "Micrograph etched 100x"),
                ("Picture 3:", "Micrograph etched 100x"),
                ("Picture 3:", "Micrograph etched 500x"),
            ],
        }

        findings = _review_captions(parsed)

        self.assertTrue(any("repeated" in message for message in messages(findings)))
        self.assertFalse(any("refers to Picture 4" in message
                             for message in messages(findings)))

    def test_magnification_is_not_harvested_as_the_job_number(self):
        """Real defect, report 6831: the legend '6831_E_1000x-7' offers two
        four-digit runs — the job AND the 1000 inside '1000x'. The
        magnification is the easier read, so it won the vote and the
        micrograph was reported as belonging to job 1000."""
        reads = ['6831_E_1000x-7', '6831_E_1000x-7', '6831_E_1000x-7']

        job, votes = _vote_job(reads)

        self.assertEqual(job, '6831')
        self.assertNotIn('1000', votes)

    def test_a_job_number_some_passes_disagree_on_is_not_reported(self):
        """Real defect, report 7227 image 2: OCR read '7207' twice and '7297'
        once for a legend that actually says 7227. A wrong job accuses the
        micrograph of belonging to another report, so a contested read must
        come back unreadable rather than confidently wrong."""
        job, votes = _vote_job(['7207_E_200x-1', '7207_E_200x-1', '7297_E_200x-1'])

        self.assertIsNone(job)
        self.assertEqual(votes, {'7207': 2, '7297': 1})

    def test_a_unanimously_read_job_number_is_reported(self):
        """Guard for the above: consensus must still be usable, or the whole
        legend cross-check goes silent."""
        self.assertEqual(_vote_job(['5739_E_500x-2', '5739_E_500x-2'])[0], '5739')
        self.assertEqual(_vote_job(['7227_E_200x-1'] * 4)[0], '7227')

    def test_one_ocr_number_cannot_be_promoted_to_a_fact(self):
        selected, votes = _select_magnification([
            "7646_E_500x-1",
            "7646_E_600x-1",
            "unreadable",
        ])

        self.assertIsNone(selected)
        self.assertEqual(votes, {500: 1, 600: 1})

    def test_consensus_accepts_600_without_an_allow_list(self):
        selected, votes = _select_magnification([
            "7646_E_600x-1",
            "7646 E 600x",
            "7646_E_500x-1",
        ])

        self.assertEqual(selected, 600)
        self.assertEqual(votes, {600: 2, 500: 1})

    def test_caption_wins_when_stable_ocr_disagrees(self):
        image = {
            "image": "image1.png",
            "mag": "600x",
            "ocr_mag": "600x",
            "mag_source": "ocr",
            "mag_votes": {"600x": 2, "500x": 1},
            "mag_read_count": 3,
            "id": "7646_E_600x-1",
        }
        pair = [(
            "Picture 1:",
            "Middle Wall Micrograph\nOrtho. Acid Etched 500x",
            image,
        )]

        with patch("lab_review._picture_image_pairs", return_value=pair):
            verdicts = picture_magnification_verdicts(
                [image],
                [("Picture 1:", "Ortho. Acid Etched 500x")],
                b"workbook",
            )

        self.assertEqual(image["mag"], "500x")
        self.assertEqual(image["ocr_mag"], "600x")
        self.assertEqual(image["mag_source"], "caption")
        self.assertNotIn("id", image)
        self.assertEqual(verdicts[0]["severity"], "warning")
        self.assertIn("caption states 500x", verdicts[0]["note"])
        self.assertIn("OCR consistently read 600x", verdicts[0]["note"])

    def test_matching_written_600x_is_preserved(self):
        image = {
            "image": "image1.png",
            "mag": "600x",
            "ocr_mag": "600x",
            "mag_source": "ocr",
            "mag_votes": {"600x": 3},
            "mag_read_count": 3,
        }
        pair = [("Picture 1:", "Ortho. Acid Etched 600x", image)]

        with patch("lab_review._picture_image_pairs", return_value=pair):
            verdicts = picture_magnification_verdicts(
                [image],
                [("Picture 1:", "Ortho. Acid Etched 600x")],
                b"workbook",
            )

        self.assertEqual(verdicts, [])
        self.assertEqual(image["mag"], "600x")
        self.assertEqual(image["mag_source"], "caption")

    def test_title_identity_accepts_machine_aliases(self):
        cases = [
            (
                "7420 AEN Saudi FS.7 3rd Stage Bucket Metallurgical Report.xlsx",
                "7420", "MS7001", "3rd Stage Bucket",
            ),
            (
                "7504 AEN Saudi 7FA 3rd Stage Bucket Metallurgical Report.xlsx",
                "7504", "MS7001FA", "3rd Stage Bucket",
            ),
            (
                "6579 AEN Saudi FS.7FA 3rd Stage Bucket Metallurgical Report.xlsx",
                "6579", "MS7001FA", "3rd Stage Bucket",
            ),
            (
                "7068 AEN Saudi 7FA 1st Stage Shroud_Tiles Metallurgical Report.xlsx",
                "7068", "MS7001FA", "1st Stage Shroud Tiles",
            ),
            (
                "7646 AEN Saudi V84.2 2nd Stage Vane Metallurgical Report.xlsx",
                "7646", "V84.2", "2nd Stage Vane",
            ),
        ]
        for filename, job, machine, description in cases:
            with self.subTest(filename=filename):
                parsed = {
                    "header": {
                        "job": job,
                        "machine": machine,
                        "customer": "AEN Saudi",
                    },
                    "sample": {"description": description},
                }

                findings = review_filename(filename, parsed, "metallurgical")

                self.assertFalse(any(
                    severity in ("critical", "warning")
                    for severity, category, _message in findings
                    if category == "Title identity"
                ))
                self.assertTrue(any(
                    "machine/set" in message
                    for message in messages(findings, "pass")
                ))

    def test_title_identity_rejects_stage_component_and_machine_mismatches(self):
        parsed = {
            "header": {
                "job": "7646",
                "machine": "MS7001FA",
                "customer": "AEN Saudi",
            },
            "sample": {"description": "3rd Stage Bucket"},
        }

        findings = review_filename(
            "7646 AEN Saudi V84.2 2nd Stage Vane Metallurgical Report.xlsx",
            parsed,
            "metallurgical",
        )
        critical = messages(findings, "critical")

        self.assertTrue(any("title says Stage 2" in message for message in critical))
        self.assertTrue(any(
            'title component "vane"' in message
            for message in critical
        ))
        self.assertTrue(any("title machine/set" in message for message in critical))

    def test_identity_and_magnification_warnings_are_marked_on_report(self):
        title_note = (
            'Report title says Stage 1, but the internal component description '
            'says Stage 2 ("2nd Stage Vane").'
        )
        parsed = {
            "title_findings": [("critical", "Title identity", title_note)],
            "header": {"job": "7646", "machine": "V84.2", "customer": "AEN"},
            "sample": {
                "description": "2nd Stage Vane",
                "material": "Rene80",
                "result": "Acceptable",
            },
            "pictures": [
                ("Picture 1:", "Middle Wall Micrograph\nOrtho. Acid Etched 500x"),
            ],
            "photo_magnification": [{
                "index": 0,
                "label": "Picture 1:",
                "severity": "warning",
                "note": (
                    "Picture 1: written caption states 500x, while legend OCR "
                    "consistently read 600x in 2 of 3 preprocessing passes. "
                    "Treat the caption as the recorded value and verify the "
                    "burned-in legend; OCR may be wrong."
                ),
            }],
            "comment": "A sufficiently long metallurgical discussion for this report.",
            "micrograph_count": 1,
            "signoff": {"met_lab": "Lab", "mat_eng": "Engineer", "date": "2026-07-30"},
            "loc": {
                "sheet": "MET",
                "header": {
                    "job": {"value": (4, 4)},
                    "machine": {"value": (3, 12)},
                    "customer": {"value": (3, 4)},
                },
                "sample": {
                    "description": {"value": (9, 4)},
                    "material": {"value": (9, 10)},
                    "result": {"value": (9, 12)},
                },
                "pictures": [{"value": (20, 4)}],
                "signoff": {},
            },
        }

        highlights = collect_highlights(parsed)

        self.assertTrue(any(
            item["category"] == "Title identity"
            and item["cell"] == (9, 4)
            for item in highlights
        ))
        self.assertTrue(any(
            item["category"] == "Magnification"
            and item["cell"] == (20, 4)
            and item["severity"] == "warning"
            and "600x" in item["note"]
            for item in highlights
        ))

    # ── Recovered from claude/tool-error-detection-7vkrc9 (31 Jul 2026), a
    # branch whose work was never merged. Each of these pins a defect found on
    # a real report; the two marked "already correct in main" were re-fixed
    # independently by the extraction rebuild and are kept as guards. ────────
    def test_nominal_duplicate_header_is_flagged(self):
        # _composition() already computed duplicate_headers/entries for the
        # NOMINAL table (mirroring the Actual side), but _review_composition
        # only ever read composition_meta['actual'] — a mislabeled/duplicated
        # column in the spec itself was silently invisible, while the identical
        # fault on the Actual side is a critical.
        nominal = {"Al": 3.0, "Cr": 14.0}
        actual = {"Al": 2.9, "Cr": 13.5}
        meta = {
            "nominal": {
                "duplicate_headers": ["Cr"],
                "entries": [
                    {"element": "Al", "raw": "3.0", "value": 3.0, "row": 13, "col": 6},
                    {"element": "Cr", "raw": "14.0", "value": 14.0, "row": 13, "col": 7},
                    {"element": "Cr", "raw": "14.0", "value": 14.0, "row": 13, "col": 8},
                ],
            },
            "actual": {"duplicate_headers": [], "entries": []},
        }

        findings = _review_composition(nominal, actual, meta)

        self.assertTrue(any(
            "Nominal composition table repeats element header" in message
            for message in messages(findings, "critical")
        ))

    def test_nominal_duplicate_header_is_marked_on_its_cell(self):
        parsed = {
            "composition_meta": {
                "nominal": {
                    "duplicate_headers": ["Cr"],
                    "entries": [
                        {"element": "Cr", "raw": "14.0", "value": 14.0,
                         "row": 13, "col": 8},
                    ],
                },
                "actual": {"duplicate_headers": [], "entries": []},
            },
        }

        highlights = collect_highlights(parsed)

        self.assertTrue(any(
            highlight["cell"] == (13, 8) and highlight["severity"] == "critical"
            and "Nominal composition table repeats" in highlight["note"]
            for highlight in highlights
        ))

    def test_hyphenated_unetched_caption_is_recognised(self):
        # _UNETCHED_PAT required the contiguous word "unetched", so the
        # hyphenated spelling used in real AEG captions ("Un-etched 25x") was
        # never surfaced as an explicit unetched / as-polished caption.
        parsed = {
            "pictures": [
                ("Picture 1:", "Micrograph- Shroud Tip Overview\nUn-etched 25x"),
            ],
            "comment": "",
        }

        findings = _review_captions(parsed)

        self.assertTrue(any(
            "states unetched / as-polished" in message
            for message in messages(findings, "info")
        ))

    def test_frame_5_machine_alias_is_recognised(self):
        # _canon_machine restricted the GE frame digit to [679]; a real report
        # in a 27-report batch used "FS.5" (GE Frame 5 exists — MS5001), which
        # failed to resolve at all, silently dropping the machine/set
        # cross-check for that report instead of matching or flagging it.
        self.assertEqual(_canon_machine("AEN Saudi FS.5 1st Stage Bucket"), "MS5001")
        self.assertEqual(_canon_machine("MS 5001"), "MS5001")

    def test_fs_dot_frame_variant_suffix_is_captured(self):
        # A real report's title was "FS.7FA" — a fully correct, explicit match
        # for an internal Machine Type of "MS7001FA". The FS.<frame> branch
        # never captured a trailing suffix, so the title was truncated to
        # "MS7001" and then reported as disagreeing with its own content.
        self.assertEqual(_canon_machine("FS.7FA"), "7FA")
        self.assertEqual(_canon_machine("MS7001FA"), "7FA")
        self.assertEqual(_canon_machine("FS.7"), "MS7001")

    def test_fs_dot_frame_variant_matches_internal_machine_type(self):
        parsed = {
            "header": {"job": "6579", "machine": "MS7001FA", "customer": "AEN Saudi"},
            "sample": {"description": "3rd Stage Bucket"},
        }

        findings = review_filename(
            "6579 AEN Saudi FS.7FA 3rd Stage Bucket Metallurgical Report.xlsx",
            parsed,
            "metallurgical",
        )

        self.assertFalse(any(
            severity in ("critical", "warning") and category == "Title identity"
            for severity, category, _message in findings
        ))
        self.assertTrue(any(
            "machine/set" in message for message in messages(findings, "pass")
        ))

    def test_duplicate_actual_composition_across_different_jobs_is_flagged(self):
        # Found on a real 27-report batch: two different AEG job numbers
        # (different customer refs, serials, engineers, dates) carried a
        # byte-identical Actual composition on every matched element — almost
        # certainly a copy-paste, since independent EDS/ICP results are not
        # expected to agree exactly. A single-report review can never see this.
        actual = {
            "Ni": 59.53, "Cr": 13.57, "Co": 9.14, "Mo": 1.46,
            "W": 4.73, "Al": 2.49, "Ti": 4.81, "Ta": 3.05, "Fe": 0.08,
        }
        reports = [
            ("6630.xlsx", {"header": {"job": "6630"}, "actual": dict(actual)}),
            ("6991.xlsx", {"header": {"job": "6991"},
                           "actual": dict(actual, Cu=0.07)}),
            ("unrelated.xlsx", {"header": {"job": "7000"},
                                "actual": {"Ni": 61.0, "Cr": 12.0, "Co": 9.0,
                                           "Mo": 1.5, "W": 4.0}}),
        ]

        findings = find_duplicate_compositions(reports)

        self.assertEqual(len(findings), 1)
        severity, category, message = findings[0]
        self.assertEqual(severity, "critical")
        self.assertEqual(category, "Composition")
        self.assertIn("6630.xlsx", message)
        self.assertIn("6991.xlsx", message)
        self.assertIn("9 matched element", message)

    def test_same_job_number_is_not_a_duplicate_composition_signal(self):
        # Two uploads of one job are a revision or a re-upload, which
        # add_version_findings covers — not a copy-paste signal.
        actual = {"Ni": 60.0, "Cr": 14.0, "Co": 9.5, "Mo": 1.5, "W": 3.8}
        reports = [
            ("a.xlsx", {"header": {"job": "7000"}, "actual": dict(actual)}),
            ("b.xlsx", {"header": {"job": "7000"}, "actual": dict(actual)}),
        ]

        self.assertEqual(find_duplicate_compositions(reports), [])

    def test_plural_pics_range_reference_is_parsed(self):
        # Already correct in main (_picture_references), re-fixed independently
        # by the rebuild. Kept as a guard: the original regex matched only the
        # singular "Pic"/"Picture", so a real comment writing "Ref. pics. 9–10"
        # could reference a picture that does not exist and slip through.
        parsed = {
            "pictures": [(f"Picture {n}:", "Etched 500x") for n in range(1, 5)],
            "comment": "Solution HT was performed (Ref. pics. 9–10).",
        }

        findings = _review_captions(parsed)

        self.assertTrue(any(
            "Comment refers to Picture 10" in message
            and "highest present caption is Picture 4" in message
            for message in messages(findings, "warning")
        ))

    def test_underscore_separated_filename_matches_stage_and_component(self):
        # Already correct in main (_component_identity normalises '_' and '-').
        # Kept as a guard: regex '_' is a \w character, so \b-anchored
        # component/stage patterns never matched across it and every
        # underscore-joined filename looked like it disagreed with its content.
        parsed = {
            "header": {"job": "7398", "machine": "MS7001", "customer": "AEN SAUDI"},
            "sample": {"description": "1st Stage Bucket"},
        }

        findings = review_filename(
            "7398__AEN_Saudi_FS.7_1st_Stage_Bucket_Metallurgical_Report__AEG_final.xlsx",
            parsed,
            "metallurgical",
        )

        self.assertFalse(any(
            severity in ("critical", "warning") and category == "Title identity"
            for severity, category, _message in findings
        ))

    # ── Second salvage pass, from claude/tool-audit-improvements-v7bj10
    # (6 Jul 2026). These four were NOT re-fixed by the extraction rebuild —
    # I had assumed "the file was rewritten" meant "the defect was re-fixed",
    # which is the same assumption that lost the branches in the first place. ──
    def test_a_rejected_result_cell_does_not_contradict_a_rejecting_comment(self):
        # "Not acceptable" contains "accept", so the Result cell graded positive
        # AND negative at once. A report whose cell and comment agreed on a
        # rejection was then warned about for contradicting itself.
        parsed = {
            "sample": {"result": "Not acceptable"},
            "comment": "The material is not suitable for further service and is "
                       "rejected; cracking exceeds the repair limit.",
            "coating": {},
        }

        findings = _review_comment(parsed)

        self.assertFalse(any(
            "but the comment indicates the part is NOT suitable" in message
            for message in messages(findings)
        ))

    def test_a_coating_layout_the_parser_missed_is_not_read_as_no_coating(self):
        # Every coating field None meant "no coating recorded", so a report the
        # parser simply failed to read was accused of contradicting its own
        # comment — the tool blaming the report for its own gap.
        parsed = {
            "sample": {"result": "See comment"},
            "comment": "The set of buckets was received with an MCrAlY coating "
                       "which remains intact over the airfoil.",
            "coating": {},          # parser found nothing at all
        }

        findings = _review_comment(parsed)

        self.assertFalse(any(
            "indicates no coating" in message for message in messages(findings)
        ))

    def test_a_coating_cell_that_really_says_no_still_contradicts_the_comment(self):
        parsed = {
            "sample": {"result": "See comment"},
            "comment": "The set of buckets was received with an MCrAlY coating.",
            "coating": {"present": "No", "type": None,
                        "received": None, "outgoing": None},
        }

        findings = _review_comment(parsed)

        self.assertTrue(any(
            "indicates no coating" in message for message in messages(findings)
        ))

    def test_hardness_in_different_scales_is_not_compared(self):
        # A 355 HV post-solution reading against a 42 HRC pre-solution reading
        # is not a comparison; it read as "hardness went up by 313".
        findings = _review_hardness(
            {"pre": {"value": 42.0, "unit": "HRC"},
             "post": {"value": 355.0, "unit": "HV"}},
            "GTD 111")

        self.assertFalse(any(
            "exceeds pre-solution" in message for message in messages(findings)
        ))

    def test_normal_scatter_on_a_vickers_reading_is_not_flagged(self):
        # The ±2 guard band is right for HRC and far too tight for HV, where a
        # few points of scatter between two indents is routine.
        findings = _review_hardness(
            {"pre": {"value": 350.0, "unit": "HV"},
             "post": {"value": 356.0, "unit": "HV"}},
            "GTD 111")

        self.assertFalse(any(
            "exceeds pre-solution" in message for message in messages(findings)
        ))

    def test_a_real_rise_in_the_same_scale_is_still_flagged(self):
        findings = _review_hardness(
            {"pre": {"value": 41.4, "unit": "HRC"},
             "post": {"value": 47.0, "unit": "HRC"}},
            "GTD 111")

        self.assertTrue(any(
            "exceeds pre-solution" in message for message in messages(findings)
        ))

    def test_sheet_scans_are_bounded(self):
        # One stray cell at Excel's last row inflates ws.max_row to ~1,048,576;
        # every label scan then became a minutes-long, worker-freezing loop.
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "AEG. Job No:"
        sheet["B1"] = "6943"
        sheet.cell(row=1_048_576, column=1, value="stray")

        self.assertGreaterEqual(sheet.max_row, 1_048_576)
        scanned = sum(1 for _row in lab_review._scan_rows(sheet))
        self.assertLessEqual(scanned, lab_review._SCAN_ROWS_CAP)
        # The report's own data is still inside the scanned window.
        self.assertIsNotNone(lab_review._find_label(sheet, r"AEG\.?\s*Job\s*No"))

    def test_a_measurement_point_index_row_is_not_thickness_data(self):
        # Real corpus report 6983 carries a '1 2 3 4 5' point-index row in its
        # coating table. It escaped notice only because it sits above the first
        # MIN/MAX row; in a multi-block table it inherits the forward-filled
        # limits and every index is reported as a thickness out of range.
        self.assertTrue(lab_review._looks_like_point_header([1.0, 2.0, 3.0, 4.0, 5.0]))
        self.assertTrue(lab_review._looks_like_point_header([0, 1, 2, 3]))
        # Real thickness readings must never be mistaken for an index row.
        self.assertFalse(
            lab_review._looks_like_point_header([0.045, 0.04, 0.041, 0.039]))
        self.assertFalse(lab_review._looks_like_point_header([1.0, 2.0]))
        self.assertFalse(lab_review._looks_like_point_header([2.0, 3.0, 4.0]))
        self.assertFalse(lab_review._looks_like_point_header([1.0, 2.0, 4.0, 5.0]))

    def test_a_point_index_row_below_the_limits_is_not_range_checked(self):
        # The failure the check above prevents, end to end: a second measurement
        # block repeats the point header, and by then MIN/MAX are forward-filled.
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Coating Coverage Assessment"
        sheet["A3"] = "Measurements"
        sheet["F3"] = "Average Values"
        sheet["G3"] = "MIN"
        sheet["H3"] = "MAX"
        sheet["A4"] = "Design limit"
        for column, value in zip("ABCDE", [0.045, 0.04, 0.041, 0.039, 0.043]):
            sheet[f"{column}5"] = value
        sheet["G5"], sheet["H5"] = 0.03, 0.05
        # Second block: the point header now sits under a limits row.
        for column, value in zip("ABCDE", [1, 2, 3, 4, 5]):
            sheet[f"{column}7"] = value
        for column, value in zip("ABCDE", [0.044, 0.048, 0.043, 0.045, 0.041]):
            sheet[f"{column}8"] = value

        parsed = lab_review.parse_coating(workbook)
        findings = lab_review.review_coating(parsed)

        self.assertEqual([row["values"] for row in parsed["rows"]],
                         [[0.045, 0.04, 0.041, 0.039, 0.043],
                          [0.044, 0.048, 0.043, 0.045, 0.041]])
        self.assertFalse(any(
            "outside" in message or "out of range" in message
            for message in messages(findings)
        ))


if __name__ == "__main__":
    unittest.main()
