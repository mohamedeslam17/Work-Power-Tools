"""Executable specification for the extraction rebuild.

Every test here is a defect that was reproduced against the shipped parser and
recorded in docs/REBUILD.md as D1–D10. They are the acceptance criteria for
Phase 1: when the extraction layer is rebuilt, these pass.

HOW THIS FILE WORKS
-------------------
Tests for defects that are still open carry ``@unittest.expectedFailure``, so
the suite stays green while the work is in progress.

When you fix a defect, its test starts passing and unittest reports an
UNEXPECTED SUCCESS, which fails the run. That is the signal, not a problem:
**delete the decorator for that test in the same commit as the fix.** A red
suite saying "unexpected success" means you finished something.

Tests WITHOUT the decorator are behaviour that already works today and must
keep working. Do not weaken them.

TARGET MODEL
------------
The shape these tests assume is specified in docs/REBUILD.md § "Document
model". If you adopt a better shape, change these tests in the same commit and
record the deviation in docs/FEEDBACK.md — do not silently reinterpret them.
"""
import io
import unittest

import openpyxl
from openpyxl import Workbook

import lab_review as L


def _wb_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(wb):
    return openpyxl.load_workbook(io.BytesIO(_wb_bytes(wb)), data_only=True)


def _met_sheet():
    """A minimal but well-formed metallurgical report.

    Row 4 is deliberately left empty: it is the slot the decoy-prose tests use,
    so that a decoy sits ABOVE the real Machine Type label (the sheet scan is
    row-major) without any other field's value sitting to its right. Without
    that empty gutter a decoy test can pass for the wrong reason — the field
    absorbs the neighbouring value instead of going blank.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "METALLURGICAL EXAMINATION"
    ws["A2"] = "Customer:"
    ws["C2"] = "ACME Power"
    ws["A3"] = "AEG Job No:"
    ws["C3"] = "J-1001"
    ws["A5"] = "Machine Type:"
    ws["C5"] = "V94.3A"
    ws["A6"] = "Sample nr"
    ws["B6"] = "Description"
    ws["C6"] = "S/N"
    ws["D6"] = "Location"
    ws["E6"] = "Material"
    ws["F6"] = "Result"
    ws["A7"] = 1
    ws["B7"] = "Blade"
    ws["C7"] = "SN-001"
    ws["D7"] = "Root"
    ws["E7"] = "IN738"
    ws["F7"] = "Acceptable"
    return wb, ws


# ── D1 · extraction status must be distinguishable ────────────────────────
class ExtractionStatusTests(unittest.TestCase):
    """A field the parser could not locate must not look like a blank field.

    This is the root defect. Today both return None, so the reviewer emits
    "Machine type is blank or a placeholder" for a value that is plainly
    present in the report but was missed by the label scan.
    """

    @unittest.expectedFailure
    def test_found_empty_and_not_located_are_three_distinct_states(self):
        wb, ws = _met_sheet()                       # machine present
        found = L.parse_metallurgical(_load(wb))

        wb2, ws2 = _met_sheet()
        ws2["C5"] = None                            # label present, value blank
        empty = L.parse_metallurgical(_load(wb2))

        wb3, ws3 = _met_sheet()
        ws3["A5"] = None                            # no machine label at all
        ws3["C5"] = None
        absent = L.parse_metallurgical(_load(wb3))

        self.assertEqual(found["fields"]["machine"]["status"], "found")
        self.assertEqual(empty["fields"]["machine"]["status"], "empty")
        self.assertEqual(absent["fields"]["machine"]["status"], "not_located")

    @unittest.expectedFailure
    def test_located_field_records_its_source_cell(self):
        wb, _ = _met_sheet()
        parsed = L.parse_metallurgical(_load(wb))
        self.assertEqual(parsed["fields"]["machine"]["cell"], "C5")


# ── D2 · prose must not hijack a label ────────────────────────────────────
class LabelResolutionTests(unittest.TestCase):

    @unittest.expectedFailure
    def test_prose_mentioning_a_label_does_not_erase_the_field(self):
        """D2. One narrative line containing the words 'Machine Type' makes the
        whole-sheet scan match the prose instead of the real field."""
        wb, ws = _met_sheet()
        ws["A4"] = "Machine Type per customer drawing pack"   # scanned first
        parsed = L.parse_metallurgical(_load(wb))
        self.assertEqual(parsed["header"].get("machine"), "V94.3A")

    @unittest.expectedFailure
    def test_prose_decoy_does_not_produce_a_blank_field_finding(self):
        """D2, end to end: the fabricated finding must not reach the reviewer.

        Asserts the value positively as well as the absence of the finding. The
        negative assertion alone is not enough — a parser that resolves the
        decoy to some *other* neighbouring value also suppresses the finding
        while being more wrong, not less.
        """
        wb, ws = _met_sheet()
        ws["A4"] = "Machine Type per customer drawing pack"
        _rtype, parsed, findings = L.review_report(
            "J-1001 metallurgical.xlsx", _wb_bytes(wb), ocr=False)
        self.assertEqual(parsed["header"].get("machine"), "V94.3A")
        blank_claims = [m for _s, _c, m in findings if "achine type is blank" in m]
        self.assertEqual(blank_claims, [])

    @unittest.expectedFailure
    def test_decoy_does_not_absorb_a_neighbouring_field_value(self):
        """D9, value variant. A decoy label whose row carries another field's
        value adopts that value: machine silently becomes the job number.

        This is the most dangerous shape of the defect — the field is non-empty,
        so no completeness check fires, and the wrong value flows into the
        title-identity and traceability checks as though it were read correctly.
        """
        wb, ws = _met_sheet()
        ws["A3"] = "Machine Type per customer drawing pack"   # displaces the job label
        parsed = L.parse_metallurgical(_load(wb))
        self.assertNotEqual(parsed["header"].get("machine"), "J-1001")

    @unittest.expectedFailure
    def test_blank_field_does_not_absorb_the_next_label(self):
        """D9. A genuinely blank field followed by another label on the same row
        reads that label as its own value — a confidently wrong string that no
        completeness check can catch, because it is not empty."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "METALLURGICAL EXAMINATION"
        ws["A2"] = "Sample nr"
        ws["A3"] = "Machine Type:"          # value cell deliberately empty
        ws["D3"] = "AEG Job No:"
        ws["E3"] = "J-1001"
        parsed = L.parse_metallurgical(_load(wb))
        self.assertIsNone(parsed["header"].get("machine"))
        self.assertEqual(parsed["header"].get("job"), "J-1001")

    @unittest.expectedFailure
    def test_value_stacked_below_its_label_is_read(self):
        """D10. Header fields only ever scan rightward, so a template variant
        that stacks the value under the label loses every header field."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "METALLURGICAL EXAMINATION"
        ws["A2"] = "Sample nr"
        ws["A3"] = "Machine Type:"
        ws["A4"] = "V94.3A"
        parsed = L.parse_metallurgical(_load(wb))
        self.assertEqual(parsed["header"].get("machine"), "V94.3A")


# ── D3 · every sample must be parsed ──────────────────────────────────────
class MultiSampleTests(unittest.TestCase):
    """⚠ SHAPE WARNING — read before optimising for these fixtures.

    The one-row-per-sample layout below is SYNTHETIC AND UNVERIFIED. It was
    written before the real corpus was recovered, and no real AEG report in
    corpus/ uses it.

    The real shape packs every sample into a SINGLE cell, whitespace and
    newline separated — see tests/test_corpus_regression.py
    ::RealMultiSampleTests, which is the authoritative D3 spec. Build for that
    first. These row-based tests are kept because the layout is plausible in
    other templates and the defect is the same one either way, but if the two
    ever conflict, the corpus wins.
    """

    @unittest.expectedFailure
    def test_every_sample_row_is_parsed(self):
        """D3, synthetic row-based variant. Only the first row under 'Sample nr'
        is read, so a rejection on sample 2 is invisible to all twenty checks."""
        wb, ws = _met_sheet()
        ws["A8"] = 2
        ws["B8"] = "Blade"
        ws["C8"] = "SN-002"
        ws["D8"] = "Tip"
        ws["E8"] = "IN738"
        ws["F8"] = "REJECT"
        ws["A9"] = 3
        ws["B9"] = "Vane"
        ws["C9"] = "SN-003"
        ws["D9"] = "Mid"
        ws["E9"] = "GTD-111"
        ws["F9"] = "Acceptable"

        parsed = L.parse_metallurgical(_load(wb))
        self.assertEqual(len(parsed["samples"]), 3)
        self.assertEqual(
            [s.get("serial") for s in parsed["samples"]],
            ["SN-001", "SN-002", "SN-003"])
        self.assertIn("REJECT", [s.get("result") for s in parsed["samples"]])
        self.assertIn("GTD-111", [s.get("material") for s in parsed["samples"]])

    @unittest.expectedFailure
    def test_a_rejected_later_sample_is_not_silently_passed(self):
        """D3, end to end. A report containing a rejection must not review clean."""
        wb, ws = _met_sheet()
        ws["A8"] = 2
        ws["C8"] = "SN-002"
        ws["E8"] = "IN738"
        ws["F8"] = "REJECT"
        _rtype, _parsed, findings = L.review_report(
            "J-1001 metallurgical.xlsx", _wb_bytes(wb), ocr=False)
        self.assertTrue(
            any("SN-002" in m or "REJECT" in m.upper() for _s, _c, m in findings),
            "no finding referenced the rejected sample")

    def test_single_sample_report_still_parses(self):
        """Guard: the common case must not regress while fixing the above."""
        wb, _ = _met_sheet()
        parsed = L.parse_metallurgical(_load(wb))
        self.assertEqual(parsed["sample"].get("serial"), "SN-001")
        self.assertEqual(parsed["sample"].get("material"), "IN738")


# ── D4 · every composition table must be parsed ───────────────────────────
class CompositionTableTests(unittest.TestCase):

    @unittest.expectedFailure
    def test_every_actual_table_is_extracted(self):
        """D4. Only the first '(Actual)' table is read. A second sample with
        grossly off-specification chemistry never reaches the checks."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "METALLURGICAL EXAMINATION"
        ws["A2"] = "Sample nr"
        for row, label, vals in (
                (5, "Composition (Nominal)", (61.0, 16.0, 3.4)),
                (7, "Composition (Actual)", (60.8, 15.9, 3.3)),
                (20, "Composition (Actual)", (41.0, 28.0, 0.2))):
            ws.cell(row=row, column=1, value=label)
            for offset, (el, val) in enumerate(zip(("Ni", "Cr", "Ti"), vals)):
                ws.cell(row=row, column=2 + offset, value=el)
                ws.cell(row=row + 1, column=2 + offset, value=val)

        parsed = L.parse_metallurgical(_load(wb))
        tables = parsed["composition"]["actual_tables"]
        self.assertEqual(len(tables), 2)
        self.assertAlmostEqual(tables[1]["values"]["Ni"], 41.0)
        self.assertAlmostEqual(tables[1]["values"]["Cr"], 28.0)

    def test_single_composition_table_still_parses(self):
        """Guard for the common case."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "METALLURGICAL EXAMINATION"
        ws["A2"] = "Sample nr"
        ws["A5"] = "Composition (Actual)"
        for offset, (el, val) in enumerate(zip(("Ni", "Cr", "Ti"), (60.8, 15.9, 3.3))):
            ws.cell(row=5, column=2 + offset, value=el)
            ws.cell(row=6, column=2 + offset, value=val)
        parsed = L.parse_metallurgical(_load(wb))
        self.assertAlmostEqual(parsed["actual"]["Ni"], 60.8)


# ── D5 · numeric coercion must not invent measurements ────────────────────
class NumericCoercionTests(unittest.TestCase):

    @unittest.expectedFailure
    def test_prose_does_not_yield_a_number(self):
        """D5. 'N/A (rev 2)' becomes 2.0 and enters the review as a hardness
        measurement with the same standing as a real reading."""
        for text in ("see note 3", "N/A (rev 2)", "rev 4 pending", "note 7"):
            with self.subTest(text=text):
                self.assertIsNone(L._num(text))

    @unittest.expectedFailure
    def test_below_detection_limit_is_not_a_plain_zero_point_zero_one(self):
        """D5. '<0.01' currently returns 0.01, losing the below-LOD meaning that
        the composition checks depend on."""
        self.assertIsNone(L._num("<0.01"))

    def test_european_and_english_decimals_still_parse(self):
        """Guard: this already works and must keep working."""
        self.assertAlmostEqual(L._num("12,5"), 12.5)
        self.assertAlmostEqual(L._num("1.234,56"), 1234.56)
        self.assertAlmostEqual(L._num("1,234.56"), 1234.56)
        self.assertAlmostEqual(L._num("42"), 42.0)
        self.assertIsNone(L._num("Not tested"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
